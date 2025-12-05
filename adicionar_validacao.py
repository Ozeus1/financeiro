import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

def adicionar_listas_suspensas_excel():
    """
    Script para adicionar listas de validação de dados a um arquivo Excel.
    """
    # --- Listas de opções fornecidas ---
    meios_pagamento_lista = [
        "Boleto", "Cartão Azul", "Cartão BB", "Cartão C6", "Cartão Gol","Cartão Latam", "Cartão Mercado Pago",
        "Cartão Nubank", "Cartão pão de açucar", "Cartão Unlimited",
        "Cartão Unlimited Master", "Cartão Unlimited Visa", "Pix",
        "Transferência", "Dinheiro"
    ]

    categorias_lista = [
        "Academia", "Alimentação", "Cagece","Condomínio", "Celulares", "Combustível", "Educação",
        "Energia","Estacionamento","Feira", "Funcionários", "Gás", "IA e propaganda Instagram","Internet", "Lanche Aline", "Lazer","Manutenção carro", "Mercado", "Moradia","Netflix e Afins",
        "Plano de Saúde", "Pontos e Viagens", "RBV negócio", "Remédios", "Restaurantes","salão",
        "Saúde", "Transporte", "Vestuário"
    ]

    root = tk.Tk()
    root.withdraw() # Oculta a janela principal do tkinter

    # Passo 1: Solicitar ao usuário o arquivo Excel de entrada
    caminho_arquivo = filedialog.askopenfilename(
        title="Selecione a planilha Excel para adicionar a validação",
        filetypes=[("Arquivos Excel", "*.xlsx")]
    )

    if not caminho_arquivo:
        messagebox.showinfo("Operação Cancelada", "Nenhum arquivo foi selecionado.")
        return

    try:
        # Passo 2: Carregar a planilha
        workbook = openpyxl.load_workbook(caminho_arquivo)
        planilha_principal = workbook.active

        # Passo 3: Criar uma nova planilha para armazenar as listas de validação
        nome_planilha_validacao = '_ListasDeValidacao'
        if nome_planilha_validacao in workbook.sheetnames:
            del workbook[nome_planilha_validacao] # Remove se já existir para evitar erros
        
        planilha_validacao = workbook.create_sheet(title=nome_planilha_validacao)
        planilha_validacao.sheet_state = 'hidden' # Oculta a planilha para não poluir o arquivo

        # Preenche a planilha de validação
        for i, item in enumerate(sorted(meios_pagamento_lista), 1):
            planilha_validacao[f'A{i}'] = item
        for i, item in enumerate(sorted(categorias_lista), 1):
            planilha_validacao[f'B{i}'] = item

        # Passo 4: Localizar as colunas "Meio de Pagamento" e "Categoria"
        col_meio_pagamento = None
        col_categoria = None
        for celula in planilha_principal[1]: # Itera sobre a primeira linha (cabeçalho)
            if celula.value == 'Meio de Pagamento':
                col_meio_pagamento = celula.column_letter
            elif celula.value == 'Categoria':
                col_categoria = celula.column_letter
        
        if not all([col_meio_pagamento, col_categoria]):
            messagebox.showerror("Erro", "Não foi possível encontrar as colunas 'Meio de Pagamento' e/ou 'Categoria' na planilha.")
            return

        # Passo 5: Criar e aplicar a validação de dados
        # Fórmula que aponta para a lista na planilha oculta
        formula_meios = f"'{nome_planilha_validacao}'!$A$1:$A${len(meios_pagamento_lista)}"
        formula_categorias = f"'{nome_planilha_validacao}'!$B$1:$B${len(categorias_lista)}"
        
        # Cria o objeto de validação
        dv_meios = DataValidation(type="list", formula1=formula_meios, allow_blank=True)
        dv_categorias = DataValidation(type="list", formula1=formula_categorias, allow_blank=True)

        # Adiciona a validação à planilha principal
        planilha_principal.add_data_validation(dv_meios)
        planilha_principal.add_data_validation(dv_categorias)

        # Aplica a regra a todas as células das colunas, a partir da linha 2
        dv_meios.add(f'{col_meio_pagamento}2:{col_meio_pagamento}1048576')
        dv_categorias.add(f'{col_categoria}2:{col_categoria}1048576')

        # Passo 6: Salvar o novo arquivo
        novo_caminho = filedialog.asksaveasfilename(
            title="Salvar planilha com validação como...",
            defaultextension=".xlsx",
            filetypes=[("Arquivos Excel", "*.xlsx")],
            initialfile=caminho_arquivo.replace(".xlsx", "_com_validacao.xlsx")
        )

        if not novo_caminho:
            messagebox.showinfo("Operação Cancelada", "O arquivo não foi salvo.")
            return

        workbook.save(novo_caminho)
        messagebox.showinfo("Sucesso", f"Validação adicionada com sucesso!\nO novo arquivo foi salvo em: {novo_caminho}")

    except Exception as e:
        messagebox.showerror("Erro Inesperado", f"Ocorreu um erro ao processar o arquivo: {e}")

# Executa a função principal
if __name__ == "__main__":
    adicionar_listas_suspensas_excel()