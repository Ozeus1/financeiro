import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

class RelatorioOrcadoVsGasto:
    def __init__(self, parent):
        self.root = tk.Toplevel(parent)
        self.root.title("Relatório Mensal: Orçado vs. Gasto")
        self.root.geometry("900x650")  # Aumentei a altura para o totalizador
        self.root.resizable(True, True)
        self.root.transient(parent)
        self.root.grab_set()

        self.db_conn = sqlite3.connect('financas.db')
        self.report_data_df = pd.DataFrame()      # DataFrame com todos os dados
        self.displayed_data_df = pd.DataFrame()   # DataFrame com dados filtrados (para exportação)

        self.criar_widgets()
        self.popular_filtro_mes_e_carregar_dados()

    def criar_widgets(self):
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Frame para filtros e botões
        controles_frame = ttk.Frame(main_frame)
        controles_frame.pack(fill=tk.X, pady=(0, 10))

        # Filtro de Mês
        ttk.Label(controles_frame, text="Filtrar por Mês:").pack(side=tk.LEFT, padx=(0, 5))
        self.combo_mes = ttk.Combobox(controles_frame, state="readonly", width=20)
        self.combo_mes.pack(side=tk.LEFT, padx=(0, 10))
        self.combo_mes.bind("<<ComboboxSelected>>", self.on_month_change)

        # Botões
        self.btn_exportar = ttk.Button(controles_frame, text="Exportar para Excel", command=self.exportar_para_excel)
        self.btn_exportar.pack(side=tk.LEFT, padx=(0, 5))

        self.btn_fechar = ttk.Button(controles_frame, text="Fechar", command=self.root.destroy)
        self.btn_fechar.pack(side=tk.LEFT)
        
        # Frame da Tabela
        tree_frame = ttk.LabelFrame(main_frame, text="Detalhes do Relatório")
        tree_frame.pack(fill=tk.BOTH, expand=True)

        # Treeview
        self.tree = ttk.Treeview(
            tree_frame,
            columns=('mes_ano', 'categoria', 'orcado', 'gasto', 'saldo'),
            show='headings'
        )
        self.tree.heading('mes_ano', text='Mês/Ano')
        self.tree.heading('categoria', text='Categoria')
        self.tree.heading('orcado', text='Orçado (R$)')
        self.tree.heading('gasto', text='Gasto (R$)')
        self.tree.heading('saldo', text='Saldo (R$)')

        self.tree.column('mes_ano', anchor=tk.W, width=100)
        self.tree.column('categoria', anchor=tk.W, width=200)
        self.tree.column('orcado', anchor=tk.E, width=120)
        self.tree.column('gasto', anchor=tk.E, width=120)
        self.tree.column('saldo', anchor=tk.E, width=120)

        scrollbar_y = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        scrollbar_x = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        self.tree.pack(fill=tk.BOTH, expand=True)

        # --- NOVO: Frame Totalizador ---
        total_frame = ttk.LabelFrame(main_frame, text="Resumo do Período Selecionado")
        total_frame.pack(fill=tk.X, pady=(10, 0))

        font_bold = ("Segoe UI", 10, "bold")

        ttk.Label(total_frame, text="Total Orçado:", font=font_bold).grid(row=0, column=0, padx=5, pady=5, sticky="e")
        self.lbl_total_orcado = ttk.Label(total_frame, text="R$ 0,00", font=font_bold, width=15, anchor="e")
        self.lbl_total_orcado.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        ttk.Label(total_frame, text="Total Gasto:", font=font_bold).grid(row=0, column=2, padx=5, pady=5, sticky="e")
        self.lbl_total_gasto = ttk.Label(total_frame, text="R$ 0,00", font=font_bold, width=15, anchor="e")
        self.lbl_total_gasto.grid(row=0, column=3, padx=5, pady=5, sticky="w")

        ttk.Label(total_frame, text="Diferença:", font=font_bold).grid(row=0, column=4, padx=5, pady=5, sticky="e")
        self.lbl_diferenca = ttk.Label(total_frame, text="R$ 0,00", font=font_bold, width=15, anchor="e")
        self.lbl_diferenca.grid(row=0, column=5, padx=5, pady=5, sticky="w")

    def popular_filtro_mes_e_carregar_dados(self):
        try:
            # Busca meses distintos que têm despesas para popular o filtro
            df_meses = pd.read_sql_query("""
                SELECT DISTINCT strftime('%Y-%m', data_pagamento) as AnoMes
                FROM v_despesas_compat 
                ORDER BY AnoMes DESC
            """, self.db_conn)
            
            meses = ["Todos os Meses"] + df_meses['AnoMes'].tolist()
            self.combo_mes['values'] = meses
            self.combo_mes.set("Todos os Meses")

            self.gerar_dados_completos()   # Gera o DataFrame completo
            self.atualizar_visualizacao()  # Exibe os dados iniciais (Todos os Meses)

        except Exception as e:
            messagebox.showerror("Erro Inicial", f"Erro ao carregar meses disponíveis: {e}", parent=self.root)

    def gerar_dados_completos(self):
        """Busca todos os dados do banco e os armazena em self.report_data_df."""
        try:
            # Gasto por mês / categoria
            df_expenses = pd.read_sql_query("""
                SELECT strftime('%Y-%m', data_pagamento) as AnoMes,
                       conta_despesa,
                       SUM(valor) as Gasto 
                FROM v_despesas_compat 
                GROUP BY AnoMes, conta_despesa
            """, self.db_conn)

            # Orçado por categoria
            df_budget = pd.read_sql_query(
                "SELECT conta_despesa, valor_orcado as Orcado FROM orcamento",
                self.db_conn
            )

            if df_expenses.empty and df_budget.empty:
                self.report_data_df = pd.DataFrame()
                return

            all_months = sorted(df_expenses['AnoMes'].unique(), reverse=True) if not df_expenses.empty else []
            all_categories = pd.unique(
                list(df_expenses['conta_despesa'].unique()) +
                list(df_budget['conta_despesa'].unique())
            )

            if not all_months and not df_budget.empty:
                # Caso de orçamento sem despesas
                base_data = [{'AnoMes': 'N/D', 'conta_despesa': cat} for cat in all_categories]
                df_report_base = pd.DataFrame(base_data)
            elif all_months:
                df_report_base = pd.MultiIndex.from_product(
                    [all_months, all_categories],
                    names=['AnoMes', 'conta_despesa']
                ).to_frame(index=False)
            else:
                self.report_data_df = pd.DataFrame()
                return
            
            df_report = pd.merge(
                df_report_base,
                df_expenses,
                on=['AnoMes', 'conta_despesa'],
                how='left'
            )

            if not df_budget.empty:
                df_report = pd.merge(
                    df_report,
                    df_budget,
                    on='conta_despesa',
                    how='left'
                )
            else:
                df_report['Orcado'] = 0.0

            df_report['Gasto'] = df_report['Gasto'].fillna(0.0)
            df_report['Orcado'] = df_report['Orcado'].fillna(0.0)
            df_report['Saldo'] = df_report['Orcado'] - df_report['Gasto']
            
            df_report_final = df_report[
                (df_report['Orcado'] != 0) | (df_report['Gasto'] != 0)
            ]
            self.report_data_df = df_report_final.sort_values(
                by=['AnoMes', 'conta_despesa'],
                ascending=[False, True]
            )

        except sqlite3.Error as e:
            messagebox.showerror("Erro de Banco de Dados", f"Erro ao consultar dados: {e}", parent=self.root)
            self.report_data_df = pd.DataFrame()
        except Exception as e:
            messagebox.showerror("Erro Inesperado", f"Ocorreu um erro ao gerar o relatório: {e}", parent=self.root)
            self.report_data_df = pd.DataFrame()
            import traceback
            traceback.print_exc()

    def on_month_change(self, event=None):
        self.atualizar_visualizacao()

    def atualizar_visualizacao(self):
        """Filtra os dados e atualiza o Treeview e os totalizadores."""
        for item in self.tree.get_children():
            self.tree.delete(item)

        if self.report_data_df.empty:
            messagebox.showinfo(
                "Relatório",
                "Não há dados de despesas ou orçamentos para gerar o relatório.",
                parent=self.root
            )
            self.atualizar_totalizadores(0, 0, 0)
            return

        selected_month = self.combo_mes.get()
        
        if selected_month == "Todos os Meses":
            df_to_display = self.report_data_df
        else:
            df_to_display = self.report_data_df[self.report_data_df['AnoMes'] == selected_month]
        
        self.displayed_data_df = df_to_display  # Salva para exportação

        for _, row in df_to_display.iterrows():
            saldo_val = row['Saldo']
            cor_saldo = 'red' if saldo_val < 0 else 'darkgreen' if saldo_val > 0 else 'black'
            
            self.tree.insert(
                '',
                tk.END,
                values=(
                    row['AnoMes'],
                    row['conta_despesa'],
                    f"R$ {row['Orcado']:.2f}".replace('.', ','),
                    f"R$ {row['Gasto']:.2f}".replace('.', ','),
                    f"R$ {saldo_val:.2f}".replace('.', ',')
                ),
                tags=(cor_saldo,)
            )
        
        self.tree.tag_configure('red', foreground='red')
        self.tree.tag_configure('darkgreen', foreground='darkgreen')
        self.tree.tag_configure('black', foreground='black')

        # Calcular e atualizar totais
        total_orcado = df_to_display['Orcado'].sum()
        total_gasto = df_to_display['Gasto'].sum()
        total_saldo = df_to_display['Saldo'].sum()
        self.atualizar_totalizadores(total_orcado, total_gasto, total_saldo)

    def atualizar_totalizadores(self, orcado, gasto, saldo):
        self.lbl_total_orcado.config(
            text=f"R$ {orcado:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        )
        self.lbl_total_gasto.config(
            text=f"R$ {gasto:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        )
        self.lbl_diferenca.config(
            text=f"R$ {saldo:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        )

        if saldo < 0:
            self.lbl_diferenca.config(foreground="red")
        elif saldo > 0:
            self.lbl_diferenca.config(foreground="darkgreen")
        else:
            self.lbl_diferenca.config(foreground="black")

    def exportar_para_excel(self):
        if not hasattr(self, 'displayed_data_df') or self.displayed_data_df.empty:
            messagebox.showwarning(
                "Sem Dados",
                "Não há dados na visualização atual para exportar.",
                parent=self.root
            )
            return

        try:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Arquivos Excel", "*.xlsx")],
                title="Salvar Relatório Orçado vs. Gasto",
                initialfile=f"Relatorio_Orcado_vs_Gasto_{self.combo_mes.get()}.xlsx"
            )

            if not file_path:
                return

            df_export = self.displayed_data_df.copy()
            df_export.rename(
                columns={
                    'AnoMes': 'Mês/Ano',
                    'conta_despesa': 'Categoria',
                    'Orcado': 'Valor Orçado (R$)',
                    'Gasto': 'Valor Gasto (R$)',
                    'Saldo': 'Saldo (R$)'
                },
                inplace=True
            )

            wb = Workbook()
            ws = wb.active
            ws.title = "Orcado_vs_Gasto"

            # Título
            ws['A1'] = f"Relatório Mensal: Orçado vs. Gasto ({self.combo_mes.get()})"
            ws.merge_cells('A1:E1')
            ws['A1'].font = Font(size=16, bold=True, color="000080")
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.row_dimensions[1].height = 25
            
            # Cabeçalhos
            headers = list(df_export.columns)
            for col_num, header_title in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_num, value=header_title)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="000080",
                    end_color="000080",
                    fill_type="solid"
                )
                cell.alignment = Alignment(horizontal='center')
                ws.column_dimensions[get_column_letter(col_num)].width = 20

            # Dados
            for r_idx, row in enumerate(df_export.values.tolist(), 4):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if isinstance(value, (int, float)):
                        cell.number_format = 'R$ #,##0.00'
                        if c_idx == headers.index('Saldo (R$)') + 1:
                            if value < 0:
                                cell.font = Font(color="FF0000")
                            elif value > 0:
                                cell.font = Font(color="008000")
            
            # Linha de totais
            total_row_idx = ws.max_row + 2
            total_orcado = df_export['Valor Orçado (R$)'].sum()
            total_gasto = df_export['Valor Gasto (R$)'].sum()
            total_saldo = df_export['Saldo (R$)'].sum()

            ws.cell(row=total_row_idx, column=2, value="TOTAIS").font = Font(bold=True)
            ws.cell(row=total_row_idx, column=3, value=total_orcado).number_format = 'R$ #,##0.00'
            ws.cell(row=total_row_idx, column=4, value=total_gasto).number_format = 'R$ #,##0.00'
            saldo_cell = ws.cell(row=total_row_idx, column=5, value=total_saldo)
            saldo_cell.number_format = 'R$ #,##0.00'
            if total_saldo < 0:
                saldo_cell.font = Font(color="FF0000", bold=True)
            elif total_saldo > 0:
                saldo_cell.font = Font(color="008000", bold=True)
            else:
                saldo_cell.font = Font(bold=True)

            # Ajustar largura das colunas
            for col_num, column_title in enumerate(headers, 1):
                max_length = len(column_title)
                column_letter = get_column_letter(col_num)
                for cell_tuple in ws[column_letter]:
                    try:
                        if len(str(cell_tuple.value)) > max_length:
                            max_length = len(str(cell_tuple.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width if adjusted_width < 50 else 50

            wb.save(file_path)
            messagebox.showinfo(
                "Exportação Concluída",
                f"Relatório exportado com sucesso para:\n{file_path}",
                parent=self.root
            )

        except Exception as e:
            messagebox.showerror("Erro na Exportação", f"Não foi possível exportar o relatório: {e}", parent=self.root)
            import traceback
            traceback.print_exc()

    def __del__(self):
        if hasattr(self, 'db_conn') and self.db_conn:
            self.db_conn.close()


def iniciar_relatorio_orcado_vs_gasto(parent):
    RelatorioOrcadoVsGasto(parent)
