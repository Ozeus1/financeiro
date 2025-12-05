import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
from datetime import datetime
from dateutil.relativedelta import relativedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Novas importações para o gráfico
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

# Estrutura simplificada para armazenar os detalhes de uma parcela
class DetalheFatura:
    def __init__(self, despesa_id, descricao, valor_parcela, num_parcela_atual, total_parcelas, data_compra):
        self.despesa_id = despesa_id
        self.descricao = descricao
        self.valor_parcela = valor_parcela
        self.num_parcela_atual = num_parcela_atual
        self.total_parcelas = total_parcelas
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
from datetime import datetime
from dateutil.relativedelta import relativedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Novas importações para o gráfico
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

# Estrutura simplificada para armazenar os detalhes de uma parcela
class DetalheFatura:
    def __init__(self, despesa_id, descricao, valor_parcela, num_parcela_atual, total_parcelas, data_compra):
        self.despesa_id = despesa_id
        self.descricao = descricao
        self.valor_parcela = valor_parcela
        self.num_parcela_atual = num_parcela_atual
        self.total_parcelas = total_parcelas
        self.data_compra = data_compra
        # ID único da parcela para referência interna no Treeview
        self.unique_id_parcela = f"{self.despesa_id}_{self.num_parcela_atual}"


class RelatorioPrevisaoFaturas:
    def __init__(self, parent):
        self.root = tk.Toplevel(parent)
        self.root.title("Relatório de Previsão de Faturas de Cartão")
        self.root.geometry("1200x800")
        self.root.resizable(True, True)
        self.root.transient(parent)
        self.root.grab_set()

        self.db_conn = sqlite3.connect('financas.db')
        self.db_cursor = self.db_conn.cursor()

        self.cartoes_com_fechamento = {}
        self.despesas_originais = pd.DataFrame()
        self.previsao_faturas = {}
        self.consolidado_mensal = {}

        ### NOVO: Variáveis para controlar a ordenação da tabela de detalhes ###
        self.detalhes_sort_column = 'descricao'
        self.detalhes_sort_reverse = False
        
        self.carregar_dados_iniciais()
        self.criar_widgets()

        if self.cartoes_com_fechamento:
            self.combo_cartoes.set("Todos os Cartões (Consolidado)")
            self.on_cartao_selecionado()
        else:
            messagebox.showinfo("Aviso",
                                "Nenhum cartão com data de fechamento cadastrada foi encontrado.\n"
                                "Cadastre-os em 'Configurar Fechamento de Cartões' para usar este relatório.",
                                parent=self.root)

    def carregar_dados_iniciais(self):
        try:
            self.db_cursor.execute("SELECT meio_pagamento, data_fechamento FROM fechamento_cartoes")
            self.cartoes_com_fechamento = {row[0]: row[1] for row in self.db_cursor.fetchall()}

            # Use v_despesas_compat and filter by user_id
            query_despesas = "SELECT id, descricao, meio_pagamento, valor, num_parcelas, data_pagamento FROM v_despesas_compat"
            self.despesas_originais = pd.read_sql_query(query_despesas, self.db_conn)
            if not self.despesas_originais.empty:
                self.despesas_originais['data_pagamento'] = pd.to_datetime(self.despesas_originais['data_pagamento'])

        except Exception as e:
            messagebox.showerror("Erro ao Carregar Dados", f"Falha ao carregar dados iniciais: {e}", parent=self.root)
            self.despesas_originais = pd.DataFrame()

    def calcular_mes_fatura(self, data_compra, dia_fechamento_cartao):
        if data_compra.day > dia_fechamento_cartao:
            return (data_compra + relativedelta(months=1)).replace(day=1)
        else:
            return data_compra.replace(day=1)

    def processar_previsao_para_cartao(self, nome_cartao):
        if nome_cartao in self.previsao_faturas:
            return

        self.previsao_faturas[nome_cartao] = {}
        if nome_cartao not in self.cartoes_com_fechamento or self.despesas_originais.empty:
            return

        dia_fechamento = self.cartoes_com_fechamento[nome_cartao]
        despesas_do_cartao = self.despesas_originais[self.despesas_originais['meio_pagamento'] == nome_cartao]

        for _, despesa in despesas_do_cartao.iterrows():
            data_compra_dt = despesa['data_pagamento']
            valor_total = float(despesa['valor'])
            num_parcelas_total = int(despesa['num_parcelas'])
            valor_parcela = valor_total / num_parcelas_total if num_parcelas_total > 0 else 0

            primeiro_mes_fatura_dt = self.calcular_mes_fatura(data_compra_dt, dia_fechamento)

            for i in range(num_parcelas_total):
                mes_fatura_atual_dt = primeiro_mes_fatura_dt + relativedelta(months=i)
                chave_fatura = mes_fatura_atual_dt.strftime('%Y-%m')

                if chave_fatura not in self.previsao_faturas[nome_cartao]:
                    self.previsao_faturas[nome_cartao][chave_fatura] = {'total': 0.0, 'detalhes': []}

                detalhe = DetalheFatura(
                    despesa_id=despesa['id'],
                    descricao=despesa['descricao'],
                    valor_parcela=valor_parcela,
                    num_parcela_atual=i + 1,
                    total_parcelas=num_parcelas_total,
                    data_compra=data_compra_dt
                )
                self.previsao_faturas[nome_cartao][chave_fatura]['detalhes'].append(detalhe)

    def processar_previsao_consolidada(self):
        self.consolidado_mensal = {}

        for nome_cartao in self.cartoes_com_fechamento.keys():
            self.processar_previsao_para_cartao(nome_cartao)

            if nome_cartao in self.previsao_faturas:
                for chave_fatura, data_fatura in self.previsao_faturas[nome_cartao].items():
                    total_fatura_cartao = sum(d.valor_parcela for d in data_fatura['detalhes'])
                    self.previsao_faturas[nome_cartao][chave_fatura]['total'] = total_fatura_cartao

        for faturas_cartao in self.previsao_faturas.values():
            for mes, dados_fatura in faturas_cartao.items():
                self.consolidado_mensal[mes] = self.consolidado_mensal.get(mes, 0.0) + dados_fatura.get('total', 0.0)

    def criar_widgets(self):
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True)

        frame_controles = ttk.Frame(main_frame, padding=10)
        frame_controles.pack(fill=tk.X)

        ttk.Label(frame_controles, text="Selecionar Visão:").pack(side=tk.LEFT, padx=(0, 5))
        opcoes_combo = ["Todos os Cartões (Consolidado)"] + list(self.cartoes_com_fechamento.keys())
        self.combo_cartoes = ttk.Combobox(frame_controles, values=opcoes_combo, state="readonly", width=30)
        self.combo_cartoes.pack(side=tk.LEFT, padx=(0, 20))
        self.combo_cartoes.bind("<<ComboboxSelected>>", self.on_cartao_selecionado)

        self.btn_exportar_excel = ttk.Button(frame_controles, text="Exportar Previsão Completa", command=self.exportar_para_excel)
        self.btn_exportar_excel.pack(side=tk.LEFT, padx=(10, 0))

        paned_window = ttk.PanedWindow(main_frame, orient=tk.VERTICAL)
        paned_window.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        top_pane = ttk.PanedWindow(paned_window, orient=tk.HORIZONTAL)
        paned_window.add(top_pane, weight=1)

        frame_faturas_mensais = ttk.LabelFrame(top_pane, text="Faturas Previstas (Mês a Mês)", padding=10)
        top_pane.add(frame_faturas_mensais, weight=1)

        self.tree_faturas_mensais = ttk.Treeview(frame_faturas_mensais, columns=('mes_fatura', 'valor_previsto'), show='headings')
        self.tree_faturas_mensais.heading('mes_fatura', text='Mês da Fatura')
        self.tree_faturas_mensais.heading('valor_previsto', text='Valor Previsto (R$)')
        self.tree_faturas_mensais.column('mes_fatura', width=120, anchor='w')
        self.tree_faturas_mensais.column('valor_previsto', width=150, anchor='e')
        self.tree_faturas_mensais.pack(fill=tk.BOTH, expand=True)
        self.tree_faturas_mensais.bind("<<TreeviewSelect>>", self.on_fatura_mensal_selecionada)

        self.frame_detalhes_fatura = ttk.LabelFrame(top_pane, text="Detalhes da Fatura do Mês", padding=10)
        top_pane.add(self.frame_detalhes_fatura, weight=2)
        
        # ### NOVO: Frame para o botão de exportar detalhes ###
        frame_botoes_detalhes = ttk.Frame(self.frame_detalhes_fatura)
        frame_botoes_detalhes.pack(fill=tk.X, pady=(0, 5))

        self.btn_exportar_detalhes = ttk.Button(frame_botoes_detalhes, text="Exportar Detalhes", command=self.exportar_detalhes_para_excel)
        self.btn_exportar_detalhes.pack(side=tk.RIGHT)

        colunas_detalhes = ('descricao', 'data_compra', 'parcela_info', 'valor_parcela')
        self.tree_detalhes_fatura = ttk.Treeview(self.frame_detalhes_fatura,
                                                 columns=colunas_detalhes,
                                                 show='headings')
        self.tree_detalhes_fatura.heading('descricao', text='Descrição')
        self.tree_detalhes_fatura.heading('data_compra', text='Data Compra')
        self.tree_detalhes_fatura.heading('parcela_info', text='Parcela')
        self.tree_detalhes_fatura.heading('valor_parcela', text='Valor Parcela (R$)')

        # ### ALTERAÇÃO: Adiciona o comando de ordenação aos cabeçalhos ###
        for col in colunas_detalhes:
            self.tree_detalhes_fatura.heading(col, command=lambda c=col: self.definir_ordenacao_detalhes(c))

        self.tree_detalhes_fatura.column('descricao', width=280, anchor='w')
        self.tree_detalhes_fatura.column('data_compra', width=100, anchor='center')
        self.tree_detalhes_fatura.column('parcela_info', width=80, anchor='center')
        self.tree_detalhes_fatura.column('valor_parcela', width=120, anchor='e')
        self.tree_detalhes_fatura.pack(fill=tk.BOTH, expand=True)

        # Frame para análise e gráfico
        bottom_pane = ttk.LabelFrame(paned_window, text="Análise de Gastos Futuros", padding=10)
        paned_window.add(bottom_pane, weight=1)

        frame_analise_controles = ttk.Frame(bottom_pane)
        frame_analise_controles.pack(fill=tk.X, pady=5)

        ttk.Label(frame_analise_controles, text="Mês Inicial para Análise:").pack(side=tk.LEFT, padx=(0, 5))

        self.mes_analise_var = tk.StringVar()
        self.combo_mes_analise = ttk.Combobox(frame_analise_controles, textvariable=self.mes_analise_var, state="readonly", width=12)
        self.combo_mes_analise.pack(side=tk.LEFT, padx=5)
        self.combo_mes_analise.bind("<<ComboboxSelected>>", self.atualizar_total_e_grafico)

        self.lbl_total_devido_texto = ttk.Label(frame_analise_controles, text="Valor Total Devido (a partir do mês informado):", font=('Helvetica', 10, 'bold'))
        self.lbl_total_devido_texto.pack(side=tk.LEFT, padx=(20, 5))

        self.total_devido_var = tk.StringVar(value="R$ 0,00")
        self.lbl_total_devido_valor = ttk.Label(frame_analise_controles, textvariable=self.total_devido_var, font=('Helvetica', 10, 'bold'), foreground='blue')
        self.lbl_total_devido_valor.pack(side=tk.LEFT)

        # Widget do Gráfico
        self.figura_grafico = Figure(figsize=(8, 4), dpi=100)
        self.ax_grafico = self.figura_grafico.add_subplot(111)

        self.canvas_grafico = FigureCanvasTkAgg(self.figura_grafico, master=bottom_pane)
        self.canvas_grafico.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True, pady=(10,0))
        self.canvas_grafico.draw()

    ### NOVO: Método para definir a coluna e a direção da ordenação ###
    def definir_ordenacao_detalhes(self, col):
        if self.detalhes_sort_column == col:
            self.detalhes_sort_reverse = not self.detalhes_sort_reverse
        else:
            self.detalhes_sort_column = col
            self.detalhes_sort_reverse = False
        
        # Re-popula a tabela com a nova ordenação
        self.on_fatura_mensal_selecionada()

    def atualizar_combo_meses(self):
        visao_selecionada = self.combo_cartoes.get()
        meses_disponiveis = []

        if visao_selecionada == "Todos os Cartões (Consolidado)":
            if self.consolidado_mensal:
                meses_disponiveis = sorted(self.consolidado_mensal.keys())
        else:
            if visao_selecionada in self.previsao_faturas:
                meses_disponiveis = sorted(self.previsao_faturas[visao_selecionada].keys())

        self.combo_mes_analise['values'] = meses_disponiveis

        if meses_disponiveis:
            mes_anterior = (datetime.now() - relativedelta(months=1)).strftime('%Y-%m')
            if mes_anterior in meses_disponiveis:
                self.mes_analise_var.set(mes_anterior)
            else:
                self.mes_analise_var.set(meses_disponiveis[0])
        else:
            self.mes_analise_var.set('')

    def atualizar_total_e_grafico(self, event=None):
        mes_inicio_str = self.mes_analise_var.get()
        if not mes_inicio_str:
            self.total_devido_var.set("R$ 0,00")
            self.ax_grafico.clear()
            self.ax_grafico.set_title("Nenhum mês selecionado para análise")
            self.canvas_grafico.draw()
            return

        visao_selecionada = self.combo_cartoes.get()
        dados_fonte = {}
        titulo_grafico = "Total Devido por Mês"

        if visao_selecionada == "Todos os Cartões (Consolidado)":
            dados_fonte = self.consolidado_mensal
            titulo_grafico = f'Total Devido por Mês (Consolidado)'
        else:
            dados_brutos_cartao = self.previsao_faturas.get(visao_selecionada, {})
            dados_fonte = {mes: data['total'] for mes, data in dados_brutos_cartao.items()}
            titulo_grafico = f'Total Devido por Mês ({visao_selecionada})'

        meses_filtrados = {m: v for m, v in dados_fonte.items() if m >= mes_inicio_str}
        total_devido = sum(meses_filtrados.values())
        self.total_devido_var.set(f"R$ {total_devido:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))

        self.ax_grafico.clear()
        if meses_filtrados:
            chaves_ordenadas = sorted(meses_filtrados.keys())
            valores = [meses_filtrados[m] for m in chaves_ordenadas]
            bars = self.ax_grafico.bar(chaves_ordenadas, valores, color='skyblue')
            self.ax_grafico.set_title(f'{titulo_grafico} - a partir de {mes_inicio_str}', fontsize=12)
            self.ax_grafico.set_ylabel('Valor Previsto (R$)')
            self.ax_grafico.set_xlabel('Mês da Fatura')
            self.figura_grafico.autofmt_xdate(rotation=45)
            self.ax_grafico.grid(axis='y', linestyle='--', alpha=0.7)

            for bar in bars:
                height = bar.get_height()
                label_text = f'{height:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
                self.ax_grafico.annotate(label_text,
                                         xy=(bar.get_x() + bar.get_width() / 2, height),
                                         xytext=(0, 3), textcoords="offset points",
                                         ha='center', va='bottom', fontsize=9)
        else:
             self.ax_grafico.set_title('Nenhum dado para o período selecionado', fontsize=12)

        self.figura_grafico.tight_layout()
        self.canvas_grafico.draw()

    def on_cartao_selecionado(self, event=None):
        selecionado = self.combo_cartoes.get()
        self.limpar_tree_detalhes_fatura()

        if selecionado == "Todos os Cartões (Consolidado)":
            self.frame_detalhes_fatura.config(text="Detalhes (Indisponível na visão consolidada)")
            self.btn_exportar_excel.config(text="Exportar para Excel", state=tk.DISABLED) # Alterado texto para claridade
            self.btn_exportar_detalhes.config(state=tk.DISABLED) ### NOVO ###
            self.processar_previsao_consolidada()
            self.popular_tree_consolidado()
        else:
            self.frame_detalhes_fatura.config(text="Detalhes da Fatura do Mês")
            self.btn_exportar_excel.config(text="Exportar Previsão Completa", state=tk.NORMAL) # Alterado texto para claridade
            self.btn_exportar_detalhes.config(state=tk.NORMAL) ### NOVO ###
            self.processar_previsao_para_cartao(selecionado)
            self.atualizar_tree_faturas_mensais(selecionado)
        
        self.atualizar_combo_meses()
        self.atualizar_total_e_grafico()

    def popular_tree_consolidado(self):
        self.tree_faturas_mensais.delete(*self.tree_faturas_mensais.get_children())
        if self.consolidado_mensal:
            chaves_ordenadas = sorted(self.consolidado_mensal.keys())
            for mes in chaves_ordenadas:
                total = self.consolidado_mensal[mes]
                self.tree_faturas_mensais.insert('', tk.END, iid=mes,
                                                 values=(mes, f"R$ {total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')))

    def atualizar_tree_faturas_mensais(self, nome_cartao):
        self.tree_faturas_mensais.delete(*self.tree_faturas_mensais.get_children())
        if nome_cartao not in self.previsao_faturas:
            return

        faturas_cartao = self.previsao_faturas[nome_cartao]
        for chave_fatura in faturas_cartao:
            faturas_cartao[chave_fatura]['total'] = sum(d.valor_parcela for d in faturas_cartao[chave_fatura]['detalhes'])

        chaves_ordenadas = sorted(faturas_cartao.keys())
        for chave_fatura in chaves_ordenadas:
            total_fatura = faturas_cartao[chave_fatura]['total']
            self.tree_faturas_mensais.insert('', tk.END, iid=chave_fatura,
                                             values=(chave_fatura, f"R$ {total_fatura:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')))

    def on_fatura_mensal_selecionada(self, event=None):
        if self.combo_cartoes.get() == "Todos os Cartões (Consolidado)":
            return

        selecionado = self.tree_faturas_mensais.selection()
        if not selecionado:
            return

        chave_fatura_selecionada = selecionado[0]
        cartao_atual = self.combo_cartoes.get()
        self.limpar_tree_detalhes_fatura()

        if cartao_atual in self.previsao_faturas and chave_fatura_selecionada in self.previsao_faturas[cartao_atual]:
            detalhes_da_fatura = self.previsao_faturas[cartao_atual][chave_fatura_selecionada]['detalhes']
            
            # ### ALTERAÇÃO: Lógica de ordenação aplicada antes de popular a tabela ###
            sort_key_func = None
            col = self.detalhes_sort_column
            if col == 'descricao':
                sort_key_func = lambda d: d.descricao.lower()
            elif col == 'data_compra':
                sort_key_func = lambda d: d.data_compra
            elif col == 'parcela_info':
                # Ordena pelo número da parcela atual
                sort_key_func = lambda d: d.num_parcela_atual
            elif col == 'valor_parcela':
                sort_key_func = lambda d: d.valor_parcela
            
            if sort_key_func:
                detalhes_da_fatura.sort(key=sort_key_func, reverse=self.detalhes_sort_reverse)

            for detalhe in detalhes_da_fatura:
                self.tree_detalhes_fatura.insert('', tk.END,
                    iid=detalhe.unique_id_parcela,
                    values=(
                        detalhe.descricao,
                        detalhe.data_compra.strftime('%d/%m/%Y'),
                        f"{detalhe.num_parcela_atual}/{detalhe.total_parcelas}",
                        f"R$ {detalhe.valor_parcela:.2f}".replace('.', ',')
                    ))

    def exportar_para_excel(self):
        cartao_selecionado = self.combo_cartoes.get()
        if not cartao_selecionado or cartao_selecionado == "Todos os Cartões (Consolidado)":
            messagebox.showwarning("Ação Inválida", "Por favor, selecione um cartão específico para exportar.", parent=self.root)
            return

        if not self.previsao_faturas or cartao_selecionado not in self.previsao_faturas or not self.previsao_faturas[cartao_selecionado]:
            messagebox.showwarning("Sem Dados", f"Não há dados de previsão para o cartão '{cartao_selecionado}'.", parent=self.root)
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Arquivos Excel", "*.xlsx")],
            title=f"Exportar Previsão Completa - {cartao_selecionado}",
            initialfile=f"Previsao_Faturas_{cartao_selecionado.replace(' ', '_')}.xlsx"
        )

        if not file_path:
            return

        try:
            wb = Workbook()
            faturas_do_cartao = self.previsao_faturas[cartao_selecionado]
            chaves_ordenadas = sorted(faturas_do_cartao.keys())

            ws_resumo = wb.active
            ws_resumo.title = f"Resumo - {cartao_selecionado[:20]}"
            ws_resumo['A1'] = f"Resumo das Faturas Previstas - Cartão: {cartao_selecionado}"
            ws_resumo.merge_cells('A1:B1')
            ws_resumo['A1'].font = Font(bold=True, size=14)
            ws_resumo['A1'].alignment = Alignment(horizontal='center')
            headers_resumo = ["Mês da Fatura", "Valor Previsto (R$)"]
            for col_num, header_title in enumerate(headers_resumo, 1):
                cell = ws_resumo.cell(row=3, column=col_num, value=header_title)
                cell.font = Font(bold=True); cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"); cell.alignment = Alignment(horizontal='center')
            row_idx_resumo = 4
            for chave_fatura in chaves_ordenadas:
                total_fatura = faturas_do_cartao[chave_fatura]['total']
                ws_resumo.cell(row=row_idx_resumo, column=1, value=chave_fatura)
                cell_valor = ws_resumo.cell(row=row_idx_resumo, column=2, value=total_fatura)
                cell_valor.number_format = 'R$ #,##0.00'
                row_idx_resumo += 1
            ws_resumo.column_dimensions[get_column_letter(1)].width = 20
            ws_resumo.column_dimensions[get_column_letter(2)].width = 25

            ws_detalhes = wb.create_sheet(title=f"Detalhes - {cartao_selecionado[:20]}")
            ws_detalhes['A1'] = f"Detalhes dos Lançamentos Previstos - Cartão: {cartao_selecionado}"
            ws_detalhes.merge_cells('A1:E1') # ### ALTERAÇÃO ### - Mescla até a coluna E
            ws_detalhes['A1'].font = Font(bold=True, size=14); ws_detalhes['A1'].alignment = Alignment(horizontal='center')
            headers_detalhes = ["Mês da Fatura", "Descrição", "Data Compra", "Parcela", "Valor Parcela (R$)"] # ### ALTERAÇÃO ### - Adiciona Parcela
            for col_num, header_title in enumerate(headers_detalhes, 1):
                cell = ws_detalhes.cell(row=3, column=col_num, value=header_title)
                cell.font = Font(bold=True); cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"); cell.alignment = Alignment(horizontal='center')
            row_idx_detalhes = 4
            for chave_fatura in chaves_ordenadas:
                lista_detalhes = sorted(faturas_do_cartao[chave_fatura]['detalhes'], key=lambda d: d.data_compra)
                for detalhe in lista_detalhes:
                    ws_detalhes.cell(row=row_idx_detalhes, column=1, value=chave_fatura)
                    ws_detalhes.cell(row=row_idx_detalhes, column=2, value=detalhe.descricao)
                    ws_detalhes.cell(row=row_idx_detalhes, column=3, value=detalhe.data_compra.strftime('%d/%m/%Y'))
                    ws_detalhes.cell(row=row_idx_detalhes, column=4, value=f"{detalhe.num_parcela_atual}/{detalhe.total_parcelas}") # ### ALTERAÇÃO ###
                    cell_valor_parc = ws_detalhes.cell(row=row_idx_detalhes, column=5, value=detalhe.valor_parcela) # ### ALTERAÇÃO ###
                    cell_valor_parc.number_format = 'R$ #,##0.00'
                    row_idx_detalhes += 1
            widths_detalhes = [20, 35, 15, 15, 20] # ### ALTERAÇÃO ###
            for i, width in enumerate(widths_detalhes, 1):
                ws_detalhes.column_dimensions[get_column_letter(i)].width = width
            
            wb.save(file_path)
            messagebox.showinfo("Exportação Concluída", f"Relatório exportado com sucesso para:\n{file_path}", parent=self.root)
        except Exception as e:
            messagebox.showerror("Erro na Exportação", f"Não foi possível exportar o relatório: {e}", parent=self.root)
            import traceback
            traceback.print_exc()

    ### NOVO: Método para exportar apenas os detalhes exibidos na tela ###
    def exportar_detalhes_para_excel(self):
        cartao_selecionado = self.combo_cartoes.get()
        selecao_mes = self.tree_faturas_mensais.selection()

        if not selecao_mes:
            messagebox.showwarning("Ação Inválida", "Por favor, selecione um mês na tabela 'Faturas Previstas' para exportar os detalhes.", parent=self.root)
            return

        mes_fatura = selecao_mes[0]
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Arquivos Excel", "*.xlsx")],
            title=f"Exportar Detalhes - {cartao_selecionado} - {mes_fatura}",
            initialfile=f"Detalhes_Fatura_{cartao_selecionado.replace(' ', '_')}_{mes_fatura}.xlsx"
        )

        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = f"Detalhes {mes_fatura}"

            # Título
            ws['A1'] = f"Detalhes da Fatura de {mes_fatura} - Cartão: {cartao_selecionado}"
            ws.merge_cells('A1:D1')
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')

            # Cabeçalhos
            headers = ["Descrição", "Data Compra", "Parcela", "Valor Parcela (R$)"]
            for col_num, header_title in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_num, value=header_title)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Dados
            row_idx = 4
            for item_id in self.tree_detalhes_fatura.get_children():
                valores = self.tree_detalhes_fatura.item(item_id, 'values')
                ws.cell(row=row_idx, column=1, value=valores[0]) # Descrição
                ws.cell(row=row_idx, column=2, value=valores[1]) # Data Compra
                ws.cell(row=row_idx, column=3, value=valores[2]) # Parcela
                
                # Converte o valor monetário para número para formatar no Excel
                valor_str = valores[3].replace("R$", "").replace(".", "").replace(",", ".").strip()
                try:
                    valor_float = float(valor_str)
                    cell_valor = ws.cell(row=row_idx, column=4, value=valor_float)
                    cell_valor.number_format = 'R$ #,##0.00'
                except ValueError:
                    ws.cell(row=row_idx, column=4, value=valores[3])

                row_idx += 1

            # Ajuste de largura das colunas
            widths = [40, 15, 15, 20]
            for i, width in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

            wb.save(file_path)
            messagebox.showinfo("Exportação Concluída", f"Detalhes da fatura exportados com sucesso para:\n{file_path}", parent=self.root)

        except Exception as e:
            messagebox.showerror("Erro na Exportação", f"Não foi possível exportar os detalhes: {e}", parent=self.root)
            import traceback
            traceback.print_exc()

    def limpar_tree_detalhes_fatura(self):
        self.tree_detalhes_fatura.delete(*self.tree_detalhes_fatura.get_children())

    def __del__(self):
        if self.db_conn:
            self.db_conn.close()

def iniciar_relatorio_previsao_faturas(parent):
    RelatorioPrevisaoFaturas(parent)