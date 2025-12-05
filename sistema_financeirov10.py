import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import sqlite3
from datetime import datetime
import calendar
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.ticker import FuncFormatter
import os
import locale
from tkcalendar import DateEntry
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from tkinter import filedialog


# Tenta importar módulos locais, mas não quebra se não existirem
try:
    import configuracao
    import MENUBD
    import relclaude1
    import relatorios1
    import gerenciar_orcamento
    import relatorio_orcado_vs_gasto
    import gerenciar_fechamento_cartoes
    import relatorio_previsao_faturas
    import importador_excel
    import relatorio_balanco # <--- ADICIONE ESTA LINHA
    import relatorio_balanco_fluxo_caixa
    import importador_supabase # <--- ADICIONE ESTA LINHA


except ImportError:
    print("Aviso: Alguns módulos de extensão não foram encontrados. A funcionalidade completa pode estar limitada.")
    class PlaceholderModule:
        def __getattr__(self, name):
            def placeholder_func(*args, **kwargs):
                messagebox.showerror("Módulo Faltando", f"O módulo para esta função não foi encontrado.")
            return placeholder_func
    configuracao = MENUBD = relclaude1 = relatorios1 = gerenciar_orcamento = \
    relatorio_orcado_vs_gasto = gerenciar_fechamento_cartoes = \
    relatorio_previsao_faturas = importador_excel = PlaceholderModule()

# Configurar a localização para formato de moeda brasileiro
try:
    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
except locale.Error:
    try:
        locale.setlocale(locale.LC_ALL, 'Portuguese_Brazil.1252')
    except locale.Error:
        print("Aviso: Locale 'pt_BR' não encontrado. A formatação de moeda pode não funcionar corretamente.")



# --- CLASSE PARA GERENCIAR CATEGORIAS DE RECEITA ---
class GerenciadorCategoriasReceita(tk.Toplevel):
    """Janela para gerenciar (adicionar, editar, excluir) categorias de receita."""
    def __init__(self, parent_widget, app_logic):
        super().__init__(parent_widget)
        self.transient(parent_widget)
        self.grab_set()
        self.title("Gerenciar Categorias de Receita")
        self.geometry("450x400")

        # Conexão com o banco de dados de RECEITAS
        self.conn_receitas = app_logic.conn_receitas
        self.cursor_receitas = app_logic.cursor_receitas

        # Variáveis de controle
        self.id_categoria_var = tk.StringVar()
        self.nome_categoria_var = tk.StringVar()

        # Frame principal
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Frame da lista
        frame_lista = ttk.LabelFrame(main_frame, text="Categorias Existentes")
        frame_lista.pack(fill=tk.BOTH, expand=True, pady=5)
        
        self.lista_categorias = tk.Listbox(frame_lista, height=10)
        self.lista_categorias.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.lista_categorias.bind("<<ListboxSelect>>", self.selecionar_categoria)

        scrollbar = ttk.Scrollbar(frame_lista, orient="vertical", command=self.lista_categorias.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.lista_categorias.config(yscrollcommand=scrollbar.set)
        
        # Frame do formulário e botões
        frame_form = ttk.LabelFrame(main_frame, text="Adicionar / Editar Categoria")
        frame_form.pack(fill=tk.X, pady=10)

        ttk.Label(frame_form, text="ID:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        ttk.Entry(frame_form, textvariable=self.id_categoria_var, state='readonly', width=10).grid(row=0, column=1, padx=5, pady=5, sticky="w")
        
        ttk.Label(frame_form, text="Nome:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        entry_categoria = ttk.Entry(frame_form, textvariable=self.nome_categoria_var, width=40)
        entry_categoria.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
        frame_form.columnconfigure(1, weight=1)

        frame_botoes = ttk.Frame(main_frame)
        frame_botoes.pack(fill=tk.X, pady=5)
        
        btn_adicionar = ttk.Button(frame_botoes, text="Adicionar Novo", command=self.adicionar_categoria)
        btn_adicionar.pack(side=tk.LEFT, padx=5)

        self.btn_atualizar = ttk.Button(frame_botoes, text="Salvar Edição", command=self.atualizar_categoria, state="disabled")
        self.btn_atualizar.pack(side=tk.LEFT, padx=5)

        self.btn_excluir = ttk.Button(frame_botoes, text="Excluir Selecionado", command=self.excluir_categoria, state="disabled")
        self.btn_excluir.pack(side=tk.LEFT, padx=5)
        
        btn_limpar = ttk.Button(frame_botoes, text="Limpar Formulário", command=self.limpar_campos)
        btn_limpar.pack(side=tk.LEFT, padx=5)
        
        # Carrega os dados e ativa a janela
        self.carregar_categorias()

    def carregar_categorias(self):
        """Carrega a lista e seleciona o primeiro item para edição."""
        # Limpa a seleção e o formulário antes de carregar
        self.limpar_campos()
        self.lista_categorias.delete(0, tk.END)
        
        # Busca e insere as categorias na lista
        self.cursor_receitas.execute("SELECT nome FROM categorias_receita ORDER BY nome")
        categorias = self.cursor_receitas.fetchall()
        for categoria in categorias:
            self.lista_categorias.insert(tk.END, categoria[0])
            
        # NOVA LÓGICA: Seleciona o primeiro item da lista, se houver algum
        if self.lista_categorias.size() > 0:
            self.lista_categorias.selection_set(0)
            self.selecionar_categoria() # Chama o evento para popular o formulário

    def limpar_campos(self):
        """Limpa o formulário e desativa os botões de ação."""
        self.id_categoria_var.set("")
        self.nome_categoria_var.set("")
        if self.lista_categorias.curselection():
            self.lista_categorias.selection_clear(0, tk.END)
        self.btn_atualizar.config(state="disabled")
        self.btn_excluir.config(state="disabled")

    def selecionar_categoria(self, event=None):
        """Preenche o formulário quando um item da lista é selecionado."""
        selecionado_indices = self.lista_categorias.curselection()
        if not selecionado_indices:
            # Se não houver seleção, limpa os campos.
            self.id_categoria_var.set("")
            self.nome_categoria_var.set("")
            self.btn_atualizar.config(state="disabled")
            self.btn_excluir.config(state="disabled")
            return
        
        nome_selecionado = self.lista_categorias.get(selecionado_indices[0])
        self.cursor_receitas.execute("SELECT id FROM categorias_receita WHERE nome = ?", (nome_selecionado,))
        resultado = self.cursor_receitas.fetchone()
        
        if resultado:
            self.id_categoria_var.set(resultado[0])
            self.nome_categoria_var.set(nome_selecionado)
            self.btn_atualizar.config(state="normal")
            self.btn_excluir.config(state="normal")

    def adicionar_categoria(self):
        """Adiciona uma nova categoria com base no texto do formulário."""
        nome = self.nome_categoria_var.get().strip()
        if not nome:
            messagebox.showerror("Erro", "O nome da categoria não pode estar vazio.", parent=self)
            return
        # Limpa o campo de ID para garantir que não seja uma atualização
        self.id_categoria_var.set("")
        try:
            self.cursor_receitas.execute("INSERT INTO categorias_receita (nome) VALUES (?)", (nome,))
            self.conn_receitas.commit()
            self.carregar_categorias()
        except sqlite3.IntegrityError:
            messagebox.showerror("Erro", f"A categoria '{nome}' já existe.", parent=self)
        except Exception as e:
            messagebox.showerror("Erro no Banco de Dados", f"Erro ao adicionar categoria: {e}", parent=self)

    def atualizar_categoria(self):
        """Salva as edições feitas no nome da categoria selecionada."""
        id_cat = self.id_categoria_var.get()
        novo_nome = self.nome_categoria_var.get().strip()

        if not id_cat or not novo_nome:
            messagebox.showwarning("Dados Incompletos", "Selecione uma categoria da lista para editar.", parent=self)
            return
        
        try:
            self.cursor_receitas.execute("SELECT nome FROM categorias_receita WHERE id = ?", (id_cat,))
            resultado = self.cursor_receitas.fetchone()
            if not resultado:
                messagebox.showerror("Erro", "Categoria de receita não encontrada.", parent=self)
                return
            nome_antigo = resultado[0]

            if nome_antigo == novo_nome:
                messagebox.showinfo("Informação", "O nome não foi alterado.", parent=self)
                return

            self.cursor_receitas.execute("UPDATE categorias_receita SET nome = ? WHERE id = ?", (novo_nome, id_cat))
            self.cursor_receitas.execute("UPDATE receitas SET conta_receita = ? WHERE conta_receita = ?", (novo_nome, nome_antigo))
            
            self.conn_receitas.commit()
            
            messagebox.showinfo("Sucesso", f"Categoria '{nome_antigo}' foi atualizada para '{novo_nome}' com sucesso.", parent=self)
            # Recarrega a lista e seleciona o item recém-editado
            self.carregar_categorias()
            # Encontrar o novo item na lista e selecioná-lo
            items = self.lista_categorias.get(0, tk.END)
            if novo_nome in items:
                idx = items.index(novo_nome)
                self.lista_categorias.selection_set(idx)
                self.selecionar_categoria()

        except sqlite3.IntegrityError:
            messagebox.showerror("Erro de Integridade", f"O nome '{novo_nome}' já está em uso.", parent=self)
            self.conn_receitas.rollback()
        except Exception as e:
            messagebox.showerror("Erro ao Atualizar", f"Erro no banco de dados: {e}", parent=self)
            self.conn_receitas.rollback()

    def excluir_categoria(self):
        """Exclui a categoria selecionada, se não estiver em uso."""
        id_cat = self.id_categoria_var.get()
        nome_cat = self.nome_categoria_var.get()

        if not id_cat:
            messagebox.showerror("Erro", "Selecione uma categoria da lista para excluir.", parent=self)
            return

        self.cursor_receitas.execute("SELECT COUNT(*) FROM receitas WHERE conta_receita = ?", (nome_cat,))
        count = self.cursor_receitas.fetchone()[0]

        if count > 0:
            messagebox.showwarning("Impossível Excluir", 
                                 f"A categoria '{nome_cat}' está sendo usada em {count} registros de receita e não pode ser excluída.",
                                 parent=self)
            return

        if messagebox.askyesno("Confirmar Exclusão", f"Tem certeza que deseja excluir a categoria '{nome_cat}'?", parent=self):
            try:
                self.cursor_receitas.execute("DELETE FROM categorias_receita WHERE id = ?", (id_cat,))
                self.conn_receitas.commit()
                self.carregar_categorias()
            except Exception as e:
                messagebox.showerror("Erro no Banco de Dados", f"Erro ao excluir categoria: {e}", parent=self)




###################################################
  

# --- CLASSE PARA GERENCIAR RECEITAS ---
class GerenciadorReceitas(tk.Toplevel):
    """Janela para gerenciar (cadastrar, editar, excluir) receitas."""
    def __init__(self, parent_widget, app_logic):
        super().__init__(parent_widget)
        self.transient(parent_widget)
        self.grab_set()
        self.title("Gerenciador de Receitas")
        self.geometry("950x600")

        self.parent_app = app_logic
        self.conn_receitas = self.parent_app.conn_receitas
        self.cursor_receitas = self.parent_app.cursor_receitas

        self.id_receita = tk.StringVar()
        self.descricao = tk.StringVar()
        self.meio_recebimento = tk.StringVar()
        self.conta_receita = tk.StringVar()
        self.valor = tk.DoubleVar()
        
        self.criar_widgets()
        self.carregar_receitas()

    def carregar_combos(self):
        self.cursor_receitas.execute("SELECT nome FROM categorias_receita ORDER BY nome")
        self.combo_conta_receita['values'] = [row[0] for row in self.cursor_receitas.fetchall()]
        
        self.cursor_receitas.execute("SELECT nome FROM meios_recebimento ORDER BY nome")
        self.combo_meio_recebimento['values'] = [row[0] for row in self.cursor_receitas.fetchall()]

    def criar_widgets(self):
        frame_principal = ttk.Frame(self, padding="10")
        frame_principal.pack(fill=tk.BOTH, expand=True)

        frame_form = ttk.LabelFrame(frame_principal, text="Lançamento de Receita")
        frame_form.pack(fill=tk.X, pady=10)
        frame_form.columnconfigure(1, weight=1)

        ttk.Label(frame_form, text="Descrição:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        ttk.Entry(frame_form, textvariable=self.descricao, width=40).grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(frame_form, text="Valor (R$):").grid(row=0, column=2, padx=5, pady=5, sticky="w")
        ttk.Entry(frame_form, textvariable=self.valor, width=15).grid(row=0, column=3, padx=5, pady=5, sticky="w")

        ttk.Label(frame_form, text="Categoria:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.combo_conta_receita = ttk.Combobox(frame_form, textvariable=self.conta_receita, state='readonly')
        self.combo_conta_receita.grid(row=1, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(frame_form, text="Meio Recebimento:").grid(row=1, column=2, padx=5, pady=5, sticky="w")
        self.combo_meio_recebimento = ttk.Combobox(frame_form, textvariable=self.meio_recebimento, width=20, state='readonly')
        self.combo_meio_recebimento.grid(row=1, column=3, padx=5, pady=5, sticky="w")
        
        ttk.Label(frame_form, text="Data Recebimento:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.data_entry = DateEntry(frame_form, width=12, date_pattern='dd/mm/yyyy', locale='pt_BR')
        self.data_entry.grid(row=2, column=1, padx=5, pady=5, sticky="w")
        self.data_entry.set_date(datetime.now())

        self.carregar_combos()

        frame_botoes = ttk.Frame(frame_form)
        frame_botoes.grid(row=3, column=0, columnspan=4, pady=10)
        ttk.Button(frame_botoes, text="Salvar", command=self.salvar_receita).pack(side=tk.LEFT, padx=5)
        self.btn_atualizar = ttk.Button(frame_botoes, text="Atualizar", command=self.atualizar_receita, state="disabled")
        self.btn_atualizar.pack(side=tk.LEFT, padx=5)
        self.btn_excluir = ttk.Button(frame_botoes, text="Excluir", command=self.excluir_receita, state="disabled")
        self.btn_excluir.pack(side=tk.LEFT, padx=5)
        ttk.Button(frame_botoes, text="Limpar", command=self.limpar_campos).pack(side=tk.LEFT, padx=5)

        frame_tabela = ttk.LabelFrame(frame_principal, text="Receitas Lançadas")
        frame_tabela.pack(fill=tk.BOTH, expand=True, pady=10)
        
        scrollbar = ttk.Scrollbar(frame_tabela)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tabela = ttk.Treeview(frame_tabela, yscrollcommand=scrollbar.set, selectmode="browse")
        self.tabela['columns'] = ('id', 'data', 'descricao', 'categoria', 'meio', 'valor')
        
        self.tabela.column("#0", width=0, stretch=tk.NO)
        self.tabela.column("id", anchor=tk.CENTER, width=40)
        self.tabela.column("data", anchor=tk.CENTER, width=100)
        self.tabela.column("descricao", anchor=tk.W, width=300)
        self.tabela.column("categoria", anchor=tk.W, width=150)
        self.tabela.column("meio", anchor=tk.W, width=150)
        self.tabela.column("valor", anchor=tk.E, width=120)

        self.tabela.heading("id", text="ID")
        self.tabela.heading("data", text="Data")
        self.tabela.heading("descricao", text="Descrição")
        self.tabela.heading("categoria", text="Categoria")
        self.tabela.heading("meio", text="Meio")
        self.tabela.heading("valor", text="Valor")
        
        self.tabela.pack(fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.tabela.yview)
        self.tabela.bind("<ButtonRelease-1>", self.selecionar_item)

    def carregar_receitas(self):
        for i in self.tabela.get_children():
            self.tabela.delete(i)
        self.cursor_receitas.execute("""
            SELECT id, strftime('%d/%m/%Y', data_recebimento), descricao, conta_receita, meio_recebimento, valor
            FROM receitas ORDER BY data_recebimento DESC
        """)
        for row in self.cursor_receitas.fetchall():
            valor_formatado = locale.currency(row[5], grouping=True, symbol=True)
            self.tabela.insert("", tk.END, values=(row[0], row[1], row[2], row[3], row[4], valor_formatado))

    def limpar_campos(self):
        self.id_receita.set("")
        self.descricao.set("")
        self.conta_receita.set("")
        self.meio_recebimento.set("")
        self.valor.set(0.0)
        self.data_entry.set_date(datetime.now())
        self.btn_atualizar.config(state="disabled")
        self.btn_excluir.config(state="disabled")
        if self.tabela.selection():
            self.tabela.selection_remove(self.tabela.selection())

    def salvar_receita(self):
        if not self.descricao.get() or self.valor.get() <= 0 or not self.conta_receita.get():
            messagebox.showerror("Erro de Validação", "Preencha todos os campos obrigatórios (Descrição, Valor > 0, Categoria).", parent=self)
            return

        try:
            data_recebimento = self.data_entry.get_date().strftime('%Y-%m-%d')
            self.cursor_receitas.execute("""
                INSERT INTO receitas (descricao, conta_receita, meio_recebimento, valor, data_recebimento, data_registro, num_parcelas)
                VALUES (?, ?, ?, ?, ?, date('now'), 1)
            """, (self.descricao.get(), self.conta_receita.get(), self.meio_recebimento.get(), self.valor.get(), data_recebimento))
            self.conn_receitas.commit()
            messagebox.showinfo("Sucesso", "Receita salva com sucesso!", parent=self)
            self.limpar_campos()
            self.carregar_receitas()
        except Exception as e:
            messagebox.showerror("Erro no Banco de Dados", f"Erro ao salvar receita: {e}", parent=self)

    def selecionar_item(self, event):
        selected_item = self.tabela.focus()
        if not selected_item: return
        
        values = self.tabela.item(selected_item, 'values')
        self.id_receita.set(values[0])
        
        data_obj = datetime.strptime(values[1], '%d/%m/%Y')
        self.data_entry.set_date(data_obj)
        
        self.descricao.set(values[2])
        self.conta_receita.set(values[3])
        self.meio_recebimento.set(values[4])
        
        try:
            valor_str = values[5].replace('R$', '').replace('.', '').replace(',', '.').strip()
            self.valor.set(float(valor_str))
        except (ValueError, IndexError):
             self.valor.set(0.0)

        self.btn_atualizar.config(state="normal")
        self.btn_excluir.config(state="normal")

    def atualizar_receita(self):
        receita_id = self.id_receita.get()
        if not receita_id: return
        
        if not self.descricao.get() or self.valor.get() <= 0 or not self.conta_receita.get():
            messagebox.showerror("Erro de Validação", "Preencha todos os campos obrigatórios.", parent=self)
            return
            
        try:
            data_recebimento = self.data_entry.get_date().strftime('%Y-%m-%d')
            self.cursor_receitas.execute("""
                UPDATE receitas SET descricao=?, conta_receita=?, meio_recebimento=?, valor=?, data_recebimento=?
                WHERE id=?
            """, (self.descricao.get(), self.conta_receita.get(), self.meio_recebimento.get(), self.valor.get(), data_recebimento, receita_id))
            self.conn_receitas.commit()
            messagebox.showinfo("Sucesso", "Receita atualizada com sucesso!", parent=self)
            self.limpar_campos()
            self.carregar_receitas()
        except Exception as e:
            messagebox.showerror("Erro no Banco de Dados", f"Erro ao atualizar receita: {e}", parent=self)

    def excluir_receita(self):
        receita_id = self.id_receita.get()
        if not receita_id: return

        if messagebox.askyesno("Confirmar", "Tem certeza que deseja excluir esta receita?", parent=self):
            try:
                self.cursor_receitas.execute("DELETE FROM receitas WHERE id=?", (receita_id,))
                self.conn_receitas.commit()
                messagebox.showinfo("Sucesso", "Receita excluída com sucesso!", parent=self)
                self.limpar_campos()
                self.carregar_receitas()
            except Exception as e:
                messagebox.showerror("Erro no Banco de Dados", f"Erro ao excluir receita: {e}", parent=self)


# --- CLASSE PRINCIPAL DA APLICAÇÃO ---
class SistemaFinanceiro:
    def __init__(self, root):
        self.root = root
        self.root.title("Sistema de Gerenciamento Financeiro")
        self.root.geometry("1270x750")
        self.root.resizable(True, True)
        
        self.menu_bar = tk.Menu(self.root)
        self.root.config(menu=self.menu_bar)

        arquivo_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.menu_bar.add_cascade(label="Arquivo", menu=arquivo_menu)
        arquivo_menu.add_command(label="Relatório Entre Datas", command=self.gerar_relatorio_entre_datas)
        arquivo_menu.add_command(label="Relatório Mensal por Período", command=self.gerar_relatorio_mensal_periodo)
        arquivo_menu.add_command(label="Importar de Planilha Excel...", command=self.abrir_importador_excel)
        arquivo_menu.add_command(label="Importar do Supabase...", command=self.abrir_importador_supabase) # <--- ADICIONE ESTA LINHA
        arquivo_menu.add_separator()
        arquivo_menu.add_command(label="Sair", command=self.on_closing)

        relatorios_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.menu_bar.add_cascade(label="Relatórios de Despesas", menu=relatorios_menu)
        relatorios_menu.add_command(label="Top 10 Contas (Gráfico Dinâmico)", command=self.mostrar_grafico_principais_contas)
        relatorios_menu.add_command(label="Por Categoria (Evolução)", command=self.gerar_relatorio_categoria)
        relatorios_menu.add_command(label="Por Meio de Pagamento (Evolução)", command=self.gerar_relatorio_meio_pagamento)
        relatorios_menu.add_separator()
        relatorios_menu.add_command(label="Relatórios Avançados com Gráficos", command=self.abrir_relatorios_avancados)
        relatorios_menu.add_command(label="Relatórios Simples (Listagem)", command=self.abrir_relatorios_simples)
        relatorios_menu.add_command(label="Mensal: Orçado vs. Gasto", command=self.abrir_relatorio_orcado_vs_gasto)
        relatorios_menu.add_command(label="Previsão de Faturas de Cartão", command=self.abrir_relatorio_previsao_faturas_cartao)

        config_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.menu_bar.add_cascade(label="Configuração", menu=config_menu)
        config_menu.add_command(label="Gerenciar Categorias e Pagamentos", command=self.abrir_gerenciador)
        config_menu.add_command(label="Gerenciar Banco de Dados", command=self.abrir_gerenciador2)
        config_menu.add_command(label="Atualizar Interface", command=self.atualizar_dados_interface)
        config_menu.add_command(label="Gerenciar Orçamento", command=self.abrir_gerenciador_orcamento) 
        config_menu.add_command(label="Configurar Fechamento de Cartões", command=self.abrir_gerenciador_fechamento_cartoes)

        receitas_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.menu_bar.add_cascade(label="Receitas", menu=receitas_menu)
        receitas_menu.add_command(label="Lançar/Gerenciar Receitas", command=self.abrir_gerenciador_receitas)
        receitas_menu.add_command(label="Gerenciar Categorias de Receita", command=self.abrir_gerenciador_categorias_receita)
        receitas_menu.add_separator()
        receitas_menu.add_command(label="Relatório Mensal de Receitas", command=self.gerar_relatorio_receita_mensal)
        receitas_menu.add_command(label="Relatório de Receitas por Categoria", command=self.gerar_relatorio_receita_categoria)
        receitas_menu.add_command(label="Relatório de Receitas Entre Datas", command=self.gerar_relatorio_receita_entre_datas)

        # ==================================================================
        # ADICIONE ESTE BLOCO DE CÓDIGO AQUI
        # Novo Menu "Relatórios Consolidados"
        consolidado_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.menu_bar.add_cascade(label="Relatórios Consolidados", menu=consolidado_menu)
        consolidado_menu.add_command(label="Balanço Mensal (Receita x Despesa)", command=self.abrir_relatorio_balanco)
        consolidado_menu.add_command(label="Balanço Mensal (Fluxo de Caixa)", command=self.abrir_relatorio_balancofc)

       

        ajuda_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.menu_bar.add_cascade(label="Ajuda", menu=ajuda_menu)
        ajuda_menu.add_command(label="Ajuda", command=self.mostrar_ajuda)
        ajuda_menu.add_command(label="Sobre...", command=self.mostrar_sobre)
        
        self.cor_primaria = "#4CAF50"
        self.cor_secundaria = "#F0F4C3"
        self.cor_fundo = "#F9F9F9"
        self.cor_texto = "#333333"
        
        self.configurar_estilo()
        
        self.criar_banco_dados()
        self.criar_banco_dados_receitas()
        
        self.id_despesa = tk.StringVar()
        self.descricao = tk.StringVar()
        self.meio_pagamento = tk.StringVar()
        self.conta_despesa = tk.StringVar()
        self.valor = tk.DoubleVar()
        self.num_parcelas = tk.IntVar()
        
        # INÍCIO DA MODIFICAÇÃO: Adiciona variável para mapear meses do gráfico
        self.mapa_meses_grafico = {}
        self.total_mes_selecionado_var = tk.StringVar()
        # FIM DA MODIFICAÇÃO
        
        self.criar_widgets()
        self.carregar_despesas()

    def on_closing(self):
        if messagebox.askokcancel("Sair", "Deseja sair do sistema?"):
            self.conn.close()
            self.conn_receitas.close()
            self.root.destroy()
            
   
    def abrir_importador_supabase(self):
        """Abre a janela para importar despesas do Supabase."""
        # The SupabaseImporter needs access to the main app's SQLite connection and cursor
        importador_supabase.iniciar_importador_supabase(self.root, self)
        # After closing the importer, refresh the main display to show new expenses
        self.carregar_despesas()

    
    
    def criar_banco_dados_receitas1(self):
        try:
            self.conn_receitas = sqlite3.connect('financas_receitas.db')
            self.cursor_receitas = self.conn_receitas.cursor()
            
            self.cursor_receitas.execute('''
                CREATE TABLE IF NOT EXISTS receitas (
                    id INTEGER PRIMARY KEY, descricao TEXT NOT NULL, meio_recebimento TEXT,
                    conta_receita TEXT NOT NULL, valor REAL NOT NULL, num_parcelas INTEGER DEFAULT 1,
                    data_registro DATE NOT NULL, data_recebimento DATE NOT NULL)''')
            self.cursor_receitas.execute('CREATE TABLE IF NOT EXISTS categorias_receita (id INTEGER PRIMARY KEY, nome TEXT NOT NULL UNIQUE)')
            self.cursor_receitas.execute('CREATE TABLE IF NOT EXISTS meios_recebimento (id INTEGER PRIMARY KEY, nome TEXT NOT NULL UNIQUE)')

            cat_receita_padrao = [('Salário',), ('Vendas',), ('Rendimentos',), ('Freelance',), ('Outras Receitas',)]
            meios_receita_padrao = [('Transferência Bancária',), ('PIX',), ('Dinheiro',), ('Cheque',)]

            self.cursor_receitas.executemany("INSERT OR IGNORE INTO categorias_receita (nome) VALUES (?)", cat_receita_padrao)
            self.cursor_receitas.executemany("INSERT OR IGNORE INTO meios_recebimento (nome) VALUES (?)", meios_receita_padrao)

            self.conn_receitas.commit()
        except sqlite3.Error as e:
            messagebox.showerror("Erro no BD de Receitas", f"Ocorreu um erro: {e}")
#######################################

    def criar_banco_dados_receitas(self):
        """Cria o banco de dados de receitas e insere os dados padrão APENAS se as tabelas estiverem vazias."""
        try:
            self.conn_receitas = sqlite3.connect('financas_receitas.db')
            self.cursor_receitas = self.conn_receitas.cursor()
            
            # --- ETAPA 1: Criar as tabelas (o IF NOT EXISTS garante que não haverá erro) ---
            self.cursor_receitas.execute('''
                CREATE TABLE IF NOT EXISTS receitas (
                    id INTEGER PRIMARY KEY, descricao TEXT NOT NULL, meio_recebimento TEXT,
                    conta_receita TEXT NOT NULL, valor REAL NOT NULL, num_parcelas INTEGER DEFAULT 1,
                    data_registro DATE NOT NULL, data_recebimento DATE NOT NULL)''')
            
            self.cursor_receitas.execute('''
                CREATE TABLE IF NOT EXISTS categorias_receita (
                    id INTEGER PRIMARY KEY, 
                    nome TEXT NOT NULL UNIQUE)''')

            self.cursor_receitas.execute('''
                CREATE TABLE IF NOT EXISTS meios_recebimento (
                    id INTEGER PRIMARY KEY, 
                    nome TEXT NOT NULL UNIQUE)''')

            # --- ETAPA 2: Verificar se as tabelas de configuração estão vazias antes de inserir os padrões ---

            # Verificar e popular 'categorias_receita'
            self.cursor_receitas.execute("SELECT COUNT(id) FROM categorias_receita")
            if self.cursor_receitas.fetchone()[0] == 0:
                # A tabela está vazia, então inserimos os padrões.
                cat_receita_padrao = [
                    ('Salário',), ('Vendas',), ('Rendimentos',), ('Freelance',), ('Outras Receitas',)
                ]
                self.cursor_receitas.executemany("INSERT INTO categorias_receita (nome) VALUES (?)", cat_receita_padrao)

            # Verificar e popular 'meios_recebimento'
            self.cursor_receitas.execute("SELECT COUNT(id) FROM meios_recebimento")
            if self.cursor_receitas.fetchone()[0] == 0:
                # A tabela está vazia, então inserimos os padrões.
                meios_receita_padrao = [
                    ('Transferência Bancária',), ('PIX',), ('Dinheiro',), ('Cheque',)
                ]
                self.cursor_receitas.executemany("INSERT INTO meios_recebimento (nome) VALUES (?)", meios_receita_padrao)

            self.conn_receitas.commit()

        except sqlite3.Error as e:
            messagebox.showerror("Erro no BD de Receitas", f"Ocorreu um erro na inicialização do BD de Receitas: {e}")




######################################
    # ===== FUNÇÕES DE RECEITAS =====
    def abrir_gerenciador_receitas(self):
        GerenciadorReceitas(self.root, self)

    # ==================================================================
    # ADICIONE ESTA FUNÇÃO AQUI
    def abrir_relatorio_balanco(self):
        """Abre a janela do relatório de balanço Receita x Despesa."""
        relatorio_balanco.iniciar_relatorio_balanco(self.root)
    # ==================================================================

# ==================================================================
    # ADICIONE ESTA FUNÇÃO AQUI
    def abrir_relatorio_balancofc(self):
        """Abre a janela do relatório de balanço Receita x Despesa."""
        relatorio_balanco_fluxo_caixa.iniciar_relatorio_balanco(self.root)
    # ==================================================================

    def abrir_gerenciador_categorias_receita(self):
        GerenciadorCategoriasReceita(self.root, self)

    def _criar_dialogo_datas(self, titulo, callback_gerar):
        dialog = tk.Toplevel(self.root)
        dialog.title(titulo)
        dialog.geometry("400x200")
        dialog.transient(self.root)
        dialog.grab_set()

        frame = ttk.Frame(dialog, padding="10")
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frame, text="Data Inicial:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        data_inicial_entry = DateEntry(frame, width=12, date_pattern='dd/mm/yyyy', locale='pt_BR')
        data_inicial_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        
        ttk.Label(frame, text="Data Final:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        data_final_entry = DateEntry(frame, width=12, date_pattern='dd/mm/yyyy', locale='pt_BR')
        data_final_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")
        data_final_entry.set_date(datetime.now())

        def gerar():
            data_inicial = data_inicial_entry.get_date()
            data_final = data_final_entry.get_date()
            if data_inicial > data_final:
                messagebox.showerror("Erro", "A data inicial não pode ser posterior à data final.", parent=dialog)
                return
            dialog.destroy()
            callback_gerar(data_inicial.strftime("%Y-%m-%d"), data_final.strftime("%Y-%m-%d"))

        ttk.Button(frame, text="Gerar Relatório", command=gerar).grid(row=2, column=0, columnspan=2, pady=20)

    def gerar_relatorio_receita_entre_datas(self):
        self._criar_dialogo_datas("Relatório de Receitas por Período", 
                                  lambda d1, d2: self.mostrar_relatorio_receita(d1, d2, tipo='periodo'))
                                  
    def gerar_relatorio_receita_mensal(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("Relatório Mensal de Receitas")
        dialog.geometry("300x200")
        dialog.transient(self.root)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="10")
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frame, text="Mês:").grid(row=0, column=0, padx=5, pady=5)
        mes_combo = ttk.Combobox(frame, values=[calendar.month_name[i] for i in range(1, 13)], state="readonly", width=15)
        mes_combo.current(datetime.now().month - 1)
        mes_combo.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(frame, text="Ano:").grid(row=1, column=0, padx=5, pady=5)
        ano_var = tk.IntVar(value=datetime.now().year)
        ttk.Spinbox(frame, from_=2000, to=2100, textvariable=ano_var, width=10).grid(row=1, column=1, padx=5, pady=5)

        def gerar():
            mes = mes_combo.current() + 1
            ano = ano_var.get()
            dialog.destroy()
            self.mostrar_relatorio_receita(mes, ano, tipo='mensal')

        ttk.Button(frame, text="Gerar Relatório", command=gerar).grid(row=2, column=0, columnspan=2, pady=20)

    def gerar_relatorio_receita_categoria(self):
        self.cursor_receitas.execute("SELECT nome FROM categorias_receita ORDER BY nome")
        categorias = [row[0] for row in self.cursor_receitas.fetchall()]
        if not categorias:
            messagebox.showinfo("Informação", "Nenhuma categoria de receita cadastrada.", parent=self.root)
            return
            
        dialog = tk.Toplevel(self.root)
        dialog.title("Relatório por Categoria de Receita")
        dialog.geometry("350x150")
        dialog.transient(self.root)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="10")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Selecione a Categoria:").pack(pady=5)
        cat_combo = ttk.Combobox(frame, values=categorias, state="readonly")
        cat_combo.pack(pady=5, fill=tk.X, expand=True)
        cat_combo.set(categorias[0])

        def gerar():
            categoria = cat_combo.get()
            dialog.destroy()
            self.mostrar_relatorio_receita(categoria, tipo='categoria')

        ttk.Button(frame, text="Gerar Relatório", command=gerar).pack(pady=10)

    def mostrar_relatorio_receita(self, *args, tipo):
        janela_relatorio = tk.Toplevel(self.root)
        janela_relatorio.geometry("900x600")
        frame_principal = ttk.Frame(janela_relatorio, padding="10")
        frame_principal.pack(fill=tk.BOTH, expand=True)

        query, params, titulo_relatorio = "", (), "Relatório de Receitas"
        col_headers, chart_type = ("Item", "Total (R$)"), "pie"

        if tipo == 'mensal':
            mes, ano = args
            titulo_relatorio = f"Receitas de {calendar.month_name[mes]}/{ano}"
            query = "SELECT conta_receita, SUM(valor) FROM receitas WHERE strftime('%Y-%m', data_recebimento) = ? GROUP BY conta_receita ORDER BY SUM(valor) DESC"
            params = (f"{ano}-{mes:02d}",)
        elif tipo == 'categoria':
            categoria = args[0]
            titulo_relatorio = f"Evolução da Receita: {categoria}"
            query = "SELECT strftime('%m/%Y', data_recebimento) as mes, SUM(valor) FROM receitas WHERE conta_receita = ? GROUP BY strftime('%Y-%m', data_recebimento) ORDER BY strftime('%Y-%m', data_recebimento) ASC"
            params = (categoria,)
            col_headers, chart_type = ("Mês/Ano", "Total (R$)"), "line"
        elif tipo == 'periodo':
            d1, d2 = args
            d1_fmt = datetime.strptime(d1, "%Y-%m-%d").strftime("%d/%m/%Y")
            d2_fmt = datetime.strptime(d2, "%Y-%m-%d").strftime("%d/%m/%Y")
            titulo_relatorio = f"Receitas de {d1_fmt} a {d2_fmt}"
            query = "SELECT conta_receita, SUM(valor) FROM receitas WHERE data_recebimento BETWEEN ? AND ? GROUP BY conta_receita ORDER BY SUM(valor) DESC"
            params = (d1, d2)
        
        janela_relatorio.title(titulo_relatorio)
        ttk.Label(frame_principal, text=titulo_relatorio, font=('Arial', 16, 'bold')).pack(pady=10)
        
        self.cursor_receitas.execute(query, params)
        resultados = self.cursor_receitas.fetchall()

        if not resultados:
            messagebox.showinfo("Sem Dados", "Nenhum dado encontrado.", parent=janela_relatorio)
            janela_relatorio.destroy()
            return
            
        notebook = ttk.Notebook(frame_principal)
        notebook.pack(fill=tk.BOTH, expand=True, pady=10)
        tab_tabela = ttk.Frame(notebook)
        tab_grafico = ttk.Frame(notebook)
        notebook.add(tab_tabela, text="Tabela")
        notebook.add(tab_grafico, text="Gráfico")

        tree = ttk.Treeview(tab_tabela, columns=('col1', 'col2'), show='headings')
        tree.heading('col1', text=col_headers[0])
        tree.heading('col2', text=col_headers[1])
        tree.column('col2', anchor=tk.E)
        total_geral = 0
        for item, valor in resultados:
            tree.insert("", "end", values=(item, locale.currency(valor, grouping=True)))
            total_geral += valor
        tree.pack(fill=tk.BOTH, expand=True)
        ttk.Label(tab_tabela, text=f"Total: {locale.currency(total_geral, grouping=True)}", font=("Arial", 12, "bold")).pack(pady=5, anchor="e")

        fig = plt.Figure(figsize=(7, 5), dpi=100)
        ax = fig.add_subplot(111)
        labels = [row[0] for row in resultados]
        sizes = [row[1] for row in resultados]
        
        if chart_type == 'pie':
            ax.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90, pctdistance=0.85)
            ax.axis('equal')
        elif chart_type == 'line':
            ax.plot(labels, sizes, marker='o', linestyle='-', color=self.cor_primaria)
            ax.set_ylabel("Valor (R$)")
            ax.grid(True, linestyle='--', alpha=0.7)
            plt.setp(ax.get_xticklabels(), rotation=45, ha='right')

        canvas = FigureCanvasTkAgg(fig, master=tab_grafico)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        fig.tight_layout()

    # ===== FUNÇÕES DE DESPESA (RESTANTES DO ARQUIVO ORIGINAL) =====
    
    def atualizar_dados_interface(self):
        """
        Recarrega os dados do banco de dados e atualiza os componentes
        visuais da interface principal (tabela e gráfico resumo).
        """
        try:
            self.carregar_despesas()  # Esta função já atualiza a tabela e o gráfico principal
            messagebox.showinfo("Atualização Concluída", 
                                "Os dados da interface foram atualizados com sucesso!", 
                                parent=self.root)
        except Exception as e:
            messagebox.showerror("Erro na Atualização", 
                                 f"Ocorreu um erro ao tentar atualizar os dados da interface:\n{e}", 
                                 parent=self.root)
    
    
    def abrir_relatorio_previsao_faturas_cartao(self):
        """Abre a janela de relatório de Previsão de Faturas de Cartão."""
        relatorio_previsao_faturas.iniciar_relatorio_previsao_faturas(self.root)
   
    # Coloque esta função dentro da classe SistemaFinanceiro
    def abrir_importador_excel(self):
        """Abre a janela para importar despesas de uma planilha Excel."""
        importador_excel.iniciar_importador_excel(self.root)
        # Após fechar o importador, recarregue as despesas na tela principal
        self.carregar_despesas()
        
    def abrir_gerenciador_fechamento_cartoes(self):
        """Abre a janela para gerenciar as datas de fechamento de faturas de cartões."""
        gerenciar_fechamento_cartoes.iniciar_gerenciador_fechamento_cartoes(self.root)
    
    def abrir_relatorio_orcado_vs_gasto(self):
        """Abre a janela de relatório Orçado vs. Gasto."""
        relatorio_orcado_vs_gasto.iniciar_relatorio_orcado_vs_gasto(self.root)   # Coloque esta função dentro da classe SistemaFinanceiro
        
    def abrir_gerenciador_orcamento(self):
        """Abre a janela para gerenciar o orçamento por categoria."""
        gerenciar_orcamento.iniciar_gerenciador_orcamento(self.root)
    
    def abrir_relatorios_simples(self):
        """Chama a janela de relatórios do módulo relatorios.py."""
        relatorios1.iniciar_relatorios(self.root)
        
        # Coloque esta função dentro da classe SistemaFinanceiro
    def abrir_relatorios_avancados(self):
        """Chama a janela de relatórios avançados do módulo relclaude."""
        # A função iniciar_relatorios_avancados espera a janela principal (self.root) como argumento
        relclaude1.iniciar_relatorios_avancados(self.root)
        
    # Coloque estas funções dentro da classe SistemaFinanceiro

    def abrir_relatorios_simples(self):
        """Chama a janela de relatórios simples do módulo relatorios.py."""
        relatorios1.iniciar_relatorios(self.root)

    def mostrar_sobre(self):
        """Exibe uma janela de informações sobre o programa."""
        messagebox.showinfo(
            "Sobre o Sistema de Gestão Financeira",
            "Sistema de Gerenciamento Financeiro\n\n"
            "Versão: 3.0\n"
            "Desenvolvido para auxiliar no controle de despesas pessoais.\n\n"
            "Autor: Seu Nome Aqui"
        )

    def mostrar_ajuda(self):
        """Exibe uma mensagem de ajuda básica."""
        messagebox.showinfo(
            "Ajuda",
            "Para cadastrar uma nova despesa, preencha os campos e clique em 'Salvar'.\n\n"
            "Para editar ou excluir, clique em um registro na tabela para carregar seus dados no formulário e depois use os botões 'Atualizar' ou 'Excluir'.\n\n"
            "Use os menus para acessar relatórios e configurações."
        )    
     
    
    
    def mostrar_grafico_principais_contas(self):
        """
        Exibe um gráfico das 10 principais contas de despesa, com cores e interatividade.
        Permite selecionar diferentes meses e clicar nas barras para ver detalhes.
        """
        try:
            # Importações necessárias (caso não estejam no topo do arquivo)
            from matplotlib.figure import Figure
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            from matplotlib.ticker import FuncFormatter
            import matplotlib.pyplot as plt
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            from openpyxl.utils import get_column_letter
            from tkinter import filedialog
            
            # Obter mês atual como padrão
            data_atual = datetime.now()
            mes_atual_ym = data_atual.strftime("%Y-%m")
            mes_atual_formatado = data_atual.strftime("%m/%Y")

            # Criar janela para o gráfico
            janela_grafico = tk.Toplevel(self.root)
            janela_grafico.title(f"Top 10 Contas - {mes_atual_formatado}")
            janela_grafico.geometry("900x800")
            
            # Frame principal
            frame_principal = ttk.Frame(janela_grafico, padding="10")
            frame_principal.pack(fill=tk.BOTH, expand=True)

            # --- Função aninhada para mostrar os detalhes da transação ---
            def mostrar_detalhes_transacao(categoria, mes_ym_selecionado):
                try:
                    # Criar a janela de detalhes
                    detalhes_window = tk.Toplevel(janela_grafico)
                    mes_formatado_titulo = datetime.strptime(mes_ym_selecionado, "%Y-%m").strftime("%m/%Y")
                    detalhes_window.title(f"Detalhes para '{categoria}' em {mes_formatado_titulo}")
                    detalhes_window.geometry("600x400")
                    detalhes_window.transient(janela_grafico)
                    detalhes_window.grab_set()

                    # Frame para o Treeview
                    tree_frame = ttk.Frame(detalhes_window, padding="10")
                    tree_frame.pack(fill=tk.BOTH, expand=True)

                    # Consultar as transações
                    self.cursor.execute("""
                        SELECT strftime('%d/%m/%Y', data_pagamento), descricao, valor
                        FROM despesas
                        WHERE conta_despesa = ? AND strftime('%Y-%m', data_pagamento) = ?
                        ORDER BY data_pagamento
                    """, (categoria, mes_ym_selecionado))
                    transacoes = self.cursor.fetchall()

                    # Treeview para exibir as transações
                    cols = ('data', 'descricao', 'valor')
                    tree = ttk.Treeview(tree_frame, columns=cols, show='headings')
                    tree.heading('data', text='Data')
                    tree.heading('descricao', text='Descrição')
                    tree.heading('valor', text='Valor (R$)')
                    tree.column('data', width=100, anchor=tk.CENTER)
                    tree.column('descricao', width=350)
                    tree.column('valor', width=120, anchor=tk.E)
                    
                    # Adicionar scrollbar
                    scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
                    tree.configure(yscrollcommand=scrollbar.set)
                    scrollbar.pack(side='right', fill='y')
                    tree.pack(side='left', fill='both', expand=True)


                    total_categoria = 0
                    for data, desc, valor in transacoes:
                        valor_formatado = locale.currency(valor, grouping=True)
                        tree.insert("", "end", values=(data, desc, valor_formatado))
                        total_categoria += valor
                    
                    # Adicionar um rótulo com o total
                    total_formatado = locale.currency(total_categoria, grouping=True)
                    ttk.Label(detalhes_window, text=f"Total: {total_formatado}", font=('Arial', 12, 'bold')).pack(pady=5)

                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao buscar detalhes: {e}", parent=janela_grafico)

            # --- Função de evento para cliques no gráfico ---
            def on_pick(event):
                # O nome da categoria e o mês estão armazenados no GID (Group ID) do artista
                gid = event.artist.get_gid()
                if gid:
                    categoria, mes_selecionado = gid
                    mostrar_detalhes_transacao(categoria, mes_selecionado)

            # --- Configuração da GUI para seleção de mês ---
            frame_selecao = ttk.Frame(frame_principal)
            frame_selecao.pack(fill=tk.X, pady=10)
            
            def obter_meses_disponiveis():
                try:
                    self.cursor.execute("""
                        SELECT DISTINCT strftime('%Y-%m', data_pagamento) as ano_mes,
                                        strftime('%m/%Y', data_pagamento) as mes_formatado
                        FROM despesas ORDER BY ano_mes DESC LIMIT 24
                    """)
                    return self.cursor.fetchall()
                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao obter meses: {e}")
                    return []
            
            meses_disponiveis = obter_meses_disponiveis()
            if not meses_disponiveis:
                meses_disponiveis = [(mes_atual_ym,  mes_atual_formatado)]
            
            mapa_meses = {mes_fmt: mes_ym for mes_ym, mes_fmt in meses_disponiveis}
            valores_meses_formatados = [mes_fmt for _, mes_fmt in meses_disponiveis]
            
            ttk.Label(frame_selecao, text="Selecione o mês:").pack(side=tk.LEFT, padx=5)
            mes_combo = ttk.Combobox(frame_selecao, state="readonly", width=10)
            mes_combo['values'] = valores_meses_formatados
            mes_combo.pack(side=tk.LEFT, padx=5)


            # --- INÍCIO DA ALTERAÇÃO: Lógica para selecionar o mês inicial ---
            # Tenta definir o mês atual como o padrão na combobox
            if mes_atual_formatado in valores_meses_formatados:
                mes_combo.set(mes_atual_formatado)
            elif valores_meses_formatados:
                # Se o mês atual não estiver na lista (sem dados), usa o mais recente
                mes_combo.current(0)
            # --- FIM DA ALTERAÇÃO ---



            # --- Configuração da Figura e Canvas do Matplotlib ---
            frame_grafico = ttk.Frame(frame_principal)
            frame_grafico.pack(fill=tk.BOTH, expand=True, pady=10)
            
            figura = Figure(figsize=(9, 6), dpi=100)
            ax = figura.add_subplot(111)
            
            canvas = FigureCanvasTkAgg(figura, frame_grafico)
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
            # Conectar o evento de clique ao manipulador
            canvas.mpl_connect('pick_event', on_pick)

            # --- GUI para informações adicionais (Total, Percentual) ---
            frame_info = ttk.Frame(frame_principal)
            frame_info.pack(fill=tk.X, pady=10)
            label_total = ttk.Label(frame_info, text="", font=('Arial', 12))
            label_total.pack(side=tk.LEFT, padx=10)
            label_percentual = ttk.Label(frame_info, text="", font=('Arial', 12))
            label_percentual.pack(side=tk.RIGHT, padx=10)

       
            # --- Função para atualizar o gráfico ---
            def atualizar_grafico(mes_selecionado):
                try:
                    # AQUI ESTÁ A CORREÇÃO: Usando a variável correta "mes_atual_formatado"
                    mes_exibicao = next((m_fmt for m_fmt, m_val in mapa_meses.items() if m_val == mes_selecionado), mes_atual_formatado)
                    
                    ax.clear()
                    
                    self.cursor.execute("""
                        SELECT conta_despesa, SUM(valor) as total_valor
                        FROM despesas WHERE strftime('%Y-%m', data_pagamento) = ?
                        GROUP BY conta_despesa ORDER BY total_valor DESC LIMIT 10
                    """, (mes_selecionado,))
                    resultados = self.cursor.fetchall()

                    if not resultados:
                        ax.text(0.5, 0.5, f"Nenhuma despesa encontrada para {mes_exibicao}", ha='center', va='center', fontsize=14)
                        janela_grafico.title(f"Top 10 Contas - {mes_exibicao}")
                        ax.set_title(f'Top 10 Contas de Despesa - {mes_exibicao}')
                        canvas.draw()
                        label_total.config(text="Total Top 10: R$ 0,00")
                        label_percentual.config(text="Representa 0.0% do total do mês")
                        return

                    self.cursor.execute("SELECT SUM(valor) FROM despesas WHERE strftime('%Y-%m', data_pagamento) = ?", (mes_selecionado,))
                    total_mes = self.cursor.fetchone()[0] or 0
                    
                    contas = [row[0] for row in resultados]
                    valores = [row[1] for row in resultados]
                    
                    total_top10 = sum(valores)
                    percentual_top10 = (total_top10 / total_mes * 100) if total_mes > 0 else 0
                    
                    # Definir cores diferentes para as barras
                    cores = plt.cm.get_cmap('tab10', 10).colors

                    # Criar barras horizontais e torná-las clicáveis
                    barras = ax.barh(contas[::-1], valores[::-1], color=cores, picker=True)

                    # Armazenar informações de categoria e mês em cada barra para o clique
                    for i, barra in enumerate(barras):
                        categoria = contas[::-1][i]
                        barra.set_gid((categoria, mes_selecionado))

                    # Adicionar rótulos de valor
                    for barra in barras:
                        largura = barra.get_width()
                        ax.text(largura * 1.01, barra.get_y() + barra.get_height()/2, 
                                locale.currency(largura, grouping=True),
                                va='center', ha='left', fontsize=9)
                    
                    # Configurar estilo do gráfico
                    ax.set_title(f'Top 10 Contas de Despesa - {mes_exibicao}')
                    ax.set_xlabel('Valor Total (R$)')
                    ax.set_ylabel('Conta de Despesa')
                    ax.grid(True, linestyle='--', alpha=0.7, axis='x')
                    ax.set_xlim(right=max(valores) * 1.18)  # Ajustar limite do eixo x para os rótulos

                    # Formatar o eixo x como moeda
                    def formatar_reais(x, pos):
                        return locale.currency(x, symbol=False, grouping=True)
                    ax.xaxis.set_major_formatter(FuncFormatter(formatar_reais))
                    
                    figura.tight_layout()
                    canvas.draw()
                    
                    # Atualizar informações
                    label_total.config(text=f"Total Top 10: {locale.currency(total_top10, grouping=True)}")
                    label_percentual.config(text=f"Representa {percentual_top10:.1f}% do total do mês")
                
                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao atualizar gráfico: {e}")
                    import traceback
                    traceback.print_exc()
                        
           ####################################################################
           
            
            # --- Manipulador de evento para mudança de mês ---
            def on_mes_change(event):
                try:
                    mes_selecionado_fmt = mes_combo.get()
                    mes_valor_ym = mapa_meses.get(mes_selecionado_fmt)
                    if not mes_valor_ym:
                        messagebox.showerror("Erro", "Mês inválido selecionado")
                        return
                    atualizar_grafico(mes_valor_ym)
                    janela_grafico.title(f"Top 10 Contas - {mes_selecionado_fmt}")
                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao mudar mês: {e}")
            
            mes_combo.bind("<<ComboboxSelected>>", on_mes_change)
            
            # --- Lógica para o botão de exportar ---
            def exportar_dados():
                try:
                    mes_selecionado = mes_combo.get()
                    mes_valor = mapa_meses.get(mes_selecionado)
                    if not mes_valor:
                        messagebox.showerror("Erro", "Mês inválido selecionado")
                        return
                    
                    arquivo = filedialog.asksaveasfilename(
                        defaultextension=".xlsx",
                        filetypes=[("Excel files", "*.xlsx")],
                        title="Exportar Top 10 Contas"
                    )
                    if not arquivo: return
                    
                    self.cursor.execute("""
                        SELECT conta_despesa, SUM(valor) as total_valor, COUNT(*) as total_registros
                        FROM despesas WHERE strftime('%Y-%m', data_pagamento) = ?
                        GROUP BY conta_despesa ORDER BY total_valor DESC
                    """, (mes_valor,))
                    dados = self.cursor.fetchall()
                    
                    self.cursor.execute("SELECT SUM(valor) FROM despesas WHERE strftime('%Y-%m', data_pagamento) = ?", (mes_valor,))
                    total_mes = self.cursor.fetchone()[0] or 0
                    
                    wb = Workbook()
                    ws = wb.active
                    titulo_planilha = mes_selecionado.replace("/", "-")
                    ws.title = f"Top Contas {titulo_planilha}"
                    
                    ws['A1'] = f"Relatório de Principais Contas - {mes_selecionado}"
                    ws.merge_cells('A1:D1')
                    ws['A1'].font = Font(size=14, bold=True)
                    
                    ws['A2'] = f"Total do Mês: {locale.currency(total_mes, grouping=True)}"
                    ws.merge_cells('A2:D2')
                    ws['A2'].font = Font(size=12)
                    
                    cabecalhos = ["Conta", "Total (R$)", "Qtd. Registros", "% do Mês"]
                    for col, cabecalho in enumerate(cabecalhos, start=1):
                        cell = ws.cell(row=4, column=col, value=cabecalho)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                    
                    for row_idx, (conta, valor, registros) in enumerate(dados, start=5):
                        percentual = (valor / total_mes) * 100 if total_mes > 0 else 0
                        ws.cell(row=row_idx, column=1, value=conta)
                        ws.cell(row=row_idx, column=2, value=valor)
                        ws.cell(row=row_idx, column=2).number_format = '"R$" #,##0.00'
                        ws.cell(row=row_idx, column=3, value=registros)
                        ws.cell(row=row_idx, column=4, value=f"{percentual:.1f}%")
                    
                    for col_letter in ['A', 'B', 'C', 'D']:
                        ws.column_dimensions[col_letter].width = 20
                    
                    wb.save(arquivo)
                    messagebox.showinfo("Exportar", f"Dados exportados com sucesso para {arquivo}")
                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao exportar dados: {e}")

            ttk.Button(frame_principal, text="Exportar para Excel", command=exportar_dados).pack(pady=10)

            # Gráfico inicial
             # --- INÍCIO DA ALTERAÇÃO: Carregamento do gráfico inicial ---
            # Pega o mês que foi selecionado na combobox e atualiza o gráfico
            mes_inicial_formatado = mes_combo.get()
            if mes_inicial_formatado:
                mes_inicial_ym = mapa_meses.get(mes_inicial_formatado)
                if mes_inicial_ym:
                    janela_grafico.title(f"Top 10 Contas - {mes_inicial_formatado}")
                    atualizar_grafico(mes_inicial_ym)
            # --- FIM DA ALTERAÇÃO ---
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exibir gráfico: {e}")
            import traceback
            traceback.print_exc()
        
    
    
   
       
    def configurar_estilo(self):
        """Configura o estilo visual do aplicativo"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Configurar estilo para botões
        style.configure('TButton', 
                        background=self.cor_primaria, 
                        foreground='white', 
                        font=('Arial', 10, 'bold'),
                        padding=5)
        
        # Configurar estilo para labels
        style.configure('TLabel', 
                        background=self.cor_fundo, 
                        foreground=self.cor_texto,
                        font=('Arial', 10))
        
        # Configurar estilo para entradas
        style.configure('TEntry', 
                        fieldbackground='white',
                        font=('Arial', 10))
        
        # Configurar estilo para frame
        style.configure('TFrame', background=self.cor_fundo)
        
        # Configurar estilo para treeview (tabela)
        style.configure('Treeview', 
                        background='white',
                        fieldbackground='white',
                        font=('Arial', 9))
        
        style.configure('Treeview.Heading', 
                        font=('Arial', 10, 'bold'),
                        background=self.cor_secundaria)
    
    def criar_banco_dados1(self):
        """Cria o banco de dados e a tabela se não existirem"""
        try:
            self.conn = sqlite3.connect('financas.db')
            self.cursor = self.conn.cursor()
            
            # Criar tabela principal de despesas
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS despesas (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    descricao TEXT NOT NULL,
                    meio_pagamento TEXT NOT NULL,
                    conta_despesa TEXT NOT NULL,
                    valor REAL NOT NULL,
                    num_parcelas INTEGER DEFAULT 1,
                    data_registro DATE NOT NULL,
                    data_pagamento DATE NOT NULL
                )
            ''')
            
            # Criar tabela de categorias de despesas (para futuras expansões)
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS categorias (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    nome TEXT NOT NULL UNIQUE
                )
            ''')
            
            # Inserir categorias padrão se não existirem
            categorias_padrao = [
                ('Tel. e Internet',), ('Gás',), ('Mercado',), ('Alimentação',), ('Moradia',), ('Transporte',), ('Educação',),
                ('Saúde',), ('Lazer',), ('Vestuário',), ('Funcionários',), ('Outros',)
            ]
            
            for categoria in categorias_padrao:
                try:
                    self.cursor.execute("INSERT INTO categorias (nome) VALUES (?)", categoria)
                except sqlite3.IntegrityError:
                    pass  # Categoria já existe
            
            # Criar tabela de meios de pagamento
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS meios_pagamento (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    nome TEXT NOT NULL UNIQUE
                )
            ''')
            
            # Inserir meios de pagamento padrão
            meios_padrao = [
                ('Dinheiro',), ('Cartão Unlimited',), ('Cartão C6',), ('Cartão Nubank',), ('Cartão BB',),
                ('Transferência',), ('PIX',), ('Boleto',)
            ]
            
            for meio in meios_padrao:
                try:
                    self.cursor.execute("INSERT INTO meios_pagamento (nome) VALUES (?)", meio)
                except sqlite3.IntegrityError:
                    pass  # Meio já existe
                    
            self.conn.commit()
            
        except sqlite3.Error as e:
            messagebox.showerror("Erro no Banco de Dados", f"Ocorreu um erro: {e}")
    #####################################
    
    # Em sistema_financeirov7.py, substitua a função inteira por esta:

    def criar_banco_dados(self):
        """Cria o banco de dados e as tabelas e insere os dados padrão APENAS se as tabelas estiverem vazias."""
        try:
            self.conn = sqlite3.connect('financas.db')
            self.cursor = self.conn.cursor()
            
            # --- ETAPA 1: Criar as tabelas (o IF NOT EXISTS garante que não haverá erro se já existirem) ---
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS despesas (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    descricao TEXT NOT NULL,
                    meio_pagamento TEXT NOT NULL,
                    conta_despesa TEXT NOT NULL,
                    valor REAL NOT NULL,
                    num_parcelas INTEGER DEFAULT 1,
                    data_registro DATE NOT NULL,
                    data_pagamento DATE NOT NULL
                )
            ''')
            
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS categorias (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    nome TEXT NOT NULL UNIQUE
                )
            ''')
            
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS meios_pagamento (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    nome TEXT NOT NULL UNIQUE
                )
            ''')
            
            # --- ETAPA 2: Verificar se as tabelas de configuração estão vazias antes de inserir ---

            # Verificar e popular 'meios_pagamento'
            self.cursor.execute("SELECT COUNT(id) FROM meios_pagamento")
            if self.cursor.fetchone()[0] == 0:
                # A tabela está vazia, então inserimos os padrões.
                meios_padrao = [
                    ('Dinheiro',), ('Cartão Unlimited',), ('Cartão C6',), ('Cartão Nubank',), ('Cartão BB',),
                    ('Transferência',), ('PIX',), ('Boleto',)
                ]
                self.cursor.executemany("INSERT INTO meios_pagamento (nome) VALUES (?)", meios_padrao)

            # Verificar e popular 'categorias'
            self.cursor.execute("SELECT COUNT(id) FROM categorias")
            if self.cursor.fetchone()[0] == 0:
                # A tabela está vazia, então inserimos os padrões.
                categorias_padrao = [
                    ('Tel. e Internet',), ('Gás',), ('Mercado',), ('Alimentação',), ('Moradia',), ('Transporte',), ('Educação',),
                    ('Saúde',), ('Lazer',), ('Vestuário',), ('Funcionários',), ('Outros',)
                ]
                self.cursor.executemany("INSERT INTO categorias (nome) VALUES (?)", categorias_padrao)

            self.conn.commit()
            
        except sqlite3.Error as e:
            messagebox.showerror("Erro no Banco de Dados", f"Ocorreu um erro na inicialização do BD: {e}")
        
        
        
        
    
    ####################################
    def carregar_categorias(self):
        """Carrega todas as categorias do banco de dados"""
        self.cursor.execute("SELECT nome FROM categorias ORDER BY nome")
        return [categoria[0] for categoria in self.cursor.fetchall()]
    
    def carregar_meios_pagamento(self):
        """Carrega todos os meios de pagamento do banco de dados"""
        self.cursor.execute("SELECT nome FROM meios_pagamento ORDER BY nome")
        return [meio[0] for meio in self.cursor.fetchall()]
    
    def criar_widgets(self):
        """Cria todos os elementos da interface gráfica"""
                
        
        # Frame principal dividido em duas partes
        self.frame_principal = ttk.Frame(self.root)
        self.frame_principal.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Frame lateral esquerdo (formulário)
        self.frame_form = ttk.Frame(self.frame_principal, style='TFrame')
        self.frame_form.pack(side=tk.LEFT, fill=tk.BOTH, padx=10, pady=10)
        
        # Frame direito (tabela e gráficos)
        self.frame_tabela = ttk.Frame(self.frame_principal, style='TFrame')
        self.frame_tabela.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Título
        titulo = ttk.Label(self.frame_form, text="Controle Financeiro", 
                          font=('Arial', 16, 'bold'), style='TLabel')
        titulo.grid(row=0, column=0, columnspan=2, pady=10)
        
   
        # Formulário de entrada
        # ID (Hidden)
        ttk.Label(self.frame_form, text="ID:").grid(row=1, column=0, sticky=tk.W, pady=5)
        id_entry = ttk.Entry(self.frame_form, textvariable=self.id_despesa, state='readonly', width=5)
        id_entry.grid(row=1, column=1, sticky=tk.W, pady=5)
        
        # Descrição
        ttk.Label(self.frame_form, text="Descrição:").grid(row=2, column=0, sticky=tk.W, pady=5)
        ttk.Entry(self.frame_form, textvariable=self.descricao, width=30).grid(row=2, column=1, sticky=tk.W, pady=5)
        
        # Meio de Pagamento (ComboBox)
        ttk.Label(self.frame_form, text="Meio de Pagamento:").grid(row=3, column=0, sticky=tk.W, pady=5)
        self.combo_meio_pagamento = ttk.Combobox(self.frame_form, textvariable=self.meio_pagamento, width=20)
        self.combo_meio_pagamento['values'] = self.carregar_meios_pagamento()
        self.combo_meio_pagamento.grid(row=3, column=1, sticky=tk.W, pady=5)
        
        # Conta de Despesa (ComboBox)
        ttk.Label(self.frame_form, text="Categoria:").grid(row=4, column=0, sticky=tk.W, pady=5)
        self.combo_conta_despesa = ttk.Combobox(self.frame_form, textvariable=self.conta_despesa, width=20)
        self.combo_conta_despesa['values'] = self.carregar_categorias()
        self.combo_conta_despesa.grid(row=4, column=1, sticky=tk.W, pady=5)
        
        # Valor
        ttk.Label(self.frame_form, text="Valor (R$):").grid(row=5, column=0, sticky=tk.W, pady=5)
        ttk.Entry(self.frame_form, textvariable=self.valor, width=15).grid(row=5, column=1, sticky=tk.W, pady=5)
        
        # Número de Parcelas
        ttk.Label(self.frame_form, text="Parcelas:").grid(row=6, column=0, sticky=tk.W, pady=5)
        ttk.Spinbox(self.frame_form, from_=1, to=48, textvariable=self.num_parcelas, width=5).grid(row=6, column=1, sticky=tk.W, pady=5)
        self.num_parcelas.set(1)  # Valor padrão
        
        # Data de Pagamento
      #  ttk.Label(self.frame_form, text="Data de Pagamento:").grid(row=7, column=0, sticky=tk.W, pady=5)
      #  self.data_entry = DateEntry(self.frame_form, width=12, background=self.cor_primaria,
      #                             foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
       # self.data_entry.grid(row=7, column=1, sticky=tk.W, pady=5)
      #  self.data_entry.set_date(datetime.now())
        
        self.data_entry = DateEntry(self.frame_form, width=12, background=self.cor_primaria,
                                   foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy',
                                   locale='pt_BR') # Padrão de data alterado para dd/mm/yyyy e locale pt_BR adicionado
                                   
        self.data_entry.grid(row=7, column=1, sticky=tk.W, pady=5)
        self.data_entry.set_date(datetime.now()) # Define a data atual ao iniciar
        
        
        # Frame dos Botões
        botoes_frame = ttk.Frame(self.frame_form)
        botoes_frame.grid(row=8, column=0, columnspan=2, pady=15)
        
        # Botões de ação
        self.btn_salvar = ttk.Button(botoes_frame, text="Salvar", command=self.salvar_despesa)
        self.btn_salvar.grid(row=0, column=0, padx=5)
        
        self.btn_atualizar = ttk.Button(botoes_frame, text="Atualizar", command=self.atualizar_despesa)
        self.btn_atualizar.grid(row=0, column=1, padx=5)
        self.btn_atualizar['state'] = 'disabled'
        
        self.btn_excluir = ttk.Button(botoes_frame, text="Excluir", command=self.excluir_despesa)
        self.btn_excluir.grid(row=0, column=2, padx=5)
        self.btn_excluir['state'] = 'disabled'
        
        self.btn_limpar = ttk.Button(botoes_frame, text="Limpar", command=self.limpar_campos)
        self.btn_limpar.grid(row=0, column=3, padx=5)

       # self.btn_conf = ttk.Button(botoes_frame, text="Configuração", command=self.abrir_gerenciador)
       # self.btn_conf.grid(row=0, column=4, padx=5)

                      
        # Frame de Pesquisa
        frame_pesquisa = ttk.LabelFrame(self.frame_form, text="Pesquisar")
        frame_pesquisa.grid(row=9, column=0, columnspan=2, sticky=tk.W+tk.E, pady=10)
        
        self.pesquisa_termo = tk.StringVar()
        ttk.Entry(frame_pesquisa, textvariable=self.pesquisa_termo, width=20).grid(row=0, column=0, padx=5, pady=5)
        ttk.Button(frame_pesquisa, text="Buscar", command=self.pesquisar_despesa).grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(frame_pesquisa, text="Mostrar Todos", command=self.carregar_despesas).grid(row=0, column=2, padx=5, pady=5)
        
        # Frame de Relatórios
        frame_relatorios = ttk.LabelFrame(self.frame_form, text="Relatórios")
        frame_relatorios.grid(row=10, column=0, columnspan=2, sticky=tk.W+tk.E, pady=10)
        
        ttk.Button(frame_relatorios, text="Relatório Mensal", 
                  command=self.gerar_relatorio_mensal).grid(row=0, column=0, padx=5, pady=5)
        ttk.Button(frame_relatorios, text="Relatório por Categoria", 
                  command=self.gerar_relatorio_categoria).grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(frame_relatorios, text="Exportar Excel", 
                  command=self.exportar_excel).grid(row=0, column=2, padx=5, pady=5)
       # ttk.Button(frame_relatorios, text="Relatório Meios de Pagamento", 
       #           command=self.gerar_relatorio_meios_pagamento).grid(row=0, column=3, padx=5, pady=5)
        
        # Tabela de despesas
        self.tree_frame = ttk.LabelFrame(self.frame_tabela, text="Registros de Despesas")
        self.tree_frame.pack(fill=tk.BOTH, expand=True)
        
        # Scrollbar para a tabela
        scrollbar = ttk.Scrollbar(self.tree_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Configurar a tabela (Treeview)
        self.tabela = ttk.Treeview(self.tree_frame, yscrollcommand=scrollbar.set)
        
        # Definir colunas
        self.tabela['columns'] = ('id', 'descricao', 'meio_pagamento', 'conta_despesa', 
                                 'valor', 'parcelas', 'data')
        
        # Formatar colunas
        self.tabela.column('#0', width=0, stretch=tk.NO)  # Coluna fantasma
        self.tabela.column('id', anchor=tk.CENTER, width=50)
        self.tabela.column('descricao', anchor=tk.W, width=200)
        self.tabela.column('meio_pagamento', anchor=tk.W, width=120)
        self.tabela.column('conta_despesa', anchor=tk.W, width=120)
        self.tabela.column('valor', anchor=tk.E, width=100)
        self.tabela.column('parcelas', anchor=tk.CENTER, width=70)
        self.tabela.column('data', anchor=tk.CENTER, width=100)
        
        # Definir cabeçalhos
        self.tabela.heading('#0', text='', anchor=tk.CENTER)
        self.tabela.heading('id', text='ID', anchor=tk.CENTER)
        self.tabela.heading('descricao', text='Descrição', anchor=tk.CENTER)
        self.tabela.heading('meio_pagamento', text='Meio Pagamento', anchor=tk.CENTER)
        self.tabela.heading('conta_despesa', text='Categoria', anchor=tk.CENTER)
        self.tabela.heading('valor', text='Valor (R$)', anchor=tk.CENTER)
        self.tabela.heading('parcelas', text='Parcelas', anchor=tk.CENTER)
        self.tabela.heading('data', text='Data', anchor=tk.CENTER)
        
        # Vincular evento de clique na tabela
        self.tabela.bind("<ButtonRelease-1>", self.selecionar_item)
        
        # Posicionar a tabela e conectar scrollbar
        self.tabela.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        scrollbar.config(command=self.tabela.yview)
        
        # Frame para gráficos
        # INÍCIO DA MODIFICAÇÃO: Título do LabelFrame é genérico agora
        self.frame_grafico = ttk.LabelFrame(self.frame_tabela, text="Resumo de Gastos")
        # FIM DA MODIFICAÇÃO
        self.frame_grafico.pack(fill=tk.BOTH, expand=True, pady=10)

        # INÍCIO DA MODIFICAÇÃO: Adiciona controles de seleção de mês
        frame_controles_grafico = ttk.Frame(self.frame_grafico)
        frame_controles_grafico.pack(fill=tk.X, padx=5, pady=2)
        
        ttk.Label(frame_controles_grafico, text="Mês:").pack(side=tk.LEFT, padx=(0, 5))
        self.combo_mes_grafico = ttk.Combobox(frame_controles_grafico, state="readonly", width=10)
        self.combo_mes_grafico.pack(side=tk.LEFT)
        self.combo_mes_grafico.bind("<<ComboboxSelected>>", self.atualizar_grafico)
        # FIM DA MODIFICAÇÃO
        
        # INÍCIO DA NOVA ALTERAÇÃO: Adiciona o Label para o total do mês
        label_total_mes = ttk.Label(frame_controles_grafico, textvariable=self.total_mes_selecionado_var, font=('Arial', 9, 'bold'))
        label_total_mes.pack(side=tk.LEFT, padx=(15, 0))
        # FIM DA NOVA ALTERAÇÃO


        # Criar gráfico inicial
        self.figura_grafico = plt.Figure(figsize=(5, 4), dpi=100)
        self.canvas_grafico = FigureCanvasTkAgg(self.figura_grafico, self.frame_grafico)
        self.canvas_grafico.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
        # Inicializar o gráfico
        self.atualizar_grafico()
    
    def abrir_gerenciador(self):
        """Abre o programa 2 dentro de uma nova janela Tkinter."""
        # opcionalmente pode usar Toplevel() se quiser que compartilhe a mesma root
        # # def abrir_gerenciador(self):
        # """Abre a janela de configurações"""
        config_window = configuracao.GerenciadorConfiguracoes(self.root, False)
        
        # Opcional: aguardar a janela de configuração fechar antes de atualizar dados
        self.root.wait_window(config_window.root)
        
        # Atualizar dados após fechar a janela de configuração
        self.combo_meio_pagamento['values'] = self.carregar_meios_pagamento()
        self.combo_conta_despesa['values'] = self.carregar_categorias()

    def abrir_gerenciador2(self):
        """Abre o programa 2 dentro de uma nova janela Tkinter."""
        # opcionalmente pode usar Toplevel() se quiser que compartilhe a mesma root
        # # def abrir_gerenciador(self):
        # """Abre a janela de configurações"""
        config_window = MENUBD.GerenciadorConfiguracoes2(self.root, False)
        
        # Opcional: aguardar a janela de configuração fechar antes de atualizar dados
        self.root.wait_window(config_window.root)
        
        # Atualizar dados após fechar a janela de configuração
        self.combo_meio_pagamento['values'] = self.carregar_meios_pagamento()
        self.combo_conta_despesa['values'] = self.carregar_categorias()
    


    def gerar_relatorio_entre_datas(self):
        """Gera um relatório de despesas entre duas datas selecionadas"""
        try:
            # Criar janela de diálogo
            dialog = tk.Toplevel(self.root)
            dialog.title("Relatório Entre Datas")
            dialog.geometry("400x300")
            dialog.resizable(False, False)
            
            # Frame para os controles
            frame = ttk.Frame(dialog, padding="10")
            frame.pack(fill=tk.BOTH, expand=True)
            
            # Data atual
            data_atual = datetime.now()
            
            # Variáveis para as datas
            data_inicial_var = tk.StringVar(value=data_atual.strftime("%d/%m/%Y"))
            data_final_var = tk.StringVar(value=data_atual.strftime("%d/%m/%Y"))
            
            # Campos para selecionar datas
            ttk.Label(frame, text="Data Inicial:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
            data_inicial_entry = ttk.Entry(frame, textvariable=data_inicial_var, width=15)
            data_inicial_entry.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)
            
            ttk.Label(frame, text="Data Final:").grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
            data_final_entry = ttk.Entry(frame, textvariable=data_final_var, width=15)
            data_final_entry.grid(row=1, column=1, padx=5, pady=5, sticky=tk.W)
            
            # Função auxiliar para escolher data com calendário
            def escolher_data(entry_var):
                def pegar_data():
                    data_selecionada = cal.selection_get()
                    entry_var.set(data_selecionada.strftime("%d/%m/%Y"))
                    top.destroy()
                
                top = tk.Toplevel(dialog)
                top.title("Selecionar Data")
                
                cal = calendar(top, selectmode='day', date_pattern='dd/mm/yyyy',
                            year=data_atual.year, month=data_atual.month, day=data_atual.day)
                cal.pack(pady=10)
                
                ttk.Button(top, text="Selecionar", command=pegar_data).pack(pady=5)
                
                # Centralizar e focar
                top.transient(dialog)
                top.wait_visibility()
                top.grab_set()
                top.focus_set()
            
            # Botões para abrir calendário
            ttk.Button(frame, text="📅", width=3, 
                    command=lambda: escolher_data(data_inicial_var)).grid(row=0, column=2, padx=5)
            ttk.Button(frame, text="📅", width=3, 
                    command=lambda: escolher_data(data_final_var)).grid(row=1, column=2, padx=5)
            
            # Função para gerar relatório e fechar diálogo
            def gerar():
                try:
                    # Converter datas do formato DD/MM/AAAA para AAAA-MM-DD
                    data_inicial = datetime.strptime(data_inicial_var.get(), "%d/%m/%Y").strftime("%Y-%m-%d")
                    data_final = datetime.strptime(data_final_var.get(), "%d/%m/%Y").strftime("%Y-%m-%d")
                    
                    # Validar datas
                    if data_inicial > data_final:
                        messagebox.showerror("Erro", "Data inicial deve ser anterior à data final.")
                        return
                        
                    self.mostrar_relatorio_entre_datas(data_inicial, data_final)
                    dialog.destroy()
                except ValueError:
                    messagebox.showerror("Erro", "Formato de data inválido. Use DD/MM/AAAA.")
            
            # Botão para gerar
            ttk.Button(frame, text="Gerar Relatório", command=gerar).grid(row=3, column=0, columnspan=3, pady=20)
            
            # Centralizar janela
            dialog.transient(self.root)
            dialog.wait_visibility()
            dialog.grab_set()
            dialog.focus_set()
            dialog.wait_window()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar relatório: {e}")

    def mostrar_relatorio_entre_datas(self, data_inicial, data_final):
        """Exibe o relatório para o período entre as datas selecionadas"""
        try:
            # Formatar datas para exibição
            data_inicial_formatada = datetime.strptime(data_inicial, "%Y-%m-%d").strftime("%d/%m/%Y")
            data_final_formatada = datetime.strptime(data_final, "%Y-%m-%d").strftime("%d/%m/%Y")
            
            # Consultar dados do período
            self.cursor.execute("""
                SELECT 
                    conta_despesa, 
                    SUM(valor) as total_valor,
                    COUNT(*) as total_registros
                FROM despesas
                WHERE data_pagamento BETWEEN ? AND ?
                GROUP BY conta_despesa
                ORDER BY total_valor DESC
            """, (data_inicial, data_final))
            
            resultados = self.cursor.fetchall()
            
            if not resultados:
                messagebox.showinfo("Relatório", f"Nenhuma despesa encontrada no período de {data_inicial_formatada} a {data_final_formatada}.")
                return
                
            # Calcular total geral
            total_geral = sum(row[1] for row in resultados)
            
            # Criar nova janela para o relatório
            janela_relatorio = tk.Toplevel(self.root)
            janela_relatorio.title(f"Relatório de {data_inicial_formatada} a {data_final_formatada}")
            janela_relatorio.geometry("900x600")
            
            # Frame principal
            frame_principal = ttk.Frame(janela_relatorio, padding="10")
            frame_principal.pack(fill=tk.BOTH, expand=True)
            
            # Título
            titulo = ttk.Label(frame_principal, 
                            text=f"Relatório de Despesas - {data_inicial_formatada} a {data_final_formatada}",
                            font=('Arial', 16, 'bold'))
            titulo.pack(pady=10)
            
            # Total geral
            total_label = ttk.Label(frame_principal, 
                                text=f"Total de Despesas: R$ {total_geral:.2f}".replace('.', ','),
                                font=('Arial', 12))
            total_label.pack(pady=5)
            
            # Frame com abas
            notebook = ttk.Notebook(frame_principal)
            notebook.pack(fill=tk.BOTH, expand=True, pady=10)
            
            # Aba de tabela
            tab_tabela = ttk.Frame(notebook)
            notebook.add(tab_tabela, text="Tabela")
            
            # Aba de gráfico
            tab_grafico = ttk.Frame(notebook)
            notebook.add(tab_grafico, text="Gráfico")
            
            # Criar tabela na primeira aba
            tree_frame = ttk.Frame(tab_tabela)
            tree_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
            
            # Scrollbar para a tabela
            scrollbar = ttk.Scrollbar(tree_frame)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # Tabela de resultados
            colunas = ('categoria', 'total', 'registros', 'percentual')
            tabela = ttk.Treeview(tree_frame, columns=colunas, show='headings', yscrollcommand=scrollbar.set)
            
            # Cabeçalhos
            tabela.heading('categoria', text='Categoria')
            tabela.heading('total', text='Total (R$)')
            tabela.heading('registros', text='Qtd. Registros')
            tabela.heading('percentual', text='% do Total')
            
            # Larguras
            tabela.column('categoria', width=150)
            tabela.column('total', width=100, anchor=tk.E)
            tabela.column('registros', width=100, anchor=tk.CENTER)
            tabela.column('percentual', width=100, anchor=tk.CENTER)
            
            # Preencher tabela
            for row in resultados:
                categoria, total, registros = row
                percentual = (total / total_geral) * 100
                
                tabela.insert('', tk.END, values=(
                    categoria,
                    f"R$ {total:.2f}".replace('.', ','),
                    registros,
                    f"{percentual:.1f}%"
                ))
                
            tabela.pack(fill=tk.BOTH, expand=True)
            scrollbar.config(command=tabela.yview)
            
            # Criar gráfico na segunda aba
            figura = plt.Figure(figsize=(7, 5), dpi=100)
            ax = figura.add_subplot(111)
            
            # Extrair dados para o gráfico
            categorias = [row[0] for row in resultados]
            valores = [row[1] for row in resultados]
            
            # Criar gráfico de pizza
            wedges, texts, autotexts = ax.pie(
                valores, 
                labels=None,
                autopct='%1.1f%%',
                startangle=90,
                shadow=False,
            )
            
            # Estilo do gráfico
            ax.set_title(f'Distribuição de Despesas - {data_inicial_formatada} a {data_final_formatada}')
            ax.axis('equal')  # Garantir que o gráfico seja circular
            
            # Adicionar legenda
            ax.legend(wedges, categorias, title="Categorias", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))
            
            # Adicionar o gráfico ao frame
            canvas = FigureCanvasTkAgg(figura, tab_grafico)
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

            # Função para exportar para Excel
            def exportar_para_excel():
                try:
                    # Solicitar local para salvar o arquivo
                    arquivo = filedialog.asksaveasfilename(
                        defaultextension=".xlsx",
                        filetypes=[("Excel files", "*.xlsx")],
                        title="Salvar Relatório"
                    )
                    
                    if not arquivo:
                        return
                    
                    # Criar planilha
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Relatório de Despesas"
                    
                    # Título
                    ws['A1'] = f"Relatório de Despesas - Período: {data_inicial_formatada} a {data_final_formatada}"
                    ws.merge_cells('A1:D1')
                    ws['A1'].font = Font(size=14, bold=True)
                    
                    # Total
                    ws['A2'] = f"Total de Despesas: R$ {total_geral:.2f}".replace('.', ',')
                    ws.merge_cells('A2:D2')
                    ws['A2'].font = Font(size=12)
                    
                    # Cabeçalhos
                    cabecalhos = ["Categoria", "Total (R$)", "Qtd. Registros", "% do Total"]
                    for col, cabecalho in enumerate(cabecalhos, start=1):
                        cell = ws.cell(row=4, column=col, value=cabecalho)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                    
                    # Dados
                    for row_idx, (categoria, total, registros) in enumerate(resultados, start=5):
                        percentual = (total / total_geral) * 100
                        
                        ws.cell(row=row_idx, column=1, value=categoria)
                        ws.cell(row=row_idx, column=2, value=f"R$ {total:.2f}".replace('.', ','))
                        ws.cell(row=row_idx, column=3, value=registros)
                        ws.cell(row=row_idx, column=4, value=f"{percentual:.1f}%")
                    
                    # Ajustar larguras
                    for col in range(1, 5):
                        ws.column_dimensions[get_column_letter(col)].width = 20
                    
                    # Salvar arquivo
                    wb.save(arquivo)
                    messagebox.showinfo("Exportar", f"Relatório exportado com sucesso para {arquivo}")
                    
                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao exportar relatório: {e}")
            
            # Botão para exportar
            ttk.Button(frame_principal, text="Exportar para Excel", 
                    command=exportar_para_excel).pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exibir relatório: {e}")        

    ##########################

    def gerar_relatorio_mensal_periodo(self):
        """Gera um relatório de despesas totais somadas por mês entre um intervalo de datas"""
        try:
            # Criar janela de diálogo
            dialog = tk.Toplevel(self.root)
            dialog.title("Relatório de Despesas por Mês")
            dialog.geometry("400x250")
            dialog.resizable(False, False)
            
            # Frame para os controles
            frame = ttk.Frame(dialog, padding="10")
            frame.pack(fill=tk.BOTH, expand=True)
            
            # Data atual
            data_atual = datetime.now()
            
            # Variáveis para as datas
            data_inicial_var = tk.StringVar(value=f"01/01/{data_atual.year}")
            data_final_var = tk.StringVar(value=data_atual.strftime("%d/%m/%Y"))
            
            # Campos para selecionar datas
            ttk.Label(frame, text="Data Inicial:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
            data_inicial_entry = ttk.Entry(frame, textvariable=data_inicial_var, width=15)
            data_inicial_entry.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)
            
            ttk.Label(frame, text="Data Final:").grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
            data_final_entry = ttk.Entry(frame, textvariable=data_final_var, width=15)
            data_final_entry.grid(row=1, column=1, padx=5, pady=5, sticky=tk.W)
            
            # Função auxiliar para escolher data com calendário
            def escolher_data(entry_var):
                def pegar_data():
                    data_selecionada = cal.selection_get()
                    entry_var.set(data_selecionada.strftime("%d/%m/%Y"))
                    top.destroy()
                
                top = tk.Toplevel(dialog)
                top.title("Selecionar Data")
                
                cal = calendar(top, selectmode='day', date_pattern='dd/mm/yyyy',
                            year=data_atual.year, month=data_atual.month, day=data_atual.day)
                cal.pack(pady=10)
                
                ttk.Button(top, text="Selecionar", command=pegar_data).pack(pady=5)
                
                # Centralizar e focar
                top.transient(dialog)
                top.wait_visibility()
                top.grab_set()
                top.focus_set()
            
            # Botões para abrir calendário
            ttk.Button(frame, text="📅", width=3, 
                    command=lambda: escolher_data(data_inicial_var)).grid(row=0, column=2, padx=5)
            ttk.Button(frame, text="📅", width=3, 
                    command=lambda: escolher_data(data_final_var)).grid(row=1, column=2, padx=5)
            
            # Função para gerar relatório e fechar diálogo
            def gerar():
                try:
                    # Converter datas do formato DD/MM/AAAA para AAAA-MM-DD
                    data_inicial = datetime.strptime(data_inicial_var.get(), "%d/%m/%Y").strftime("%Y-%m-%d")
                    data_final = datetime.strptime(data_final_var.get(), "%d/%m/%Y").strftime("%Y-%m-%d")
                    
                    # Validar datas
                    if data_inicial > data_final:
                        messagebox.showerror("Erro", "Data inicial deve ser anterior à data final.")
                        return
                        
                    self.mostrar_relatorio_mensal_periodo(data_inicial, data_final)
                    dialog.destroy()
                except ValueError:
                    messagebox.showerror("Erro", "Formato de data inválido. Use DD/MM/AAAA.")
            
            # Botão para gerar
            ttk.Button(frame, text="Gerar Relatório", command=gerar).grid(row=3, column=0, columnspan=3, pady=20)
            
            # Centralizar janela
            dialog.transient(self.root)
            dialog.wait_visibility()
            dialog.grab_set()
            dialog.focus_set()
            dialog.wait_window()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar relatório: {e}")

    def mostrar_relatorio_mensal_periodo(self, data_inicial, data_final):
        """Exibe o relatório de despesas totais somadas por mês dentro do período selecionado"""
        try:
            # Formatar datas para exibição
            data_inicial_formatada = datetime.strptime(data_inicial, "%Y-%m-%d").strftime("%d/%m/%Y")
            data_final_formatada = datetime.strptime(data_final, "%Y-%m-%d").strftime("%d/%m/%Y")
            
            # Consultar despesas agrupadas por mês
            self.cursor.execute("""
                SELECT 
                    strftime('%Y-%m', data_pagamento) as ano_mes,
                    strftime('%m/%Y', data_pagamento) as mes_ano_formatado,
                    SUM(valor) as total_valor,
                    COUNT(*) as total_registros
                FROM despesas
                WHERE data_pagamento BETWEEN ? AND ?
                GROUP BY ano_mes
                ORDER BY ano_mes
            """, (data_inicial, data_final))
            
            resultados = self.cursor.fetchall()
            
            if not resultados:
                messagebox.showinfo("Relatório", f"Nenhuma despesa encontrada no período de {data_inicial_formatada} a {data_final_formatada}.")
                return
                
            # Calcular total geral
            total_geral = sum(row[2] for row in resultados)
            
            # Criar nova janela para o relatório
            janela_relatorio = tk.Toplevel(self.root)
            janela_relatorio.title(f"Relatório Mensal - {data_inicial_formatada} a {data_final_formatada}")
            janela_relatorio.geometry("900x600")
            
            # Frame principal
            frame_principal = ttk.Frame(janela_relatorio, padding="10")
            frame_principal.pack(fill=tk.BOTH, expand=True)
            
            # Título
            titulo = ttk.Label(frame_principal, 
                            text=f"Despesas Mensais - {data_inicial_formatada} a {data_final_formatada}",
                            font=('Arial', 16, 'bold'))
            titulo.pack(pady=10)
            
            # Total geral
            total_label = ttk.Label(frame_principal, 
                                text=f"Total de Despesas: R$ {total_geral:.2f}".replace('.', ','),
                                font=('Arial', 12))
            total_label.pack(pady=5)
            
            # Frame com abas
            notebook = ttk.Notebook(frame_principal)
            notebook.pack(fill=tk.BOTH, expand=True, pady=10)
            
            # Aba de tabela
            tab_tabela = ttk.Frame(notebook)
            notebook.add(tab_tabela, text="Tabela")
            
            # Aba de gráfico
            tab_grafico = ttk.Frame(notebook)
            notebook.add(tab_grafico, text="Gráfico")
            
            # Criar tabela na primeira aba
            tree_frame = ttk.Frame(tab_tabela)
            tree_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
            
            # Scrollbar para a tabela
            scrollbar = ttk.Scrollbar(tree_frame)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # Tabela de resultados
            colunas = ('mes_ano', 'total', 'registros', 'percentual')
            tabela = ttk.Treeview(tree_frame, columns=colunas, show='headings', yscrollcommand=scrollbar.set)
            
            # Cabeçalhos
            tabela.heading('mes_ano', text='Mês/Ano')
            tabela.heading('total', text='Total (R$)')
            tabela.heading('registros', text='Qtd. Registros')
            tabela.heading('percentual', text='% do Total')
            
            # Larguras
            tabela.column('mes_ano', width=100, anchor=tk.CENTER)
            tabela.column('total', width=150, anchor=tk.E)
            tabela.column('registros', width=100, anchor=tk.CENTER)
            tabela.column('percentual', width=100, anchor=tk.CENTER)
            
            # Preencher tabela
            for row in resultados:
                ano_mes, mes_ano_formatado, total, registros = row
                percentual = (total / total_geral) * 100
                
                tabela.insert('', tk.END, values=(
                    mes_ano_formatado,
                    f"R$ {total:.2f}".replace('.', ','),
                    registros,
                    f"{percentual:.1f}%"
                ))
                
            tabela.pack(fill=tk.BOTH, expand=True)
            scrollbar.config(command=tabela.yview)
            
            # Criar gráfico de barras na segunda aba
            figura = plt.Figure(figsize=(7, 5), dpi=100)
            ax = figura.add_subplot(111)
            
            # Extrair dados para o gráfico
            meses = [row[1] for row in resultados]
            valores = [row[2] for row in resultados]
            
            # Criar gráfico de barras
            barras = ax.bar(meses, valores, color=self.cor_primaria)
            
            # Adicionar rótulos com valores
            for barra in barras:
                altura = barra.get_height()
                ax.text(barra.get_x() + barra.get_width()/2., altura + 0.05*max(valores),
                    f'R${altura:.2f}'.replace('.', ','),
                    ha='center', va='bottom', fontsize=9, rotation=0)
            
            # Estilo do gráfico
            ax.set_title(f'Despesas Mensais - {data_inicial_formatada} a {data_final_formatada}')
            ax.set_xlabel('Mês/Ano')
            ax.set_ylabel('Valor Total (R$)')
            ax.grid(True, linestyle='--', alpha=0.7, axis='y')
            
            # Rotacionar rótulos do eixo x para melhor visualização
            plt.setp(ax.get_xticklabels(), rotation=45, ha='right')
            
            # Ajustar layout
            figura.tight_layout()
            
            # Adicionar o gráfico ao frame
            canvas = FigureCanvasTkAgg(figura, tab_grafico)
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
            
            # Adicionar aba de gráfico de linha também
            tab_grafico_linha = ttk.Frame(notebook)
            notebook.add(tab_grafico_linha, text="Evolução")
            
            # Criar gráfico de linha na terceira aba
            figura_linha = plt.Figure(figsize=(7, 5), dpi=100)
            ax_linha = figura_linha.add_subplot(111)
            
            # Criar gráfico de linha
            ax_linha.plot(meses, valores, marker='o', linestyle='-', color=self.cor_primaria, linewidth=2)
            
            # Adicionar rótulos
            for i, valor in enumerate(valores):
                ax_linha.annotate(f'R${valor:.2f}'.replace('.', ','), 
                        (meses[i], valores[i]), 
                        textcoords="offset points",
                        xytext=(0,10), 
                        ha='center',
                        fontsize=8)
            
            # Estilo do gráfico
            ax_linha.set_title(f'Evolução de Despesas Mensais')
            ax_linha.set_xlabel('Mês/Ano')
            ax_linha.set_ylabel('Valor Total (R$)')
            ax_linha.grid(True, linestyle='--', alpha=0.7)
            
            # Rotacionar rótulos do eixo x
            plt.setp(ax_linha.get_xticklabels(), rotation=45, ha='right')
            
            # Ajustar layout
            figura_linha.tight_layout()
            
            # Adicionar o gráfico ao frame
            canvas_linha = FigureCanvasTkAgg(figura_linha, tab_grafico_linha)
            canvas_linha.get_tk_widget().pack(fill=tk.BOTH, expand=True)

            # Função para exportar para Excel
            def exportar_para_excel():
                try:
                    # Solicitar local para salvar o arquivo
                    arquivo = filedialog.asksaveasfilename(
                        defaultextension=".xlsx",
                        filetypes=[("Excel files", "*.xlsx")],
                        title="Salvar Relatório Mensal"
                    )
                    
                    if not arquivo:
                        return
                    
                    # Criar planilha
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Despesas Mensais"
                    
                    # Título
                    ws['A1'] = f"Relatório de Despesas Mensais - Período: {data_inicial_formatada} a {data_final_formatada}"
                    ws.merge_cells('A1:D1')
                    ws['A1'].font = Font(size=14, bold=True)
                    
                    # Total
                    ws['A2'] = f"Total de Despesas: R$ {total_geral:.2f}".replace('.', ',')
                    ws.merge_cells('A2:D2')
                    ws['A2'].font = Font(size=12)
                    
                    # Cabeçalhos
                    cabecalhos = ["Mês/Ano", "Total (R$)", "Qtd. Registros", "% do Total"]
                    for col, cabecalho in enumerate(cabecalhos, start=1):
                        cell = ws.cell(row=4, column=col, value=cabecalho)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                    
                    # Dados
                    for row_idx, (ano_mes, mes_ano_formatado, total, registros) in enumerate(resultados, start=5):
                        percentual = (total / total_geral) * 100
                        
                        ws.cell(row=row_idx, column=1, value=mes_ano_formatado)
                        ws.cell(row=row_idx, column=2, value=f"R$ {total:.2f}".replace('.', ','))
                        ws.cell(row=row_idx, column=3, value=registros)
                        ws.cell(row=row_idx, column=4, value=f"{percentual:.1f}%")
                    
                    # Ajustar larguras
                    for col in range(1, 5):
                        ws.column_dimensions[get_column_letter(col)].width = 20
                    
                    # Salvar arquivo
                    wb.save(arquivo)
                    messagebox.showinfo("Exportar", f"Relatório exportado com sucesso para {arquivo}")
                    
                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao exportar relatório: {e}")
            
            # Botão para exportar
            ttk.Button(frame_principal, text="Exportar para Excel", 
                    command=exportar_para_excel).pack(pady=10)
            
            # Função para buscar detalhes de um mês específico
            def mostrar_detalhe_mes(event):
                # Verificar se algum item foi selecionado
                selecao = tabela.selection()
                if not selecao:
                    return
                    
                # Obter informações do item selecionado
                item = tabela.item(selecao[0])
                mes_ano = item['values'][0]  # Formato MM/AAAA
                
                # Converter para formato YYYY-MM para consulta SQL
                mes, ano = mes_ano.split('/')
                ano_mes = f"{ano}-{mes}"
                
                # Criar nova janela para detalhes
                janela_detalhe = tk.Toplevel(janela_relatorio)
                janela_detalhe.title(f"Detalhes do mês {mes_ano}")
                janela_detalhe.geometry("800x500")
                
                # Frame principal
                frame_det = ttk.Frame(janela_detalhe, padding="10")
                frame_det.pack(fill=tk.BOTH, expand=True)
                
                # Título
                ttk.Label(frame_det, 
                        text=f"Detalhes de Despesas - {mes_ano}",
                        font=('Arial', 14, 'bold')).pack(pady=10)
                
                # Consultar despesas do mês por categoria
                self.cursor.execute("""
                    SELECT 
                        conta_despesa,
                        SUM(valor) as total_valor,
                        COUNT(*) as total_registros
                    FROM despesas
                    WHERE strftime('%Y-%m', data_pagamento) = ?
                    GROUP BY conta_despesa
                    ORDER BY total_valor DESC
                """, (ano_mes,))
                
                detalhes = self.cursor.fetchall()
                
                if not detalhes:
                    ttk.Label(frame_det, text="Nenhum detalhe disponível").pack(pady=20)
                    return
                    
                # Total do mês
                total_mes = sum(row[1] for row in detalhes)
                ttk.Label(frame_det, 
                        text=f"Total do Mês: R$ {total_mes:.2f}".replace('.', ','),
                        font=('Arial', 12)).pack(pady=5)
                
                # Frame da tabela
                tree_frame_det = ttk.Frame(frame_det)
                tree_frame_det.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
                
                # Scrollbar
                scrollbar_det = ttk.Scrollbar(tree_frame_det)
                scrollbar_det.pack(side=tk.RIGHT, fill=tk.Y)
                
                # Tabela
                colunas_det = ('categoria', 'total', 'registros', 'percentual')
                tabela_det = ttk.Treeview(tree_frame_det, columns=colunas_det, show='headings', 
                                        yscrollcommand=scrollbar_det.set)
                
                # Cabeçalhos
                tabela_det.heading('categoria', text='Categoria')
                tabela_det.heading('total', text='Total (R$)')
                tabela_det.heading('registros', text='Qtd. Registros')
                tabela_det.heading('percentual', text='% do Mês')
                
                # Larguras
                tabela_det.column('categoria', width=200)
                tabela_det.column('total', width=150, anchor=tk.E)
                tabela_det.column('registros', width=100, anchor=tk.CENTER)
                tabela_det.column('percentual', width=100, anchor=tk.CENTER)
                
                # Preencher tabela
                for categoria, total, registros in detalhes:
                    percentual = (total / total_mes) * 100
                    
                    tabela_det.insert('', tk.END, values=(
                        categoria,
                        f"R$ {total:.2f}".replace('.', ','),
                        registros,
                        f"{percentual:.1f}%"
                    ))
                    
                tabela_det.pack(fill=tk.BOTH, expand=True)
                scrollbar_det.config(command=tabela_det.yview)
            
            # Vincular duplo clique na tabela para mostrar detalhes
            tabela.bind("<Double-1>", mostrar_detalhe_mes)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exibir relatório: {e}")




    ####################

    # INÍCIO DA MODIFICAÇÃO: Nova função para popular o combobox de meses
    def _repovoar_combo_meses_grafico1(self):
        """Consulta o BD e preenche o combobox de seleção de mês para o gráfico."""
        try:
            selecao_atual = self.combo_mes_grafico.get()

            self.cursor.execute("""
                SELECT DISTINCT 
                    strftime('%Y-%m', data_pagamento) as mes_query,
                    strftime('%m/%Y', data_pagamento) as mes_display
                FROM despesas 
                ORDER BY mes_query DESC
            """)
            meses_disponiveis = self.cursor.fetchall()
            
            self.mapa_meses_grafico = {mes_display: mes_query for mes_query, mes_display in meses_disponiveis}
            valores_combo = [mes_display for _, mes_display in meses_disponiveis]
            
            self.combo_mes_grafico['values'] = valores_combo
            
            if selecao_atual in valores_combo:
                self.combo_mes_grafico.set(selecao_atual)
            elif valores_combo:
                self.combo_mes_grafico.set(valores_combo[0])
            else:
                # Caso não haja dados, define com o mês atual
                self.combo_mes_grafico.set(datetime.now().strftime('%m/%Y'))

        except Exception as e:
            print(f"Erro ao repovoar combo de meses do gráfico: {e}")
    # FIM DA MODIFICAÇÃO

    def _repovoar_combo_meses_grafico(self):
            """
            Consulta o BD, preenche o combobox de seleção de mês para o gráfico
            e define o mês atual como padrão na inicialização.
            """
            try:
                # Tenta preservar a seleção que o usuário já fez
                selecao_previa = self.combo_mes_grafico.get()

                # Obtém o mês e ano atuais nos formatos para exibição e para consulta
                mes_atual_display = datetime.now().strftime('%m/%Y')
                mes_atual_query = datetime.now().strftime('%Y-%m')

                # Busca no banco de dados todos os meses que já possuem despesas
                self.cursor.execute("""
                    SELECT DISTINCT 
                        strftime('%Y-%m', data_pagamento) as mes_query,
                        strftime('%m/%Y', data_pagamento) as mes_display
                    FROM despesas 
                    ORDER BY mes_query DESC
                """)
                meses_disponiveis = self.cursor.fetchall()
                
                # Prepara a lista de meses para o combobox e o dicionário de mapeamento
                self.mapa_meses_grafico = {display: query for query, display in meses_disponiveis}
                valores_combo = [display for _, display in meses_disponiveis]
                
                # Garante que o mês atual esteja sempre na lista de opções
                if mes_atual_display not in self.mapa_meses_grafico:
                    valores_combo.insert(0, mes_atual_display)
                    self.mapa_meses_grafico[mes_atual_display] = mes_atual_query
                
                self.combo_mes_grafico['values'] = valores_combo

                # Lógica para definir a seleção no combobox
                if selecao_previa in valores_combo:
                    # Se o usuário já havia selecionado um mês, mantém essa seleção
                    self.combo_mes_grafico.set(selecao_previa)
                else:
                    # Caso contrário (incluindo a primeira vez que o programa abre),
                    # define o mês atual como o padrão
                    self.combo_mes_grafico.set(mes_atual_display)

            except Exception as e:
                print(f"Erro ao repovoar combo de meses do gráfico: {e}")


    def carregar_despesas(self):
        """Carrega todas as despesas do banco de dados para a tabela"""
        # Limpar dados da tabela
        for item in self.tabela.get_children():
            self.tabela.delete(item)
            
        try:
            # Consultar todas as despesas
            self.cursor.execute("""
                SELECT id, descricao, meio_pagamento, conta_despesa, valor, 
                       num_parcelas, strftime('%d/%m/%Y', data_pagamento) as data_formatada
                FROM despesas
                ORDER BY data_pagamento DESC
            """)
            
            # Inserir dados na tabela
            for row in self.cursor.fetchall():
                # Formatar valor como R$
                valor_formatado = f"R$ {row[4]:.2f}".replace('.', ',')
                
                self.tabela.insert('', tk.END, values=(
                    row[0], row[1], row[2], row[3], 
                    valor_formatado, row[5], row[6]
                ))

            # INÍCIO DA MODIFICAÇÃO: Atualiza a lista de meses e o gráfico
            self._repovoar_combo_meses_grafico()
            self.atualizar_grafico()
            # FIM DA MODIFICAÇÃO
                
        except sqlite3.Error as e:
            messagebox.showerror("Erro de Banco de Dados", f"Erro ao carregar despesas: {e}")
    
    def limpar_campos(self):
        """Limpa todos os campos do formulário"""
        self.id_despesa.set('')
        self.descricao.set('')
        self.meio_pagamento.set('')
        self.conta_despesa.set('')
        self.valor.set(0.0)
        self.num_parcelas.set(1)
        self.data_entry.set_date(datetime.now())
        
        # Desabilitar botões de atualizar e excluir
        self.btn_atualizar['state'] = 'disabled'
        self.btn_excluir['state'] = 'disabled'
        
        # Habilitar botão de salvar
        self.btn_salvar['state'] = 'normal'
    
    def validar_campos(self):
        """Valida se todos os campos obrigatórios foram preenchidos"""
        if not self.descricao.get().strip():
            messagebox.showwarning("Campo Obrigatório", "Por favor, preencha a descrição.")
            return False
            
        if not self.meio_pagamento.get().strip():
            messagebox.showwarning("Campo Obrigatório", "Por favor, selecione o meio de pagamento.")
            return False
            
        if not self.conta_despesa.get().strip():
            messagebox.showwarning("Campo Obrigatório", "Por favor, selecione a categoria da despesa.")
            return False
            
        try:
            valor = float(self.valor.get())
            if valor <= 0:
                messagebox.showwarning("Valor Inválido", "O valor deve ser maior que zero.")
                return False
        except:
            messagebox.showwarning("Valor Inválido", "Por favor, informe um valor numérico válido.")
            return False
            
        return True
    
    def salvar_despesa(self):
        """Salva uma nova despesa no banco de dados"""
        if not self.validar_campos():
            return
            
        try:
            # Obter a data do campo DateEntry
            data_pagamento = self.data_entry.get_date().strftime('%Y-%m-%d')
            
            # Inserir no banco de dados
            self.cursor.execute("""
                INSERT INTO despesas (descricao, meio_pagamento, conta_despesa, valor,
                                     num_parcelas, data_registro, data_pagamento)
                VALUES (?, ?, ?, ?, ?, date('now'), ?)
            """, (
                self.descricao.get().strip(),
                self.meio_pagamento.get().strip(),
                self.conta_despesa.get().strip(),
                self.valor.get(),
                self.num_parcelas.get(),
                data_pagamento
            ))
            
            self.conn.commit()
            messagebox.showinfo("Sucesso", "Despesa registrada com sucesso!")
            
            # Limpar campos e recarregar dados
            self.limpar_campos()
            self.carregar_despesas()
            
        except sqlite3.Error as e:
            messagebox.showerror("Erro ao Salvar", f"Ocorreu um erro: {e}")
    
    def selecionar_item(self, event):
        """Seleciona um item da tabela para edição"""
        try:
            # Obter o item selecionado
            selected_item = self.tabela.selection()[0]
            item = self.tabela.item(selected_item)
            values = item['values']
            
            if not values:
                return
                
            # Limpar campos primeiro
            self.limpar_campos()
            
            # Preencher campos com os valores do item selecionado
            self.id_despesa.set(values[0])
            self.descricao.set(values[1])
            self.meio_pagamento.set(values[2])
            self.conta_despesa.set(values[3])
            
            # Converter valor formatado para número
            valor_str = values[4].replace('R$', '').replace(' ', '').replace(',', '.')
            self.valor.set(float(valor_str))
            
            self.num_parcelas.set(values[5])
            
            # Converter data formato DD/MM/YYYY para objeto date
            data_str = values[6]
            dia, mes, ano = map(int, data_str.split('/'))
            self.data_entry.set_date(datetime(ano, mes, dia))
            
            # Habilitar botões de atualizar e excluir
            self.btn_atualizar['state'] = 'normal'
            self.btn_excluir['state'] = 'normal'
            
            # Desabilitar botão de salvar
            self.btn_salvar['state'] = 'disabled'
            
        except IndexError:
            pass  # Nenhum item selecionado
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao selecionar item: {e}")
    
    def atualizar_despesa(self):
        """Atualiza um registro de despesa existente"""
        if not self.validar_campos() or not self.id_despesa.get():
            return
            
        try:
            # Confirmação do usuário
            confirmar = messagebox.askyesno("Confirmar Atualização", 
                                         "Tem certeza que deseja atualizar este registro?")
            if not confirmar:
                return
                
            # Obter a data do campo DateEntry
            data_pagamento = self.data_entry.get_date().strftime('%Y-%m-%d')
            
            # Atualizar registro no banco de dados
            self.cursor.execute("""
                UPDATE despesas SET 
                    descricao = ?, 
                    meio_pagamento = ?,
                    conta_despesa = ?,
                    valor = ?,
                    num_parcelas = ?,
                    data_pagamento = ?
                WHERE id = ?
            """, (
                self.descricao.get().strip(),
                self.meio_pagamento.get().strip(),
                self.conta_despesa.get().strip(),
                self.valor.get(),
                self.num_parcelas.get(),
                data_pagamento,
                self.id_despesa.get()
            ))
            
            self.conn.commit()
            messagebox.showinfo("Sucesso", "Registro atualizado com sucesso!")
            
            # Limpar campos e recarregar dados
            self.limpar_campos()
            self.carregar_despesas()
            
        except sqlite3.Error as e:
            messagebox.showerror("Erro ao Atualizar", f"Ocorreu um erro: {e}")
    
    def excluir_despesa(self):
        """Exclui um registro de despesa"""
        if not self.id_despesa.get():
            messagebox.showwarning("Selecione um Registro", "Por favor, selecione um registro para excluir.")
            return
            
        try:
            # Confirmação do usuário
            confirmar = messagebox.askyesno("Confirmar Exclusão", 
                                         "Tem certeza que deseja excluir este registro? Esta ação não pode ser desfeita.")
            if not confirmar:
                return
                
            # Excluir registro do banco de dados
            self.cursor.execute("DELETE FROM despesas WHERE id = ?", (self.id_despesa.get(),))
            self.conn.commit()
            
            messagebox.showinfo("Sucesso", "Registro excluído com sucesso!")
            
            # Limpar campos e recarregar dados
            self.limpar_campos()
            self.carregar_despesas()
            
        except sqlite3.Error as e:
            messagebox.showerror("Erro ao Excluir", f"Ocorreu um erro: {e}")
    
    def pesquisar_despesa(self):
        """Pesquisa despesas por termo em descrição ou categoria"""
        termo = self.pesquisa_termo.get().strip()
        
        if not termo:
            messagebox.showinfo("Pesquisa", "Por favor, digite um termo para pesquisar.")
            return
            
        # Limpar dados da tabela
        for item in self.tabela.get_children():
            self.tabela.delete(item)
            
        try:
            # Consultar despesas que correspondem ao termo
            self.cursor.execute("""
                SELECT id, descricao, meio_pagamento, conta_despesa, valor, 
                       num_parcelas, strftime('%d/%m/%Y', data_pagamento) as data_formatada
                FROM despesas
                WHERE descricao LIKE ? OR conta_despesa LIKE ? OR meio_pagamento LIKE ?
                ORDER BY data_pagamento DESC
            """, (f'%{termo}%', f'%{termo}%', f'%{termo}%'))
            
            resultados = self.cursor.fetchall()
            
            # Inserir dados na tabela
            for row in resultados:
                # Formatar valor como R$
                valor_formatado = f"R$ {row[4]:.2f}".replace('.', ',')
                
                self.tabela.insert('', tk.END, values=(
                    row[0], row[1], row[2], row[3], 
                    valor_formatado, row[5], row[6]
                ))
                
            if not resultados:
                messagebox.showinfo("Pesquisa", "Nenhum resultado encontrado para o termo informado.")
                
        except sqlite3.Error as e:
            messagebox.showerror("Erro de Pesquisa", f"Erro ao pesquisar despesas: {e}")

    # ===== INÍCIO DAS FUNÇÕES DO GRÁFICO ATUALIZADO =====
    def on_bar_pick(self, event):
        """
        Callback para o evento de clique em uma barra do gráfico.
        Abre uma nova janela com os detalhes da categoria selecionada.
        """
        # INÍCIO DA MODIFICAÇÃO: Lógica alterada para usar get_gid()
        try:
            gid = event.artist.get_gid()
            if not gid:
                return # Ignora o clique se não houver dados associados

            categoria_selecionada, ano_mes = gid
            self.mostrar_detalhes_categoria_mes(categoria_selecionada, ano_mes)
        except (AttributeError, ValueError, IndexError) as e:
            print(f"Clique ignorado ou erro ao processar: {e}")
            pass
        # FIM DA MODIFICAÇÃO


    def mostrar_detalhes_categoria_mes(self, categoria, ano_mes):
        """
        Cria uma nova janela para exibir todas as transações de uma
        categoria específica em um determinado mês.
        """
        try:
            mes_ano_formatado = datetime.strptime(ano_mes, "%Y-%m").strftime("%m/%Y")
            
            janela_detalhes = tk.Toplevel(self.root)
            janela_detalhes.title(f"Detalhes para '{categoria}' em {mes_ano_formatado}")
            janela_detalhes.geometry("600x400")
            janela_detalhes.transient(self.root)
            janela_detalhes.grab_set()

            frame_detalhes = ttk.Frame(janela_detalhes, padding="10")
            frame_detalhes.pack(fill=tk.BOTH, expand=True)

            self.cursor.execute("""
                SELECT strftime('%d/%m/%Y', data_pagamento), descricao, valor
                FROM despesas
                WHERE conta_despesa = ? AND strftime('%Y-%m', data_pagamento) = ?
                ORDER BY data_pagamento
            """, (categoria, ano_mes))
            
            transacoes = self.cursor.fetchall()

            if not transacoes:
                ttk.Label(frame_detalhes, text="Nenhuma transação encontrada.").pack()
                return

            colunas = ('data', 'descricao', 'valor')
            tabela_detalhes = ttk.Treeview(frame_detalhes, columns=colunas, show='headings')
            
            tabela_detalhes.heading('data', text='Data')
            tabela_detalhes.heading('descricao', text='Descrição')
            tabela_detalhes.heading('valor', text='Valor (R$)')

            tabela_detalhes.column('data', width=100, anchor=tk.CENTER)
            tabela_detalhes.column('descricao', width=300, anchor=tk.W)
            tabela_detalhes.column('valor', width=120, anchor=tk.E)

            total = 0
            for data, desc, valor in transacoes:
                valor_formatado = locale.currency(valor, grouping=True)
                tabela_detalhes.insert('', tk.END, values=(data, desc, valor_formatado))
                total += valor

            tabela_detalhes.pack(fill=tk.BOTH, expand=True)

            total_formatado = locale.currency(total, grouping=True)
            ttk.Label(frame_detalhes, text=f"Total: {total_formatado}", font=('Arial', 12, 'bold')).pack(pady=10)

        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível carregar os detalhes: {e}")

    # INÍCIO DA MODIFICAÇÃO: A função foi reescrita para ser dinâmica
    def atualizar_grafico(self, event=None):
        """
        Atualiza o gráfico de resumo para mostrar as 10 principais categorias de gastos
        do mês selecionado no combobox. O gráfico é interativo.
        """
        try:
            from matplotlib.ticker import FuncFormatter
            self.figura_grafico.clear()
            ax = self.figura_grafico.add_subplot(111)

            # Obter o mês selecionado no combobox
            mes_exibicao = self.combo_mes_grafico.get()
            if not mes_exibicao:
                # Se nada for selecionado, usa o mês/ano atual como padrão
                mes_exibicao = datetime.now().strftime('%m/%Y')
                mes_para_query = datetime.now().strftime('%Y-%m')
            else:
                mes_para_query = self.mapa_meses_grafico.get(mes_exibicao)

            # Atualiza o título do frame que contém o gráfico
            self.frame_grafico.config(text=f"Resumo de Gastos ({mes_exibicao})")

            # INÍCIO DA NOVA ALTERAÇÃO: Buscar e exibir o total do mês
            if mes_para_query:
                self.cursor.execute("SELECT SUM(valor) FROM despesas WHERE strftime('%Y-%m', data_pagamento) = ?", (mes_para_query,))
                total_mes = self.cursor.fetchone()[0]
                total_mes = total_mes if total_mes is not None else 0.0
                total_formatado = locale.currency(total_mes, grouping=True, symbol='R$')
                self.total_mes_selecionado_var.set(f"Total do Mês: {total_formatado}")
            else:
                self.total_mes_selecionado_var.set("Total do Mês: R$ 0,00")
            # FIM DA NOVA ALTERAÇÃO

            self.cursor.execute("""
                SELECT conta_despesa, SUM(valor) as total
                FROM despesas
                WHERE strftime('%Y-%m', data_pagamento) = ?
                GROUP BY conta_despesa
                ORDER BY total DESC
                LIMIT 10
            """, (mes_para_query,))
            
            resultados = self.cursor.fetchall()
            
            if resultados:
                categorias = [row[0] for row in resultados]
                valores = [row[1] for row in resultados]
                
                cores_grafico = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd',
                                 '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf']
                
                barras = ax.bar(categorias, valores, color=cores_grafico[:len(categorias)], picker=True)

                # Adiciona o 'gid' a cada barra para o evento de clique
                for i, barra in enumerate(barras):
                    barra.set_gid((categorias[i], mes_para_query))
                    altura = barra.get_height()
                    ax.text(barra.get_x() + barra.get_width() / 2., altura,
                           locale.currency(altura, symbol=True, grouping=True),
                           ha='center', va='bottom', rotation=0, fontsize=8)

                ax.set_title(f'Top 10 Categorias ({mes_exibicao})')
                ax.set_ylabel('Valor Total (R$)')
                ax.set_xlabel('Categoria')
                
                formatter = FuncFormatter(lambda y, _: locale.currency(y, symbol=True, grouping=True))
                ax.yaxis.set_major_formatter(formatter)

                plt.setp(ax.get_xticklabels(), rotation=45, ha='right', fontsize=8)
                
                ax.margins(y=0.15)
                
            else:
                ax.text(0.5, 0.5, f'Sem dados para o mês {mes_exibicao}', 
                       horizontalalignment='center', verticalalignment='center',
                       transform=ax.transAxes)
                ax.set_title(f'Top 10 Categorias ({mes_exibicao})')
                ax.axis('off')

            self.canvas_grafico.mpl_connect('pick_event', self.on_bar_pick)
            self.figura_grafico.tight_layout()
            self.canvas_grafico.draw()
            
        except sqlite3.Error as e:
            print(f"Erro de banco de dados ao atualizar gráfico: {e}")
        except Exception as e:
            import traceback
            print(f"Erro geral ao atualizar gráfico: {e}")
            traceback.print_exc()
    # FIM DA MODIFICAÇÃO
    
    def gerar_relatorio_mensal(self):
        """Gera um relatório mensal de despesas"""
        try:
            # Solicitar mês e ano ao usuário
            meses = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                   "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
            
            mes_atual = datetime.now().month
            ano_atual = datetime.now().year
            
            # Criar janela de diálogo
            dialog = tk.Toplevel(self.root)
            dialog.title("Relatório Mensal")
            dialog.geometry("300x200")
            dialog.resizable(False, False)
            
            # Variáveis para selecionar mês e ano
            mes_var = tk.IntVar(value=mes_atual)
            ano_var = tk.IntVar(value=ano_atual)
            
            # Frame para os controles
            frame = ttk.Frame(dialog, padding="10")
            frame.pack(fill=tk.BOTH, expand=True)
            
            # Combobox para o mês
            ttk.Label(frame, text="Mês:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
            mes_combo = ttk.Combobox(frame, values=meses, width=15)
            mes_combo.current(mes_atual - 1)  # Ajustar para índice baseado em zero
            mes_combo.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)
            
            # Spinbox para o ano
            ttk.Label(frame, text="Ano:").grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
            ano_spin = ttk.Spinbox(frame, from_=2000, to=2100, textvariable=ano_var, width=10)
            ano_spin.grid(row=1, column=1, padx=5, pady=5, sticky=tk.W)
            
            # Função para gerar relatório e fechar diálogo
            def gerar():
                mes_indice = meses.index(mes_combo.get()) + 1
                ano = ano_var.get()
                self.mostrar_relatorio_mensal(mes_indice, ano)
                dialog.destroy()
            
            # Botão para gerar
            ttk.Button(frame, text="Gerar Relatório", command=gerar).grid(row=2, column=0, columnspan=2, pady=20)
            
            # Centralizar janela
            dialog.transient(self.root)
            dialog.wait_visibility()
            dialog.grab_set()
            dialog.focus_set()
            dialog.wait_window()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar relatório: {e}")
   
    def mostrar_relatorio_mensal(self, mes, ano):
        """Exibe o relatório mensal para o mês e ano selecionados, com gráfico de pizza interativo."""
        try:
            # Importações locais para manter a função mais independente, se necessário
            from matplotlib.figure import Figure
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            import matplotlib.pyplot as plt

            # Calcular início e fim do mês
            primeiro_dia = f"{ano}-{mes:02d}-01"
            ultimo_dia = f"{ano}-{mes:02d}-{calendar.monthrange(ano, mes)[1]}"
            
            # Consultar dados do mês
            self.cursor.execute("""
                SELECT 
                    conta_despesa, 
                    SUM(valor) as total_valor,
                    COUNT(*) as total_registros
                FROM despesas
                WHERE data_pagamento BETWEEN ? AND ?
                GROUP BY conta_despesa
                ORDER BY total_valor DESC
            """, (primeiro_dia, ultimo_dia))
            
            resultados = self.cursor.fetchall()
            
            if not resultados:
                messagebox.showinfo("Relatório", f"Nenhuma despesa encontrada para {calendar.month_name[mes]} de {ano}.")
                return
                
            # Calcular total geral
            total_geral = sum(row[1] for row in resultados)
            
            # Criar nova janela para o relatório
            janela_relatorio = tk.Toplevel(self.root)
            janela_relatorio.title(f"Relatório de {calendar.month_name[mes]} de {ano}")
            janela_relatorio.geometry("900x700") # Aumentado um pouco para melhor visualização da legenda
            
            # Frame principal
            frame_principal = ttk.Frame(janela_relatorio, padding="10")
            frame_principal.pack(fill=tk.BOTH, expand=True)
            
            # Título
            titulo_texto = f"Relatório de Despesas - {calendar.month_name[mes]} de {ano}"
            titulo_label = ttk.Label(frame_principal, text=titulo_texto, font=('Arial', 16, 'bold'))
            titulo_label.pack(pady=10)
            
            # Total geral
            total_texto = f"Total de Despesas: {locale.currency(total_geral, grouping=True)}"
            total_label_widget = ttk.Label(frame_principal, text=total_texto, font=('Arial', 12))
            total_label_widget.pack(pady=5)
            
            # Frame com abas
            notebook = ttk.Notebook(frame_principal)
            notebook.pack(fill=tk.BOTH, expand=True, pady=10)
            
            # Aba de tabela
            tab_tabela = ttk.Frame(notebook)
            notebook.add(tab_tabela, text="Tabela")
            
            # Aba de gráfico
            tab_grafico = ttk.Frame(notebook)
            notebook.add(tab_grafico, text="Gráfico")
            
            # --- Início: Funções aninhadas para interatividade do gráfico de pizza ---
            def mostrar_detalhes_transacao_pizza(categoria_clicada, p_ano, p_mes):
                try:
                    detalhes_window = tk.Toplevel(janela_relatorio)
                    mes_nome = calendar.month_name[p_mes]
                    detalhes_window.title(f"Detalhes: {categoria_clicada} ({mes_nome}/{p_ano})")
                    detalhes_window.geometry("600x400")
                    detalhes_window.transient(janela_relatorio) # Mantém sobre a janela principal
                    detalhes_window.grab_set() # Foca nesta janela

                    tree_frame = ttk.Frame(detalhes_window, padding="10")
                    tree_frame.pack(fill=tk.BOTH, expand=True)

                    mes_ano_sql = f"{p_ano}-{p_mes:02d}"
                    self.cursor.execute("""
                        SELECT strftime('%d/%m/%Y', data_pagamento), descricao, valor
                        FROM despesas
                        WHERE conta_despesa = ? AND strftime('%Y-%m', data_pagamento) = ?
                        ORDER BY data_pagamento ASC
                    """, (categoria_clicada, mes_ano_sql))
                    transacoes = self.cursor.fetchall()

                    cols = ('data', 'descricao', 'valor')
                    tree = ttk.Treeview(tree_frame, columns=cols, show='headings')
                    tree.heading('data', text='Data')
                    tree.heading('descricao', text='Descrição')
                    tree.heading('valor', text='Valor')
                    tree.column('data', width=100, anchor=tk.CENTER)
                    tree.column('descricao', width=350, anchor=tk.W)
                    tree.column('valor', width=120, anchor=tk.E)
                    
                    scrollbar_detalhes = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
                    tree.configure(yscrollcommand=scrollbar_detalhes.set)
                    scrollbar_detalhes.pack(side=tk.RIGHT, fill=tk.Y)
                    tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

                    soma_categoria = 0
                    if transacoes:
                        for data_t, desc_t, valor_t in transacoes:
                            tree.insert("", "end", values=(data_t, desc_t, locale.currency(valor_t, grouping=True)))
                            soma_categoria += valor_t
                    else:
                        tree.insert("", "end", values=("", "Nenhuma transação encontrada.", ""))

                    ttk.Label(detalhes_window, text=f"Total para {categoria_clicada}: {locale.currency(soma_categoria, grouping=True)}", font=('Arial', 10, 'bold')).pack(pady=5)
                
                except Exception as e_detail:
                    messagebox.showerror("Erro Detalhes", f"Não foi possível mostrar os detalhes: {e_detail}", parent=janela_relatorio)

            def on_slice_pick(event):
                gid = event.artist.get_gid()
                if gid:
                    categoria_selecionada, ano_selecionado, mes_selecionado = gid
                    mostrar_detalhes_transacao_pizza(categoria_selecionada, ano_selecionado, mes_selecionado)
            # --- Fim: Funções aninhadas ---

            # Criar tabela na primeira aba
            tree_frame_tabela = ttk.Frame(tab_tabela)
            tree_frame_tabela.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
            
            scrollbar_tabela = ttk.Scrollbar(tree_frame_tabela)
            scrollbar_tabela.pack(side=tk.RIGHT, fill=tk.Y)
            
            colunas_tabela = ('categoria', 'total', 'registros', 'percentual')
            tabela_widget = ttk.Treeview(tree_frame_tabela, columns=colunas_tabela, show='headings', yscrollcommand=scrollbar_tabela.set)
            
            tabela_widget.heading('categoria', text='Categoria')
            tabela_widget.heading('total', text='Total (R$)')
            tabela_widget.heading('registros', text='Qtd. Registros')
            tabela_widget.heading('percentual', text='% do Total')
            
            tabela_widget.column('categoria', width=250, anchor=tk.W) # Aumentado para nomes de categoria
            tabela_widget.column('total', width=120, anchor=tk.E)
            tabela_widget.column('registros', width=100, anchor=tk.CENTER)
            tabela_widget.column('percentual', width=100, anchor=tk.CENTER)
            
            for row in resultados:
                categoria, total, registros = row
                percentual = (total / total_geral) * 100 if total_geral else 0
                tabela_widget.insert('', tk.END, values=(
                    categoria,
                    locale.currency(total, grouping=True),
                    registros,
                    f"{percentual:.1f}%"
                ))
                
            tabela_widget.pack(fill=tk.BOTH, expand=True)
            scrollbar_tabela.config(command=tabela_widget.yview)
            
            # Criar gráfico de pizza na segunda aba
            figura_pizza = Figure(figsize=(8, 6), dpi=100) # Ajustado tamanho
            ax_pizza = figura_pizza.add_subplot(111)
            
            categorias_grafico = [row[0] for row in resultados]
            valores_grafico = [row[1] for row in resultados]

            # Gerar cores distintas para o gráfico de pizza
            num_categorias = len(categorias_grafico)
            # Usar um colormap que tenha cores variadas. 'tab20' é uma boa opção.
            if num_categorias > 0 :
                cores = plt.cm.get_cmap('tab20', num_categorias).colors 
            else:
                cores = []

            wedges, texts, autotexts = ax_pizza.pie(
                valores_grafico, 
                autopct='%1.1f%%',
                startangle=140, # Ângulo inicial para melhor visualização das porcentagens
                colors=cores,
                pctdistance=0.85 # Distância das porcentagens do centro
            )

            # Tornar cada fatia clicável e definir seu GID (Group ID)
            for i, wedge in enumerate(wedges):
                wedge.set_picker(True)
                wedge.set_gid((categorias_grafico[i], ano, mes)) # Passa (categoria, ano_int, mes_int)

            # Estilo do gráfico de pizza
            ax_pizza.set_title(f'Distribuição de Despesas - {calendar.month_name[mes]} de {ano}', pad=20)
            ax_pizza.axis('equal')  # Garante que o gráfico seja circular
            
            # Adicionar legenda à direita
            # Ajustar bbox_to_anchor para garantir que a legenda caiba e não sobreponha o gráfico
            ax_pizza.legend(wedges, categorias_grafico, title="Categorias", 
                            loc="center left", bbox_to_anchor=(1, 0.5), fontsize='small')
            
            figura_pizza.tight_layout(rect=[0, 0, 0.85, 1]) # Ajustar layout para dar espaço à legenda

            canvas_pizza = FigureCanvasTkAgg(figura_pizza, tab_grafico)
            canvas_pizza_widget = canvas_pizza.get_tk_widget()
            canvas_pizza_widget.pack(fill=tk.BOTH, expand=True)
            canvas_pizza.mpl_connect('pick_event', on_slice_pick) # Conectar evento de clique
            
            # Botão para exportar
            ttk.Button(frame_principal, text="Exportar para Excel", 
                    command=lambda: self.exportar_relatorio_excel(mes, ano)).pack(pady=10)
                
        except Exception as e:
            messagebox.showerror("Erro ao Exibir Relatório", f"Ocorreu um erro: {e}")
            import traceback
            traceback.print_exc()
        
      
            
    def gerar_relatorio_meio_pagamento(self):
        """Gera um relatório de despesas por meio de pagamento"""
        try:
            # Obter todos os meios de pagamento
            self.cursor.execute("SELECT DISTINCT meio_pagamento FROM despesas ORDER BY meio_pagamento")
            meios_pagamento = [row[0] for row in self.cursor.fetchall()]
            
            if not meios_pagamento:
                messagebox.showinfo("Relatório", "Não há meios de pagamento cadastrados.")
                return
                
            # Criar janela de diálogo
            dialog = tk.Toplevel(self.root)
            dialog.title("Relatório por Meio de Pagamento")
            dialog.geometry("300x150")
            dialog.resizable(False, False)
            
            # Frame para os controles
            frame = ttk.Frame(dialog, padding="10")
            frame.pack(fill=tk.BOTH, expand=True)
            
            # Combobox para meio de pagamento
            ttk.Label(frame, text="Selecione:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
            meio_pagamento_combo = ttk.Combobox(frame, values=meios_pagamento, width=20)
            meio_pagamento_combo.current(0)
            meio_pagamento_combo.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)
            
            # Função para gerar relatório e fechar diálogo
            def gerar():
                meio_pagamento = meio_pagamento_combo.get()
                self.mostrar_relatorio_meio_pagamento(meio_pagamento)
                dialog.destroy()
            
            # Botão para gerar
            ttk.Button(frame, text="Gerar Relatório", command=gerar).grid(row=2, column=0, columnspan=2, pady=20)
            
            # Centralizar janela
            dialog.transient(self.root)
            dialog.wait_visibility()
            dialog.grab_set()
            dialog.focus_set()
            dialog.wait_window()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar relatório: {e}")

    def mostrar_relatorio_meio_pagamento(self, meio_pagamento):
        """Exibe o relatório por meio de pagamento selecionado"""
        try:
            # Consultar dados do meio de pagamento
            self.cursor.execute("""
                SELECT 
                    strftime('%Y-%m', data_pagamento) as mes_ano,
                    strftime('%m/%Y', data_pagamento) as mes_ano_formatado,
                    SUM(valor) as total_valor
                FROM despesas
                WHERE meio_pagamento = ?
                GROUP BY mes_ano
                ORDER BY mes_ano DESC
            """, (meio_pagamento,))
            
            resultados = self.cursor.fetchall()
            
            if not resultados:
                messagebox.showinfo("Relatório", f"Nenhuma despesa encontrada para o meio de pagamento {meio_pagamento}.")
                return
                
            # Calcular total geral
            total_geral = sum(row[2] for row in resultados)
            
            # Criar nova janela para o relatório
            janela_relatorio = tk.Toplevel(self.root)
            janela_relatorio.title(f"Relatório do Meio de Pagamento: {meio_pagamento}")
            janela_relatorio.geometry("800x600")
            
            # Frame principal
            frame_principal = ttk.Frame(janela_relatorio, padding="10")
            frame_principal.pack(fill=tk.BOTH, expand=True)
            
            # Título
            titulo = ttk.Label(frame_principal, 
                            text=f"Relatório de Despesas - Meio de Pagamento: {meio_pagamento}",
                            font=('Arial', 16, 'bold'))
            titulo.pack(pady=10)
            
            # Total geral
            total_label = ttk.Label(frame_principal, 
                                text=f"Total de Despesas: R$ {total_geral:.2f}".replace('.', ','),
                                font=('Arial', 12))
            total_label.pack(pady=5)
            
            # Frame com abas
            notebook = ttk.Notebook(frame_principal)
            notebook.pack(fill=tk.BOTH, expand=True, pady=10)
            
            # Aba de tabela
            tab_tabela = ttk.Frame(notebook)
            notebook.add(tab_tabela, text="Tabela")
            
            # Aba de gráfico
            tab_grafico = ttk.Frame(notebook)
            notebook.add(tab_grafico, text="Gráfico")
            
            # Criar tabela na primeira aba
            tree_frame = ttk.Frame(tab_tabela)
            tree_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
            
            # Scrollbar para a tabela
            scrollbar = ttk.Scrollbar(tree_frame)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # Tabela de resultados
            colunas = ('mes_ano', 'total')
            tabela = ttk.Treeview(tree_frame, columns=colunas, show='headings', yscrollcommand=scrollbar.set)
            
            # Cabeçalhos
            tabela.heading('mes_ano', text='Mês/Ano')
            tabela.heading('total', text='Total (R$)')
            
            # Larguras
            tabela.column('mes_ano', width=150, anchor=tk.CENTER)
            tabela.column('total', width=150, anchor=tk.E)
            
            # Preencher tabela
            for row in resultados:
                mes_ano_key, mes_ano_formatado, total = row
                
                tabela.insert('', tk.END, values=(
                    mes_ano_formatado,
                    f"R$ {total:.2f}".replace('.', ',')
                ))
                
            tabela.pack(fill=tk.BOTH, expand=True)
            scrollbar.config(command=tabela.yview)
            
            # Criar gráfico na segunda aba
            figura = plt.Figure(figsize=(7, 5), dpi=100)
            ax = figura.add_subplot(111)
            
            # Extrair dados para o gráfico (invertendo a ordem para cronológica)
            meses = [row[1] for row in reversed(resultados)]
            valores = [row[2] for row in reversed(resultados)]
            
            # Criar gráfico de linha
            ax.plot(meses, valores, marker='o', linestyle='-', color=self.cor_primaria, linewidth=2)
            
            # Adicionar rótulos
            for i, valor in enumerate(valores):
                ax.annotate(f'R${valor:.2f}'.replace('.', ','), 
                        (meses[i], valores[i]), 
                        textcoords="offset points",
                        xytext=(0,10), 
                        ha='center',
                        fontsize=8)
            
            # Estilo do gráfico
            ax.set_title(f'Evolução de Despesas - {meio_pagamento}')
            ax.set_xlabel('Mês/Ano')
            ax.set_ylabel('Valor Total (R$)')
            ax.grid(True, linestyle='--', alpha=0.7)
            
            # Rotacionar rótulos do eixo x
            plt.setp(ax.get_xticklabels(), rotation=45, ha='right')
            
            # Ajustar layout
            figura.tight_layout()
            
            # Adicionar o gráfico ao frame
            canvas = FigureCanvasTkAgg(figura, tab_grafico)
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
            
            # Botão para exportar
            ttk.Button(frame_principal, text="Exportar para Excel", 
                    command=lambda: self.exportar_meio_pagamento_excel(meio_pagamento)).pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exibir relatório: {e}")

    def exportar_meio_pagamento_excel(self, meio_pagamento):
        """Exporta o relatório de meio de pagamento para Excel"""
        try:
            # Selecionar o local para salvar
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"Relatorio_MeioPagamento_{meio_pagamento.replace(' ', '_')}"
            )
            
            if not file_path:
                return
                
            # Consultar dados
            self.cursor.execute("""
                SELECT 
                    strftime('%Y-%m', data_pagamento) as mes_ano,
                    strftime('%m/%Y', data_pagamento) as mes_ano_formatado,
                    SUM(valor) as total_valor
                FROM despesas
                WHERE meio_pagamento = ?
                GROUP BY mes_ano
                ORDER BY mes_ano DESC
            """, (meio_pagamento,))
            
            resultados = self.cursor.fetchall()
            
            # Criar workbook e worksheet
            workbook = xlsxwriter.Workbook(file_path)
            worksheet = workbook.add_worksheet("Relatório")
            
            # Formatos
            titulo_formato = workbook.add_format({
                'bold': True,
                'font_size': 14,
                'align': 'center',
                'valign': 'vcenter'
            })
            
            cabecalho_formato = workbook.add_format({
                'bold': True,
                'font_size': 12,
                'align': 'center',
                'valign': 'vcenter',
                'bg_color': '#f2f2f2',
                'border': 1
            })
            
            data_formato = workbook.add_format({
                'align': 'center',
                'border': 1
            })
            
            valor_formato = workbook.add_format({
                'num_format': 'R$ #,##0.00',
                'align': 'right',
                'border': 1
            })
            
            total_formato = workbook.add_format({
                'num_format': 'R$ #,##0.00',
                'align': 'right',
                'bold': True,
                'bg_color': '#f2f2f2',
                'border': 1
            })
            
            # Título
            worksheet.merge_range('A1:C1', f"Relatório de Despesas - Meio de Pagamento: {meio_pagamento}", titulo_formato)
            worksheet.set_row(0, 30)
            
            # Cabeçalhos
            worksheet.write('A3', 'Mês/Ano', cabecalho_formato)
            worksheet.write('B3', 'Total', cabecalho_formato)
            
            # Definir larguras das colunas
            worksheet.set_column('A:A', 15)
            worksheet.set_column('B:B', 20)
            
            # Preencher dados
            linha = 3
            total_geral = 0
            
            for row in resultados:
                mes_ano_key, mes_ano_formatado, total = row
                
                worksheet.write(linha, 0, mes_ano_formatado, data_formato)
                worksheet.write(linha, 1, total, valor_formato)
                
                total_geral += total
                linha += 1
            
            # Total geral
            worksheet.write(linha + 1, 0, 'Total Geral', cabecalho_formato)
            worksheet.write(linha + 1, 1, total_geral, total_formato)
            
            # Criar gráfico
            grafico = workbook.add_chart({'type': 'line'})
            
            # Adicionar série ao gráfico - precisa inverter a ordem para cronológica
            grafico.add_series({
                'name': 'Despesas',
                'categories': [worksheet.name, 3, 0, linha - 1, 0],
                'values': [worksheet.name, 3, 1, linha - 1, 1],
                'marker': {'type': 'circle', 'size': 8},
                'line': {'width': 2.5},
            })
            
            # Configurar gráfico
            grafico.set_title({'name': f'Evolução de Despesas - {meio_pagamento}'})
            grafico.set_x_axis({
                'name': 'Mês/Ano',
                'num_font': {'rotation': -45},
            })
            grafico.set_y_axis({'name': 'Valor Total (R$)'})
            grafico.set_style(10)
            grafico.set_size({'width': 720, 'height': 400})
            
            # Inserir gráfico na planilha
            worksheet.insert_chart('D3', grafico)
            
            workbook.close()
            messagebox.showinfo("Exportação", "Relatório exportado com sucesso!")
                
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar para Excel: {e}")

    def gerar_relatorio_categoria(self):
        """Gera um relatório de despesas por categoria"""
        try:
            # Obter todas as categorias
            categorias = self.carregar_categorias()
            
            if not categorias:
                messagebox.showinfo("Relatório", "Não há categorias cadastradas.")
                return
                
            # Criar janela de diálogo
            dialog = tk.Toplevel(self.root)
            dialog.title("Relatório por Categoria")
            dialog.geometry("300x150")
            dialog.resizable(False, False)
            
            # Frame para os controles
            frame = ttk.Frame(dialog, padding="10")
            frame.pack(fill=tk.BOTH, expand=True)
            
            # Combobox para categoria
            ttk.Label(frame, text="Selecione a Categoria:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
            categoria_combo = ttk.Combobox(frame, values=categorias, width=20)
            categoria_combo.current(0)
            categoria_combo.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)
            
            # Função para gerar relatório e fechar diálogo
            def gerar():
                categoria = categoria_combo.get()
                self.mostrar_relatorio_categoria(categoria)
                dialog.destroy()
            
            # Botão para gerar
            ttk.Button(frame, text="Gerar Relatório", command=gerar).grid(row=2, column=0, columnspan=2, pady=20)
            
            # Centralizar janela
            dialog.transient(self.root)
            dialog.wait_visibility()
            dialog.grab_set()
            dialog.focus_set()
            dialog.wait_window()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar relatório: {e}")
    
    def mostrar_relatorio_categoria(self, categoria):
        """Exibe o relatório por categoria selecionada"""
        try:
            # Consultar dados da categoria
            self.cursor.execute("""
                SELECT 
                    strftime('%Y-%m', data_pagamento) as mes_ano,
                    strftime('%m/%Y', data_pagamento) as mes_ano_formatado,
                    SUM(valor) as total_valor
                FROM despesas
                WHERE conta_despesa = ?
                GROUP BY mes_ano
                ORDER BY mes_ano DESC
            """, (categoria,))
            
            resultados = self.cursor.fetchall()
            
            if not resultados:
                messagebox.showinfo("Relatório", f"Nenhuma despesa encontrada para a categoria {categoria}.")
                return
                
            # Calcular total geral
            total_geral = sum(row[2] for row in resultados)
            
            # Criar nova janela para o relatório
            janela_relatorio = tk.Toplevel(self.root)
            janela_relatorio.title(f"Relatório da Categoria: {categoria}")
            janela_relatorio.geometry("800x600")
            
            # Frame principal
            frame_principal = ttk.Frame(janela_relatorio, padding="10")
            frame_principal.pack(fill=tk.BOTH, expand=True)
            
            # Título
            titulo = ttk.Label(frame_principal, 
                              text=f"Relatório de Despesas - Categoria: {categoria}",
                              font=('Arial', 16, 'bold'))
            titulo.pack(pady=10)
            
            # Total geral
            total_label = ttk.Label(frame_principal, 
                                  text=f"Total de Despesas: R$ {total_geral:.2f}".replace('.', ','),
                                  font=('Arial', 12))
            total_label.pack(pady=5)
            
            # Frame com abas
            notebook = ttk.Notebook(frame_principal)
            notebook.pack(fill=tk.BOTH, expand=True, pady=10)
            
            # Aba de tabela
            tab_tabela = ttk.Frame(notebook)
            notebook.add(tab_tabela, text="Tabela")
            
            # Aba de gráfico
            tab_grafico = ttk.Frame(notebook)
            notebook.add(tab_grafico, text="Gráfico")
            
            # Criar tabela na primeira aba
            tree_frame = ttk.Frame(tab_tabela)
            tree_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
            
            # Scrollbar para a tabela
            scrollbar = ttk.Scrollbar(tree_frame)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # Tabela de resultados
            colunas = ('mes_ano', 'total')
            tabela = ttk.Treeview(tree_frame, columns=colunas, show='headings', yscrollcommand=scrollbar.set)
            
            # Cabeçalhos
            tabela.heading('mes_ano', text='Mês/Ano')
            tabela.heading('total', text='Total (R$)')
            
            # Larguras
            tabela.column('mes_ano', width=150, anchor=tk.CENTER)
            tabela.column('total', width=150, anchor=tk.E)
            
            # Preencher tabela
            for row in resultados:
                mes_ano_key, mes_ano_formatado, total = row
                
                tabela.insert('', tk.END, values=(
                    mes_ano_formatado,
                    f"R$ {total:.2f}".replace('.', ',')
                ))
                
            tabela.pack(fill=tk.BOTH, expand=True)
            scrollbar.config(command=tabela.yview)
            
            # Criar gráfico na segunda aba
            figura = plt.Figure(figsize=(7, 5), dpi=100)
            ax = figura.add_subplot(111)
            
            # Extrair dados para o gráfico (invertendo a ordem para cronológica)
            meses = [row[1] for row in reversed(resultados)]
            valores = [row[2] for row in reversed(resultados)]
            
            # Criar gráfico de linha
            ax.plot(meses, valores, marker='o', linestyle='-', color=self.cor_primaria, linewidth=2)
            
            # Adicionar rótulos
            for i, valor in enumerate(valores):
                ax.annotate(f'R${valor:.2f}'.replace('.', ','), 
                          (meses[i], valores[i]), 
                          textcoords="offset points",
                          xytext=(0,10), 
                          ha='center',
                          fontsize=8)
            
            # Estilo do gráfico
            ax.set_title(f'Evolução de Despesas - {categoria}')
            ax.set_xlabel('Mês/Ano')
            ax.set_ylabel('Valor Total (R$)')
            ax.grid(True, linestyle='--', alpha=0.7)
            
            # Rotacionar rótulos do eixo x
            plt.setp(ax.get_xticklabels(), rotation=45, ha='right')
            
            # Ajustar layout
            figura.tight_layout()
            
            # Adicionar o gráfico ao frame
            canvas = FigureCanvasTkAgg(figura, tab_grafico)
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
            
            # Botão para exportar
            ttk.Button(frame_principal, text="Exportar para Excel", 
                    command=lambda: self.exportar_categoria_excel(categoria)).pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exibir relatório: {e}")
    
    def exportar_excel(self):
        """Exporta todas as despesas para um arquivo Excel"""
        try:
            # Verificar se pandas está instalado
            try:
                import pandas as pd
            except ImportError:
                resposta = messagebox.askyesno(
                    "Biblioteca Necessária", 
                    "É necessário instalar a biblioteca pandas para exportar para Excel. Deseja instalar agora?"
                )
                if resposta:
                    import subprocess
                    subprocess.check_call(['pip', 'install', 'pandas', 'openpyxl'])
                    import pandas as pd
                else:
                    return
            
            # Consultar todas as despesas
            self.cursor.execute("""
                SELECT 
                    id, descricao, meio_pagamento, conta_despesa, valor, 
                    num_parcelas, strftime('%d/%m/%Y', data_pagamento) as data_pagamento,
                    strftime('%d/%m/%Y', data_registro) as data_registro
                FROM despesas
                ORDER BY data_pagamento DESC
            """)
            
            # Converter resultado para DataFrame
            colunas = ['ID', 'Descrição', 'Meio de Pagamento', 'Categoria', 
                     'Valor (R$)', 'Parcelas', 'Data de Pagamento', 'Data de Registro']
            
            df = pd.DataFrame(self.cursor.fetchall(), columns=colunas)
            
            # Solicitar local para salvar o arquivo
            from tkinter import filedialog
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Salvar Como"
            )
            
            if not file_path:
                return  # Usuário cancelou
                
            # Exportar para Excel
            df.to_excel(file_path, index=False, sheet_name='Despesas')
            
            messagebox.showinfo("Exportação Concluída", 
                             f"Dados exportados com sucesso para {file_path}")
            
        except Exception as e:
            messagebox.showerror("Erro na Exportação", f"Erro ao exportar dados: {e}")
    
    def exportar_relatorio_excel(self, mes, ano):
        """Exporta o relatório mensal para Excel"""
        try:
            # Calcular início e fim do mês
            primeiro_dia = f"{ano}-{mes:02d}-01"
            ultimo_dia = f"{ano}-{mes:02d}-{calendar.monthrange(ano, mes)[1]}"
            
            # Consultar dados do mês
            self.cursor.execute("""
                SELECT 
                    conta_despesa as Categoria, 
                    SUM(valor) as 'Total Valor',
                    COUNT(*) as 'Quantidade'
                FROM despesas
                WHERE data_pagamento BETWEEN ? AND ?
                GROUP BY conta_despesa
                ORDER BY 'Total Valor' DESC
            """, (primeiro_dia, ultimo_dia))
            
            # Converter para DataFrame
            df_resumo = pd.DataFrame(self.cursor.fetchall(), 
                                   columns=['Categoria', 'Total (R$)', 'Quantidade'])
            
            # Consultar detalhes das despesas
            self.cursor.execute("""
                SELECT 
                    id as ID,
                    descricao as Descrição,
                    meio_pagamento as 'Meio de Pagamento',
                    conta_despesa as Categoria,
                    valor as 'Valor (R$)',
                    num_parcelas as Parcelas,
                    strftime('%d/%m/%Y', data_pagamento) as 'Data de Pagamento'
                FROM despesas
                WHERE data_pagamento BETWEEN ? AND ?
                ORDER BY data_pagamento
            """, (primeiro_dia, ultimo_dia))
            
            df_detalhes = pd.DataFrame(self.cursor.fetchall(), 
                                     columns=['ID', 'Descrição', 'Meio de Pagamento', 'Categoria', 
                                             'Valor (R$)', 'Parcelas', 'Data de Pagamento'])
            
            # Solicitar local para salvar o arquivo
            from tkinter import filedialog
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Salvar Como"
            )
            
            if not file_path:
                return  # Usuário cancelou
                
            # Criar arquivo Excel com múltiplas planilhas
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df_resumo.to_excel(writer, sheet_name='Resumo', index=False)
                df_detalhes.to_excel(writer, sheet_name='Detalhes', index=False)
            
            messagebox.showinfo("Exportação Concluída", 
                             f"Relatório exportado com sucesso para {file_path}")
            
        except Exception as e:
            messagebox.showerror("Erro na Exportação", f"Erro ao exportar relatório: {e}")
    
    def exportar_categoria_excel(self, categoria):
        """Exporta o relatório por categoria para Excel"""
        try:
            # Consultar resumo por mês
            self.cursor.execute("""
                SELECT 
                    strftime('%m/%Y', data_pagamento) as 'Mês/Ano',
                    SUM(valor) as 'Total Valor',
                    COUNT(*) as 'Quantidade'
                FROM despesas
                WHERE conta_despesa = ?
                GROUP BY strftime('%Y-%m', data_pagamento)
                ORDER BY strftime('%Y-%m', data_pagamento) DESC
            """, (categoria,))
            
            # Converter para DataFrame
            df_resumo = pd.DataFrame(self.cursor.fetchall(), 
                                   columns=['Mês/Ano', 'Total (R$)', 'Quantidade'])
            
            # Consultar detalhes das despesas
            self.cursor.execute("""
                SELECT 
                    id as ID,
                    descricao as Descrição,
                    meio_pagamento as 'Meio de Pagamento',
                    valor as 'Valor (R$)',
                    num_parcelas as Parcelas,
                    strftime('%d/%m/%Y', data_pagamento) as 'Data de Pagamento'
                FROM despesas
                WHERE conta_despesa = ?
                ORDER BY data_pagamento DESC
            """, (categoria,))
            
            df_detalhes = pd.DataFrame(self.cursor.fetchall(), 
                                     columns=['ID', 'Descrição', 'Meio de Pagamento', 
                                             'Valor (R$)', 'Parcelas', 'Data de Pagamento'])
            
            # Solicitar local para salvar o arquivo
            from tkinter import filedialog
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Salvar Como"
            )
            
            if not file_path:
                return  # Usuário cancelou
                
            # Criar arquivo Excel com múltiplas planilhas
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df_resumo.to_excel(writer, sheet_name=f'Resumo - {categoria}', index=False)
                df_detalhes.to_excel(writer, sheet_name='Detalhes', index=False)
            
            messagebox.showinfo("Exportação Concluída", 
                             f"Relatório exportado com sucesso para {file_path}")
            
        except Exception as e:
            messagebox.showerror("Erro na Exportação", f"Erro ao exportar relatório: {e}")

# Função principal
def main():
    root = tk.Tk()
    app = SistemaFinanceiro(root)
    root.mainloop()

if __name__ == "__main__":
    main()