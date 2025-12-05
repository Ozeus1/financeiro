import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import sqlite3
from datetime import datetime
import calendar
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.ticker import FuncFormatter
import os
import locale
# ### CORREÇÃO 1: Importar a classe Calendar além da DateEntry ###
from tkcalendar import DateEntry, Calendar
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from tkinter import filedialog
from werkzeug.security import generate_password_hash, check_password_hash

class SessaoUsuario:
    def __init__(self):
        self.user_id = None
        self.username = None
        self.nivel_acesso = None
        self.autenticado = False
    
    def login(self, user_id, username, nivel_acesso):
        self.user_id = user_id
        self.username = username
        self.nivel_acesso = nivel_acesso
        self.autenticado = True
    
    def logout(self):
        self.user_id = None
        self.username = None
        self.nivel_acesso = None
        self.autenticado = False


# Tenta importar módulos locais, mas não quebra se não existirem
try:
    import configuracao
    import MENUBD
    import relclaude1
    import relatorios1
    import gerenciar_orcamento
    import relatorio_orcado_vs_gasto
    import gerenciar_fechamento_cartoes
    import relatorio_previsao_faturas
    import importador_excel
    import relatorio_balanco # <--- ADICIONE ESTA LINHA
    import relatorio_balanco_fluxo_caixa
    import importador_supabase # <--- ADICIONE ESTA LINHA


except ImportError:
    print("Aviso: Alguns módulos de extensão não foram encontrados. A funcionalidade completa pode estar limitada.")
    class PlaceholderModule:
        def __getattr__(self, name):
            def placeholder_func(*args, **kwargs):
                messagebox.showerror("Módulo Faltando", f"O módulo para esta função não foi encontrado.")
            return placeholder_func
    configuracao = MENUBD = relclaude1 = relatorios1 = gerenciar_orcamento = \
    relatorio_orcado_vs_gasto = gerenciar_fechamento_cartoes = \
    relatorio_previsao_faturas = importador_excel = PlaceholderModule()

# Configurar a localização para formato de moeda brasileiro
try:
    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
except locale.Error:
    try:
        locale.setlocale(locale.LC_ALL, 'Portuguese_Brazil.1252')
    except locale.Error:
        print("Aviso: Locale 'pt_BR' não encontrado. A formatação de moeda pode não funcionar corretamente.")


# --- FUNÇÕES PARA CONVERSÃO DE VALORES COM VÍRGULA ---
def converter_para_float(valor_str):
    """Converte string com vírgula ou ponto para float"""
    try:
        # Remove espaços e substitui vírgula por ponto
        valor_limpo = str(valor_str).strip().replace(',', '.')
        return float(valor_limpo)
    except (ValueError, AttributeError):
        return 0.0

def validar_entrada_numerica(novo_valor):
    """Valida entrada numérica permitindo números, vírgula e ponto"""
    if novo_valor == "":
        return True
    # Permite números, vírgula, ponto e sinal negativo
    if all(c in '0123456789.,-' for c in novo_valor):
        # Verifica se não há mais de uma vírgula ou ponto
        if novo_valor.count(',') <= 1 and novo_valor.count('.') <= 1:
            return True
    return False


# --- CLASSE PARA GERENCIAR CATEGORIAS DE RECEITA ---
class GerenciadorCategoriasReceita(tk.Toplevel):
    """Janela para gerenciar (adicionar, editar, excluir) categorias de receita."""
    def __init__(self, parent_widget, app_logic):
        super().__init__(parent_widget)
        self.transient(parent_widget)
        self.grab_set()
        self.title("Gerenciar Categorias de Receita")
        self.geometry("450x400")

        # Conexão com o banco de dados de RECEITAS
        self.conn_receitas = app_logic.conn_receitas
        self.cursor_receitas = app_logic.cursor_receitas

        # Variáveis de controle
        self.id_categoria_var = tk.StringVar()
        self.nome_categoria_var = tk.StringVar()

        # Frame principal
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Frame da lista
        frame_lista = ttk.LabelFrame(main_frame, text="Categorias Existentes")
        frame_lista.pack(fill=tk.BOTH, expand=True, pady=5)
        
        self.lista_categorias = tk.Listbox(frame_lista, height=10)
        self.lista_categorias.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.lista_categorias.bind("<<ListboxSelect>>", self.selecionar_categoria)

        scrollbar = ttk.Scrollbar(frame_lista, orient="vertical", command=self.lista_categorias.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.lista_categorias.config(yscrollcommand=scrollbar.set)
        
        # Frame do formulário e botões
        frame_form = ttk.LabelFrame(main_frame, text="Adicionar / Editar Categoria")
        frame_form.pack(fill=tk.X, pady=10)

        ttk.Label(frame_form, text="ID:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        ttk.Entry(frame_form, textvariable=self.id_categoria_var, state='readonly', width=10).grid(row=0, column=1, padx=5, pady=5, sticky="w")
        
        ttk.Label(frame_form, text="Nome:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        entry_categoria = ttk.Entry(frame_form, textvariable=self.nome_categoria_var, width=40)
        entry_categoria.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
        frame_form.columnconfigure(1, weight=1)

        frame_botoes = ttk.Frame(main_frame)
        frame_botoes.pack(fill=tk.X, pady=5)
        
        btn_adicionar = ttk.Button(frame_botoes, text="Adicionar Novo", command=self.adicionar_categoria)
        btn_adicionar.pack(side=tk.LEFT, padx=5)

        self.btn_atualizar = ttk.Button(frame_botoes, text="Salvar Edição", command=self.atualizar_categoria, state="disabled")
        self.btn_atualizar.pack(side=tk.LEFT, padx=5)

        self.btn_excluir = ttk.Button(frame_botoes, text="Excluir Selecionado", command=self.excluir_categoria, state="disabled")
        self.btn_excluir.pack(side=tk.LEFT, padx=5)
        
        btn_limpar = ttk.Button(frame_botoes, text="Limpar Formulário", command=self.limpar_campos)
        btn_limpar.pack(side=tk.LEFT, padx=5)
        
        # Carrega os dados e ativa a janela
        self.carregar_categorias()

    def carregar_categorias(self):
        """Carrega a lista e seleciona o primeiro item para edição."""
        # Limpa a seleção e o formulário antes de carregar
        self.limpar_campos()
        self.lista_categorias.delete(0, tk.END)
        
        # Busca e insere as categorias na lista
        self.cursor_receitas.execute("SELECT nome FROM categorias_receita ORDER BY nome")
        categorias = self.cursor_receitas.fetchall()
        for categoria in categorias:
            self.lista_categorias.insert(tk.END, categoria[0])
            
        # NOVA LÓGICA: Seleciona o primeiro item da lista, se houver algum
        if self.lista_categorias.size() > 0:
            self.lista_categorias.selection_set(0)
            self.selecionar_categoria() # Chama o evento para popular o formulário

    def limpar_campos(self):
        """Limpa o formulário e desativa os botões de ação."""
        self.id_categoria_var.set("")
        self.nome_categoria_var.set("")
        if self.lista_categorias.curselection():
            self.lista_categorias.selection_clear(0, tk.END)
        self.btn_atualizar.config(state="disabled")
        self.btn_excluir.config(state="disabled")

    def selecionar_categoria(self, event=None):
        """Preenche o formulário quando um item da lista é selecionado."""
        selecionado_indices = self.lista_categorias.curselection()
        if not selecionado_indices:
            # Se não houver seleção, limpa os campos.
            self.id_categoria_var.set("")
            self.nome_categoria_var.set("")
            self.btn_atualizar.config(state="disabled")
            self.btn_excluir.config(state="disabled")
            return
        
        nome_selecionado = self.lista_categorias.get(selecionado_indices[0])
        self.cursor_receitas.execute("SELECT id FROM categorias_receita WHERE nome = ?", (nome_selecionado,))
        resultado = self.cursor_receitas.fetchone()
        
        if resultado:
            self.id_categoria_var.set(resultado[0])
            self.nome_categoria_var.set(nome_selecionado)
            self.btn_atualizar.config(state="normal")
            self.btn_excluir.config(state="normal")

    def adicionar_categoria(self):
        """Adiciona uma nova categoria com base no texto do formulário."""
        nome = self.nome_categoria_var.get().strip()
        if not nome:
            messagebox.showerror("Erro", "O nome da categoria não pode estar vazio.", parent=self)
            return
        # Limpa o campo de ID para garantir que não seja uma atualização
        self.id_categoria_var.set("")
        try:
            self.cursor_receitas.execute("INSERT INTO categorias_receita (nome) VALUES (?)", (nome,))
            self.conn_receitas.commit()
            self.carregar_categorias()
        except sqlite3.IntegrityError:
            messagebox.showerror("Erro", f"A categoria '{nome}' já existe.", parent=self)
        except Exception as e:
            messagebox.showerror("Erro no Banco de Dados", f"Erro ao adicionar categoria: {e}", parent=self)

    def atualizar_categoria(self):
        """Salva as edições feitas no nome da categoria selecionada."""
        id_cat = self.id_categoria_var.get()
        novo_nome = self.nome_categoria_var.get().strip()

        if not id_cat or not novo_nome:
            messagebox.showwarning("Dados Incompletos", "Selecione uma categoria da lista para editar.", parent=self)
            return
        
        try:
            self.cursor_receitas.execute("SELECT nome FROM categorias_receita WHERE id = ?", (id_cat,))
            resultado = self.cursor_receitas.fetchone()
            if not resultado:
                messagebox.showerror("Erro", "Categoria de receita não encontrada.", parent=self)
                return
            nome_antigo = resultado[0]

            if nome_antigo == novo_nome:
                messagebox.showinfo("Informação", "O nome não foi alterado.", parent=self)
                return

            self.cursor_receitas.execute("UPDATE categorias_receita SET nome = ? WHERE id = ?", (novo_nome, id_cat))
            self.cursor_receitas.execute("UPDATE receitas SET conta_receita = ? WHERE conta_receita = ?", (novo_nome, nome_antigo))
            
            self.conn_receitas.commit()
            
            messagebox.showinfo("Sucesso", f"Categoria '{nome_antigo}' foi atualizada para '{novo_nome}' com sucesso.", parent=self)
            # Recarrega a lista e seleciona o item recém-editado
            self.carregar_categorias()
            # Encontrar o novo item na lista e selecioná-lo
            items = self.lista_categorias.get(0, tk.END)
            if novo_nome in items:
                idx = items.index(novo_nome)
                self.lista_categorias.selection_set(idx)
                self.selecionar_categoria()

        except sqlite3.IntegrityError:
            messagebox.showerror("Erro de Integridade", f"O nome '{novo_nome}' já está em uso.", parent=self)
            self.conn_receitas.rollback()
        except Exception as e:
            messagebox.showerror("Erro ao Atualizar", f"Erro no banco de dados: {e}", parent=self)
            self.conn_receitas.rollback()

    def excluir_categoria(self):
        """Exclui a categoria selecionada, se não estiver em uso."""
        id_cat = self.id_categoria_var.get()
        nome_cat = self.nome_categoria_var.get()

        if not id_cat:
            messagebox.showerror("Erro", "Selecione uma categoria da lista para excluir.", parent=self)
            return

        self.cursor_receitas.execute("SELECT COUNT(*) FROM receitas WHERE conta_receita = ?", (nome_cat,))
        count = self.cursor_receitas.fetchone()[0]

        if count > 0:
            messagebox.showwarning("Impossível Excluir", 
                                 f"A categoria '{nome_cat}' está sendo usada em {count} registros de receita e não pode ser excluída.",
                                 parent=self)
            return

        if messagebox.askyesno("Confirmar Exclusão", f"Tem certeza que deseja excluir a categoria '{nome_cat}'?", parent=self):
            try:
                self.cursor_receitas.execute("DELETE FROM categorias_receita WHERE id = ?", (id_cat,))
                self.conn_receitas.commit()
                self.carregar_categorias()
            except Exception as e:
                messagebox.showerror("Erro no Banco de Dados", f"Erro ao excluir categoria: {e}", parent=self)




###################################################
  

# --- CLASSE PARA GERENCIAR RECEITAS ---
class GerenciadorReceitas(tk.Toplevel):
    """Janela para gerenciar (cadastrar, editar, excluir) receitas."""
    def __init__(self, parent_widget, app_logic):
        super().__init__(parent_widget)
        self.transient(parent_widget)
        self.grab_set()
        self.title("Gerenciador de Receitas")
        self.geometry("950x600")

        self.parent_app = app_logic
        self.conn_receitas = self.parent_app.conn_receitas
        self.cursor_receitas = self.parent_app.cursor_receitas

        self.id_receita = tk.StringVar()
        self.descricao = tk.StringVar()
        self.meio_recebimento = tk.StringVar()
        self.conta_receita = tk.StringVar()
        self.valor = tk.StringVar()
        
        self.criar_widgets()
        self.carregar_receitas()

    def carregar_combos(self):
        self.cursor_receitas.execute("SELECT nome FROM categorias_receita ORDER BY nome")
        self.combo_conta_receita['values'] = [row[0] for row in self.cursor_receitas.fetchall()]
        
        self.cursor_receitas.execute("SELECT nome FROM meios_recebimento ORDER BY nome")
        self.combo_meio_recebimento['values'] = [row[0] for row in self.cursor_receitas.fetchall()]

    def criar_widgets(self):
        frame_principal = ttk.Frame(self, padding="10")
        frame_principal.pack(fill=tk.BOTH, expand=True)

        frame_form = ttk.LabelFrame(frame_principal, text="Lançamento de Receita")
        frame_form.pack(fill=tk.X, pady=10)
        frame_form.columnconfigure(1, weight=1)

        ttk.Label(frame_form, text="Descrição:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        ttk.Entry(frame_form, textvariable=self.descricao, width=40).grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(frame_form, text="Valor (R$):").grid(row=0, column=2, padx=5, pady=5, sticky="w")
        vcmd = (self.register(validar_entrada_numerica), '%P')
        ttk.Entry(frame_form, textvariable=self.valor, width=15, validate='key', validatecommand=vcmd).grid(row=0, column=3, padx=5, pady=5, sticky="w")

        ttk.Label(frame_form, text="Categoria:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.combo_conta_receita = ttk.Combobox(frame_form, textvariable=self.conta_receita, state='readonly')
        self.combo_conta_receita.grid(row=1, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(frame_form, text="Meio Recebimento:").grid(row=1, column=2, padx=5, pady=5, sticky="w")
        self.combo_meio_recebimento = ttk.Combobox(frame_form, textvariable=self.meio_recebimento, width=20, state='readonly')
        self.combo_meio_recebimento.grid(row=1, column=3, padx=5, pady=5, sticky="w")
        
        ttk.Label(frame_form, text="Data Recebimento:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.data_entry = DateEntry(frame_form, width=12, date_pattern='dd/mm/yyyy', locale='pt_BR')
        self.data_entry.grid(row=2, column=1, padx=5, pady=5, sticky="w")
        self.data_entry.set_date(datetime.now())

        self.carregar_combos()

        frame_botoes = ttk.Frame(frame_form)
        frame_botoes.grid(row=3, column=0, columnspan=4, pady=10)
        ttk.Button(frame_botoes, text="Salvar", command=self.salvar_receita).pack(side=tk.LEFT, padx=5)
        self.btn_atualizar = ttk.Button(frame_botoes, text="Atualizar", command=self.atualizar_receita, state="disabled")
        self.btn_atualizar.pack(side=tk.LEFT, padx=5)
        self.btn_excluir = ttk.Button(frame_botoes, text="Excluir", command=self.excluir_receita, state="disabled")
        self.btn_excluir.pack(side=tk.LEFT, padx=5)
        ttk.Button(frame_botoes, text="Limpar", command=self.limpar_campos).pack(side=tk.LEFT, padx=5)

        frame_tabela = ttk.LabelFrame(frame_principal, text="Receitas Lançadas")
        frame_tabela.pack(fill=tk.BOTH, expand=True, pady=10)
        
        scrollbar = ttk.Scrollbar(frame_tabela)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tabela = ttk.Treeview(frame_tabela, yscrollcommand=scrollbar.set, selectmode="browse")
        self.tabela['columns'] = ('id', 'data', 'descricao', 'categoria', 'meio', 'valor')
        
        self.tabela.column("#0", width=0, stretch=tk.NO)
        self.tabela.column("id", anchor=tk.CENTER, width=40)
        self.tabela.column("data", anchor=tk.CENTER, width=100)
        self.tabela.column("descricao", anchor=tk.W, width=300)
        self.tabela.column("categoria", anchor=tk.W, width=150)
        self.tabela.column("meio", anchor=tk.W, width=150)
        self.tabela.column("valor", anchor=tk.E, width=120)

        self.tabela.heading("id", text="ID")
        self.tabela.heading("data", text="Data")
        self.tabela.heading("descricao", text="Descrição")
        self.tabela.heading("categoria", text="Categoria")
        self.tabela.heading("meio", text="Meio")
        self.tabela.heading("valor", text="Valor")
        
        self.tabela.pack(fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.tabela.yview)
        self.tabela.bind("<ButtonRelease-1>", self.selecionar_item)

    def carregar_receitas(self):
        for i in self.tabela.get_children():
            self.tabela.delete(i)
        self.cursor_receitas.execute("""
            SELECT id, strftime('%d/%m/%Y', data_recebimento), descricao, conta_receita, meio_recebimento, valor
            FROM receitas ORDER BY data_recebimento DESC
        """)
        for row in self.cursor_receitas.fetchall():
            valor_formatado = locale.currency(row[5], grouping=True, symbol=True)
            self.tabela.insert("", tk.END, values=(row[0], row[1], row[2], row[3], row[4], valor_formatado))

    def limpar_campos(self):
        self.id_receita.set("")
        self.descricao.set("")
        self.conta_receita.set("")
        self.meio_recebimento.set("")
        self.valor.set(0.0)
        self.data_entry.set_date(datetime.now())
        self.btn_atualizar.config(state="disabled")
        self.btn_excluir.config(state="disabled")
        if self.tabela.selection():
            self.tabela.selection_remove(self.tabela.selection())

    def salvar_receita(self):
        valor_convertido = converter_para_float(self.valor.get())
        if not self.descricao.get() or valor_convertido <= 0 or not self.conta_receita.get():
            messagebox.showerror("Erro de Validação", "Preencha todos os campos obrigatórios (Descrição, Valor > 0, Categoria).", parent=self)
            return

        try:
            data_recebimento = self.data_entry.get_date().strftime('%Y-%m-%d')
            self.cursor_receitas.execute("""
                INSERT INTO receitas (descricao, conta_receita, meio_recebimento, valor, data_recebimento, data_registro, num_parcelas)
                VALUES (?, ?, ?, ?, ?, date('now'), 1)
            """, (self.descricao.get(), self.conta_receita.get(), self.meio_recebimento.get(), valor_convertido, data_recebimento))
            self.conn_receitas.commit()
            messagebox.showinfo("Sucesso", "Receita salva com sucesso!", parent=self)
            self.limpar_campos()
            self.carregar_receitas()
        except Exception as e:
            messagebox.showerror("Erro no Banco de Dados", f"Erro ao salvar receita: {e}", parent=self)

    def selecionar_item(self, event):
        selected_item = self.tabela.focus()
        if not selected_item: return
        
        values = self.tabela.item(selected_item, 'values')
        self.id_receita.set(values[0])
        
        data_obj = datetime.strptime(values[1], '%d/%m/%Y')
        self.data_entry.set_date(data_obj)
        
        self.descricao.set(values[2])
        self.conta_receita.set(values[3])
        self.meio_recebimento.set(values[4])
        
        try:
            valor_str = values[5].replace('R$', '').replace('.', '').replace(',', '.').strip()
            self.valor.set(float(valor_str))
        except (ValueError, IndexError):
             self.valor.set(0.0)

        self.btn_atualizar.config(state="normal")
        self.btn_excluir.config(state="normal")

    def atualizar_receita(self):
        receita_id = self.id_receita.get()
        if not receita_id: return

        valor_convertido = converter_para_float(self.valor.get())
        if not self.descricao.get() or valor_convertido <= 0 or not self.conta_receita.get():
            messagebox.showerror("Erro de Validação", "Preencha todos os campos obrigatórios.", parent=self)
            return

        try:
            data_recebimento = self.data_entry.get_date().strftime('%Y-%m-%d')
            self.cursor_receitas.execute("""
                UPDATE receitas SET descricao=?, conta_receita=?, meio_recebimento=?, valor=?, data_recebimento=?
                WHERE id=?
            """, (self.descricao.get(), self.conta_receita.get(), self.meio_recebimento.get(), valor_convertido, data_recebimento, receita_id))
            self.conn_receitas.commit()
            messagebox.showinfo("Sucesso", "Receita atualizada com sucesso!", parent=self)
            self.limpar_campos()
            self.carregar_receitas()
        except Exception as e:
            messagebox.showerror("Erro no Banco de Dados", f"Erro ao atualizar receita: {e}", parent=self)

    def excluir_receita(self):
        receita_id = self.id_receita.get()
        if not receita_id: return

        if messagebox.askyesno("Confirmar", "Tem certeza que deseja excluir esta receita?", parent=self):
            try:
                self.cursor_receitas.execute("DELETE FROM receitas WHERE id=?", (receita_id,))
                self.conn_receitas.commit()
                messagebox.showinfo("Sucesso", "Receita excluída com sucesso!", parent=self)
                self.limpar_campos()
                self.carregar_receitas()
            except Exception as e:
                messagebox.showerror("Erro no Banco de Dados", f"Erro ao excluir receita: {e}", parent=self)


# --- CLASSE PRINCIPAL DA APLICAÇÃO ---
# --- DIÁLOGO DE LOGIN ---
class LoginDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Login - Sistema Financeiro")
        self.geometry("300x250")
        self.resizable(False, False)
        self.parent = parent
        self.sessao = None
        
        # Centralizar na tela
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry('{}x{}+{}+{}'.format(width, height, x, y))
        
        self.configure(bg='#f0f0f0')
        
        # Estilo
        style = ttk.Style()
        style.configure('Login.TLabel', background='#f0f0f0', font=('Arial', 10))
        style.configure('Login.TButton', font=('Arial', 10, 'bold'))
        
        # Widgets
        frame = ttk.Frame(self, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Usuário:", style='Login.TLabel').pack(anchor=tk.W, pady=(0, 5))
        self.user_entry = ttk.Entry(frame)
        self.user_entry.pack(fill=tk.X, pady=(0, 15))
        self.user_entry.focus()
        
        ttk.Label(frame, text="Senha:", style='Login.TLabel').pack(anchor=tk.W, pady=(0, 5))
        self.pass_entry = ttk.Entry(frame, show="*")
        self.pass_entry.pack(fill=tk.X, pady=(0, 20))
        self.pass_entry.bind('<Return>', self.fazer_login)
        
        self.btn_login = ttk.Button(frame, text="Entrar", command=self.fazer_login, style='Login.TButton')
        self.btn_login.pack(fill=tk.X, pady=5)
        
        self.lbl_status = ttk.Label(frame, text="", foreground="red", style='Login.TLabel', wraplength=250)
        self.lbl_status.pack(pady=10)
        
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        self.transient(parent)
        self.grab_set()
        self.wait_window(self)
        
    def fazer_login(self, event=None):
        username = self.user_entry.get().strip()
        password = self.pass_entry.get().strip()
        
        if not username or not password:
            self.lbl_status.config(text="Preencha usuário e senha.")
            return
            
        try:
            conn = sqlite3.connect('financas.db')
            cursor = conn.cursor()
            
            cursor.execute("SELECT id, username, password_hash, nivel_acesso FROM users WHERE username = ?", (username,))
            user = cursor.fetchone()
            
            conn.close()
            
            if user and check_password_hash(user[2], password):
                self.sessao = SessaoUsuario()
                self.sessao.login(user[0], user[1], user[3])
                self.destroy()
            else:
                self.lbl_status.config(text="Usuário ou senha incorretos.")
                
        except sqlite3.Error as e:
            self.lbl_status.config(text=f"Erro de conexão: {e}")

    def on_close(self):
        self.destroy()

# --- CLASSE PRINCIPAL DA APLICAÇÃO ---
class SistemaFinanceiro:
    def __init__(self, root, sessao=None):
        self.root = root
        self.sessao = sessao if sessao else SessaoUsuario()
        self.root.title("Sistema de Gerenciamento Financeiro")
        self.root.geometry("1270x750")
        self.root.resizable(True, True)
        
        self.menu_bar = tk.Menu(self.root)
        self.root.config(menu=self.menu_bar)

        arquivo_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.menu_bar.add_cascade(label="Arquivo", menu=arquivo_menu)
        arquivo_menu.add_command(label="Importar de Planilha Excel...", command=self.abrir_importador_excel)
        arquivo_menu.add_command(label="Importar do Supabase...", command=self.abrir_importador_supabase) # <--- ADICIONE ESTA LINHA
        arquivo_menu.add_command(label="Atualizar Interface", command=self.atualizar_dados_interface)
        # Adiciona os submenus no menu Configuração
         # Submenu Cadastro
        cadastro_submenu = tk.Menu(arquivo_menu, tearoff=0)
        cadastro_submenu.add_command(label="Orçamento por Categoria", command=self.abrir_gerenciador_orcamento) 
        cadastro_submenu.add_command(label="Data de Fechamento de Cartões", command=self.abrir_gerenciador_fechamento_cartoes)
        cadastro_submenu.add_command(label="Pagamentos e Receitas (Fluxo Financeiro)", command=self.abrir_relatorio_balancofc)

        # Submenu Dados
        dados_submenu = tk.Menu(arquivo_menu, tearoff=0)
        dados_submenu.add_command(label="Categorias e Pagamentos", command=self.abrir_gerenciador)
        dados_submenu.add_command(label="Banco de Dados", command=self.abrir_gerenciador2)
        dados_submenu.add_command(label="Contas de Receita", command=self.abrir_gerenciador_categorias_receita)


        arquivo_menu.add_cascade(label="Cadastro", menu=cadastro_submenu)
        arquivo_menu.add_cascade(label="Gerenciar", menu=dados_submenu)

        arquivo_menu.add_separator()
        arquivo_menu.add_command(label="Sair", command=self.on_closing)

        # Menu Despesa
   

        relatorios_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.menu_bar.add_cascade(label="Despesas", menu=relatorios_menu)
        relatorios_menu.add_command(label="Top 10 Contas (Gráfico Dinâmico)", command=self.mostrar_grafico_principais_contas)
        relatorios_menu.add_command(label="Por Categoria (Evolução)", command=self.gerar_relatorio_categoria)
        relatorios_menu.add_command(label="Por Meio de Pagamento (Evolução)", command=self.gerar_relatorio_meio_pagamento)
        relatorios_menu.add_separator()
        relatorios_menu.add_command(label="Relatórios Avançados com Gráficos", command=self.abrir_relatorios_avancados)
        relatorios_menu.add_command(label="Relatórios Simples (Listagem)", command=self.abrir_relatorios_simples)
        relatorios_menu.add_command(label="Relatório Entre Datas", command=self.gerar_relatorio_entre_datas)
        relatorios_menu.add_command(label="Relatório Mensal por Período", command=self.gerar_relatorio_mensal_periodo)
      
             
        receitas_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.menu_bar.add_cascade(label="Receitas", menu=receitas_menu)
        receitas_menu.add_command(label="Lançar/Gerenciar Receitas", command=self.abrir_gerenciador_receitas)
        receitas_menu.add_command(label="Gerenciar Categorias de Receita", command=self.abrir_gerenciador_categorias_receita)
        receitas_menu.add_separator()
        receitas_menu.add_command(label="Relatório Mensal de Receitas", command=self.gerar_relatorio_receita_mensal)
        receitas_menu.add_command(label="Relatório de Receitas por Categoria", command=self.gerar_relatorio_receita_categoria)
        receitas_menu.add_command(label="Relatório de Receitas Entre Datas", command=self.gerar_relatorio_receita_entre_datas)

        # ==================================================================
        # Novo Menu "Análise Financeira"
        consolidado_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.menu_bar.add_cascade(label="Análise Financeira", menu=consolidado_menu)
        consolidado_menu.add_command(label="Balanço Mensal (Receita x Despesa)", command=self.abrir_relatorio_balanco)
        consolidado_menu.add_command(label="Balanço Mensal (Fluxo Financeiro)", command=self.abrir_relatorio_balancofc)

        # ==================================================================
        # Novo Menu "Cartões de Crédito"
        cartoes_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.menu_bar.add_cascade(label="Cartões de Crédito", menu=cartoes_menu)
        cartoes_menu.add_command(label="Configurar Fechamento de Cartões", command=self.abrir_gerenciador_fechamento_cartoes)
        cartoes_menu.add_command(label="Previsão de Faturas de Cartão", command=self.abrir_relatorio_previsao_faturas_cartao)

 # ==================================================================
        # Novo Menu "Orçamento"
        orcamento_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.menu_bar.add_cascade(label="Orçamento", menu=orcamento_menu)
        orcamento_menu.add_command(label="Gerenciar Orçamento", command=self.abrir_gerenciador_orcamento) 
        orcamento_menu.add_command(label="Mensal: Orçado x Gasto", command=self.abrir_relatorio_orcado_vs_gasto)



        ajuda_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.menu_bar.add_cascade(label="Ajuda", menu=ajuda_menu)
        ajuda_menu.add_command(label="Ajuda", command=self.mostrar_ajuda)
        ajuda_menu.add_command(label="Sobre...", command=self.mostrar_sobre)
        
        self.cor_primaria = "#4CAF50"
        self.cor_secundaria = "#F0F4C3"
        self.cor_fundo = "#F9F9F9"
        self.cor_texto = "#333333"
        
        self.configurar_estilo()
        
        self.criar_banco_dados()
        self.criar_banco_dados_receitas()
        
        self.id_despesa = tk.StringVar()
        self.descricao = tk.StringVar()
        self.meio_pagamento = tk.StringVar()
        self.conta_despesa = tk.StringVar()
        self.valor = tk.StringVar()
        self.num_parcelas = tk.IntVar()
        
        # INÍCIO DA MODIFICAÇÃO: Adiciona variável para mapear meses do gráfico
        self.mapa_meses_grafico = {}
        self.total_mes_selecionado_var = tk.StringVar()
        # FIM DA MODIFICAÇÃO

        # === INÍCIO NOVA ALTERAÇÃO: Variáveis para ordenação e visibilidade da tabela ===
        self._sort_state = {'col': 'data', 'reverse': True} # Ordenação padrão
        self._all_columns = {
            'id': ('ID', tk.BooleanVar(value=True)),
            'descricao': ('Descrição', tk.BooleanVar(value=True)),
            'meio_pagamento': ('Meio Pagamento', tk.BooleanVar(value=True)),
            'conta_despesa': ('Categoria', tk.BooleanVar(value=True)),
            'valor': ('Valor (R$)', tk.BooleanVar(value=True)),
            'parcelas': ('Parcelas', tk.BooleanVar(value=True)),
            'data': ('Data', tk.BooleanVar(value=True)),
        }
        # === FIM NOVA ALTERAÇÃO ===
        
        self.criar_widgets()
        self.carregar_despesas()

    def on_closing(self):
        if messagebox.askokcancel("Sair", "Deseja sair do sistema?"):
            self.conn.close()
            self.conn_receitas.close()
            self.root.destroy()
            
   
    def abrir_importador_supabase(self):
        """Abre a janela para importar despesas do Supabase."""
        # The SupabaseImporter needs access to the main app's SQLite connection and cursor
        importador_supabase.iniciar_importador_supabase(self.root, self)
        # After closing the importer, refresh the main display to show new expenses
        self.carregar_despesas()

    
    
    def criar_banco_dados_receitas1(self):
        try:
            self.conn_receitas = sqlite3.connect('financas_receitas.db')
            self.cursor_receitas = self.conn_receitas.cursor()
            
            self.cursor_receitas.execute('''
                CREATE TABLE IF NOT EXISTS receitas (
                    id INTEGER PRIMARY KEY, descricao TEXT NOT NULL, meio_recebimento TEXT,
                    conta_receita TEXT NOT NULL, valor REAL NOT NULL, num_parcelas INTEGER DEFAULT 1,
                    data_registro DATE NOT NULL, data_recebimento DATE NOT NULL)''')
            self.cursor_receitas.execute('CREATE TABLE IF NOT EXISTS categorias_receita (id INTEGER PRIMARY KEY, nome TEXT NOT NULL UNIQUE)')
            self.cursor_receitas.execute('CREATE TABLE IF NOT EXISTS meios_recebimento (id INTEGER PRIMARY KEY, nome TEXT NOT NULL UNIQUE)')

            cat_receita_padrao = [('Salário',), ('Vendas',), ('Rendimentos',), ('Freelance',), ('Outras Receitas',)]
            meios_receita_padrao = [('Transferência Bancária',), ('PIX',), ('Dinheiro',), ('Cheque',)]
    def mostrar_grafico_principais_contas(self):
        """
        Exibe um gráfico das 10 principais contas de despesa, com cores e interatividade.
        Permite selecionar diferentes meses e clicar nas barras para ver detalhes.
        """
        try:
            # Importações necessárias (caso não estejam no topo do arquivo)
            from matplotlib.figure import Figure
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            from matplotlib.ticker import FuncFormatter
            import matplotlib.pyplot as plt
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            from openpyxl.utils import get_column_letter
            from tkinter import filedialog
            
            # Obter mês atual como padrão
            data_atual = datetime.now()
            mes_atual_ym = data_atual.strftime("%Y-%m")
            mes_atual_formatado = data_atual.strftime("%m/%Y")

            # Criar janela para o gráfico
            janela_grafico = tk.Toplevel(self.root)
            janela_grafico.title(f"Top 10 Contas - {mes_atual_formatado}")
            janela_grafico.geometry("900x800")
            
            # Frame principal
            frame_principal = ttk.Frame(janela_grafico, padding="10")
            frame_principal.pack(fill=tk.BOTH, expand=True)

            # --- Função aninhada para mostrar os detalhes da transação ---
            def mostrar_detalhes_transacao(categoria, mes_ym_selecionado):
                try:
                    # Criar a janela de detalhes
                    detalhes_window = tk.Toplevel(janela_grafico)
                    mes_formatado_titulo = datetime.strptime(mes_ym_selecionado, "%Y-%m").strftime("%m/%Y")
                    detalhes_window.title(f"Detalhes para '{categoria}' em {mes_formatado_titulo}")
                    detalhes_window.geometry("600x400")
                    detalhes_window.transient(janela_grafico)
                    detalhes_window.grab_set()

                    # Frame para o Treeview
                    tree_frame = ttk.Frame(detalhes_window, padding="10")
                    tree_frame.pack(fill=tk.BOTH, expand=True)

                    # Consultar as transações
                    self.cursor.execute("""
                        SELECT strftime('%d/%m/%Y', data_pagamento), descricao, valor
                        FROM v_despesas_compat
                        WHERE conta_despesa = ? AND strftime('%Y-%m', data_pagamento) = ?
                        AND user_id = ?
                        ORDER BY data_pagamento
                    """, (categoria, mes_ym_selecionado, self.sessao.user_id))
                    transacoes = self.cursor.fetchall()

                    # Treeview para exibir as transações
                    cols = ('data', 'descricao', 'valor')
                    tree = ttk.Treeview(tree_frame, columns=cols, show='headings')
                    tree.heading('data', text='Data')
                    tree.heading('descricao', text='Descrição')
                    tree.heading('valor', text='Valor (R$)')
                    tree.column('data', width=100, anchor=tk.CENTER)
                    tree.column('descricao', width=350)
                    tree.column('valor', width=120, anchor=tk.E)
                    
                    # Adicionar scrollbar
                    scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
                    tree.configure(yscrollcommand=scrollbar.set)
                    scrollbar.pack(side='right', fill='y')
                    tree.pack(side='left', fill='both', expand=True)


                    total_categoria = 0
                    for data, desc, valor in transacoes:
                        valor_formatado = locale.currency(valor, grouping=True)
                        tree.insert("", "end", values=(data, desc, valor_formatado))
                        total_categoria += valor
                    
                    # Adicionar um rótulo com o total
                    total_formatado = locale.currency(total_categoria, grouping=True)
                    ttk.Label(detalhes_window, text=f"Total: {total_formatado}", font=('Arial', 12, 'bold')).pack(pady=5)

                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao buscar detalhes: {e}", parent=janela_grafico)

            # --- Função de evento para cliques no gráfico ---
            def on_pick(event):
                # O nome da categoria e o mês estão armazenados no GID (Group ID) do artista
                gid = event.artist.get_gid()
                if gid:
                    categoria, mes_selecionado = gid
                    mostrar_detalhes_transacao(categoria, mes_selecionado)

            # --- Configuração da GUI para seleção de mês ---
            frame_selecao = ttk.Frame(frame_principal)
            frame_selecao.pack(fill=tk.X, pady=10)
            
            def obter_meses_disponiveis():
                try:
                    self.cursor.execute("""
                        SELECT DISTINCT strftime('%Y-%m', data_pagamento) as ano_mes,
                                        strftime('%m/%Y', data_pagamento) as mes_formatado
                        FROM v_despesas_compat 
                        WHERE user_id = ?
                        ORDER BY ano_mes DESC LIMIT 24
                    """, (self.sessao.user_id,))
                    return self.cursor.fetchall()
                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao obter meses: {e}")
                    return []
            
            meses_disponiveis = obter_meses_disponiveis()
            if not meses_disponiveis:
                meses_disponiveis = [(mes_atual_ym,  mes_atual_formatado)]
            
            mapa_meses = {mes_fmt: mes_ym for mes_ym, mes_fmt in meses_disponiveis}
            valores_meses_formatados = [mes_fmt for _, mes_fmt in meses_disponiveis]
            
            ttk.Label(frame_selecao, text="Selecione o mês:").pack(side=tk.LEFT, padx=5)
            mes_combo = ttk.Combobox(frame_selecao, state="readonly", width=10)
            mes_combo['values'] = valores_meses_formatados
            mes_combo.pack(side=tk.LEFT, padx=5)


            # --- INÍCIO DA ALTERAÇÃO: Lógica para selecionar o mês inicial ---
            # Tenta definir o mês atual como o padrão na combobox
            if mes_atual_formatado in valores_meses_formatados:
                mes_combo.set(mes_atual_formatado)
            elif valores_meses_formatados:
                # Se o mês atual não estiver na lista (sem dados), usa o mais recente
                mes_combo.current(0)
            # --- FIM DA ALTERAÇÃO ---



            # --- Configuração da Figura e Canvas do Matplotlib ---
            frame_grafico = ttk.Frame(frame_principal)
            frame_grafico.pack(fill=tk.BOTH, expand=True, pady=10)
            
            figura = Figure(figsize=(9, 6), dpi=100)
            ax = figura.add_subplot(111)
            
            canvas = FigureCanvasTkAgg(figura, frame_grafico)
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
            # Conectar o evento de clique ao manipulador
            canvas.mpl_connect('pick_event', on_pick)

            # --- GUI para informações adicionais (Total, Percentual) ---
            frame_info = ttk.Frame(frame_principal)
            frame_info.pack(fill=tk.X, pady=10)
            label_total = ttk.Label(frame_info, text="", font=('Arial', 12))
            label_total.pack(side=tk.LEFT, padx=10)
            label_percentual = ttk.Label(frame_info, text="", font=('Arial', 12))
            label_percentual.pack(side=tk.RIGHT, padx=10)

       
            # --- Função para atualizar o gráfico ---
            def atualizar_grafico(mes_selecionado):
                try:
                    # AQUI ESTÁ A CORREÇÃO: Usando a variável correta "mes_atual_formatado"
                    mes_exibicao = next((m_fmt for m_fmt, m_val in mapa_meses.items() if m_val == mes_selecionado), mes_atual_formatado)
                    
                    ax.clear()
                    
                    self.cursor.execute("""
                        SELECT conta_despesa, SUM(valor) as total_valor
                        FROM v_despesas_compat 
                        WHERE strftime('%Y-%m', data_pagamento) = ?
                        AND user_id = ?
                        GROUP BY conta_despesa ORDER BY total_valor DESC LIMIT 10
                    """, (mes_selecionado, self.sessao.user_id))
                    resultados = self.cursor.fetchall()

                    if not resultados:
                        ax.text(0.5, 0.5, f"Nenhuma despesa encontrada para {mes_exibicao}", ha='center', va='center', fontsize=14)
                        janela_grafico.title(f"Top 10 Contas - {mes_exibicao}")
                        ax.set_title(f'Top 10 Contas de Despesa - {mes_exibicao}')
                        canvas.draw()
                        label_total.config(text="Total Top 10: R$ 0,00")
                        label_percentual.config(text="Representa 0.0% do total do mês")
                        return

                    self.cursor.execute("SELECT SUM(valor) FROM v_despesas_compat WHERE strftime('%Y-%m', data_pagamento) = ? AND user_id = ?", (mes_selecionado, self.sessao.user_id))
                    total_mes = self.cursor.fetchone()[0] or 0
                    
                    contas = [row[0] for row in resultados]
                    valores = [row[1] for row in resultados]
                    
                    total_top10 = sum(valores)
                    percentual_top10 = (total_top10 / total_mes * 100) if total_mes > 0 else 0
                    
                    # Definir cores diferentes para as barras
                    cores = plt.cm.get_cmap('tab10', 10).colors

                    # Criar barras horizontais e torná-las clicáveis
                    barras = ax.barh(contas[::-1], valores[::-1], color=cores, picker=True)

                    # Armazenar informações de categoria e mês em cada barra para o clique
                    for i, barra in enumerate(barras):
                        categoria = contas[::-1][i]
                        barra.set_gid((categoria, mes_selecionado))

                    # Adicionar rótulos de valor
                    for barra in barras:
                        largura = barra.get_width()
                        ax.text(largura * 1.01, barra.get_y() + barra.get_height()/2, 
                                locale.currency(largura, grouping=True),
                                va='center', ha='left', fontsize=9)
                    
                    # Configurar estilo do gráfico
                    ax.set_title(f'Top 10 Contas de Despesa - {mes_exibicao}')
                    ax.set_xlabel('Valor Total (R$)')
                    ax.set_ylabel('Conta de Despesa')
                    ax.grid(True, linestyle='--', alpha=0.7, axis='x')
                    ax.set_xlim(right=max(valores) * 1.18)  # Ajustar limite do eixo x para os rótulos

                    # Formatar o eixo x como moeda
                    def formatar_reais(x, pos):
                        return locale.currency(x, symbol=False, grouping=True)
                    ax.xaxis.set_major_formatter(FuncFormatter(formatar_reais))
                    
                    figura.tight_layout()
                    canvas.draw()
                    
                    # Atualizar informações
                    label_total.config(text=f"Total Top 10: {locale.currency(total_top10, grouping=True)}")
                    label_percentual.config(text=f"Representa {percentual_top10:.1f}% do total do mês")
                
                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao atualizar gráfico: {e}")
                    import traceback
                    traceback.print_exc()
                        
           ####################################################################
           
            
            # --- Manipulador de evento para mudança de mês ---
            def on_mes_change(event):
                try:
                    mes_selecionado_fmt = mes_combo.get()
                    mes_valor_ym = mapa_meses.get(mes_selecionado_fmt)
                    if not mes_valor_ym:
                        messagebox.showerror("Erro", "Mês inválido selecionado")
                        return
                    atualizar_grafico(mes_valor_ym)
                    janela_grafico.title(f"Top 10 Contas - {mes_selecionado_fmt}")
                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao mudar mês: {e}")
            
            mes_combo.bind("<<ComboboxSelected>>", on_mes_change)
            
            # --- Lógica para o botão de exportar ---
            def exportar_dados():
                try:
                    mes_selecionado = mes_combo.get()
                    mes_valor = mapa_meses.get(mes_selecionado)
                    if not mes_valor:
                        messagebox.showerror("Erro", "Mês inválido selecionado")
                        return
                    
                    # Consultar dados para exportação
                    self.cursor.execute("""
                        SELECT conta_despesa, SUM(valor) as total_valor
                        FROM v_despesas_compat 
                        WHERE strftime('%Y-%m', data_pagamento) = ?
                        AND user_id = ?
                        GROUP BY conta_despesa ORDER BY total_valor DESC
                    """, (mes_valor, self.sessao.user_id))
                    resultados = self.cursor.fetchall()
                    
                    if not resultados:
                        messagebox.showinfo("Exportar", "Sem dados para exportar neste mês.")
                        return
                        
                    # Solicitar local para salvar
                    arquivo = filedialog.asksaveasfilename(
                        defaultextension=".xlsx",
                        filetypes=[("Excel files", "*.xlsx")],
                        title="Salvar Relatório"
                    )
                    
                    if not arquivo:
                        return
                        
                    # Criar Excel
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Top Contas"
                    
                    ws['A1'] = f"Relatório Top Contas - {mes_selecionado}"
                    ws.merge_cells('A1:B1')
                    ws['A1'].font = Font(size=14, bold=True)
                    
                    ws['A3'] = "Conta"
                    ws['B3'] = "Valor (R$)"
                    ws['A3'].font = Font(bold=True)
                    ws['B3'].font = Font(bold=True)
                    
                    for idx, (conta, valor) in enumerate(resultados, start=4):
                        ws.cell(row=idx, column=1, value=conta)
                        ws.cell(row=idx, column=2, value=valor).number_format = '#,##0.00'
                        
                    wb.save(arquivo)
                    messagebox.showinfo("Sucesso", "Dados exportados com sucesso!")
                    
                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao exportar: {e}")

            ttk.Button(frame_principal, text="Exportar Dados", command=exportar_dados).pack(pady=10)

            # Inicializar com o mês selecionado
            if mes_combo.get():
                on_mes_change(None)

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao abrir gráfico: {e}")

    def configurar_estilo(self):
        style = ttk.Style()
        
        # Configurar estilo para botões
        style.configure('TButton', 
                        background=self.cor_primaria, 
                        foreground='white', 
                        font=('Arial', 10, 'bold'),
                        padding=5)
        
        # Configurar estilo para labels
        style.configure('TLabel', 
                        background=self.cor_fundo, 
                        foreground=self.cor_texto,
                        font=('Arial', 10))
        
        # Configurar estilo para entradas
        style.configure('TEntry', 
                        fieldbackground='white',
                        font=('Arial', 10))
        
        # Configurar estilo para frame
        style.configure('TFrame', background=self.cor_fundo)
        
        # Configurar estilo para treeview (tabela)
        style.configure('Treeview', 
                        background='white',
                        fieldbackground='white',
                        font=('Arial', 9))
        
        style.configure('Treeview.Heading', 
                        font=('Arial', 10, 'bold'),
                        background=self.cor_secundaria)

            self.cursor = self.conn.cursor()
            
            # --- ETAPA 1: Criar as tabelas do NOVO SCHEMA ---
            
            # 1. Tabela de Usuários
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS users (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT UNIQUE NOT NULL,
                    email TEXT UNIQUE NOT NULL,
                    password_hash TEXT NOT NULL,
                    nivel_acesso TEXT DEFAULT 'usuario',
                    ativo BOOLEAN DEFAULT 1,
                    data_criacao DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            ''')

            # 2. Categorias de Despesa (Normalizada)
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS categorias_despesa (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    nome TEXT UNIQUE NOT NULL,
                    ativo BOOLEAN DEFAULT 1
                )
            ''')

            # 3. Meios de Pagamento (Normalizada)
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS meios_pagamento (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    nome TEXT UNIQUE NOT NULL,
                    tipo TEXT,
                    ativo BOOLEAN DEFAULT 1
                )
            ''')

            # 4. Despesas (Com Chaves Estrangeiras)
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS despesas (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    descricao TEXT NOT NULL,
                    valor REAL NOT NULL,
                    num_parcelas INTEGER DEFAULT 1,
                    data_registro DATETIME DEFAULT CURRENT_TIMESTAMP,
                    data_pagamento DATE NOT NULL,
                    user_id INTEGER NOT NULL,
                    categoria_id INTEGER NOT NULL,
                    meio_pagamento_id INTEGER NOT NULL,
                    FOREIGN KEY (user_id) REFERENCES users(id),
                    FOREIGN KEY (categoria_id) REFERENCES categorias_despesa(id),
                    FOREIGN KEY (meio_pagamento_id) REFERENCES meios_pagamento(id)
                )
            ''')

            # --- ETAPA 2: View de Compatibilidade (O "Segredo" da Migração Rápida) ---
            self.cursor.execute("DROP VIEW IF EXISTS v_despesas_compat")
            self.cursor.execute('''
                CREATE VIEW v_despesas_compat AS
                SELECT d.id, d.descricao, d.valor, d.num_parcelas, d.data_pagamento, d.data_registro,
                       c.nome as conta_despesa,
                       m.nome as meio_pagamento,
                       d.user_id
                FROM despesas d
                JOIN categorias_despesa c ON d.categoria_id = c.id
                JOIN meios_pagamento m ON d.meio_pagamento_id = m.id
            ''')

            # --- ETAPA 3: Popular dados padrão se necessário ---

            # Usuário Admin Padrão
            self.cursor.execute("SELECT COUNT(*) FROM users")
            if self.cursor.fetchone()[0] == 0:
                senha_hash = generate_password_hash('admin123')
                self.cursor.execute("""
                    INSERT INTO users (username, email, password_hash, nivel_acesso)
                    VALUES (?, ?, ?, ?)
                """, ('admin', 'admin@financeiro.com', senha_hash, 'admin'))
                print("Usuário 'admin' criado com senha 'admin123'")

            # Categorias Padrão
            self.cursor.execute("SELECT COUNT(*) FROM categorias_despesa")
            if self.cursor.fetchone()[0] == 0:
                categorias_padrao = [
                    ('Tel. e Internet',), ('Gás',), ('Mercado',), ('Alimentação',), ('Moradia',), 
                    ('Transporte',), ('Educação',), ('Saúde',), ('Lazer',), ('Vestuário',), 
                    ('Funcionários',), ('Outros',)
                ]
                self.cursor.executemany("INSERT INTO categorias_despesa (nome) VALUES (?)", categorias_padrao)

            # Meios de Pagamento Padrão
            self.cursor.execute("SELECT COUNT(*) FROM meios_pagamento")
            if self.cursor.fetchone()[0] == 0:
                meios_padrao = [
                    ('Dinheiro', 'dinheiro'), ('Cartão Unlimited', 'cartao'), ('Cartão C6', 'cartao'), 
                    ('Cartão Nubank', 'cartao'), ('Cartão BB', 'cartao'), ('Transferência', 'transferencia'), 
                    ('PIX', 'pix'), ('Boleto', 'boleto')
                ]
                self.cursor.executemany("INSERT INTO meios_pagamento (nome, tipo) VALUES (?, ?)", meios_padrao)

            self.conn.commit()
            
        except sqlite3.Error as e:
            messagebox.showerror("Erro no Banco de Dados", f"Erro ao criar banco de dados: {e}")

    def carregar_categorias(self):
        """Carrega todas as categorias ativas do banco de dados"""
        self.cursor.execute("SELECT nome FROM categorias_despesa WHERE ativo = 1 ORDER BY nome")
        return [categoria[0] for categoria in self.cursor.fetchall()]
    
    def carregar_meios_pagamento(self):
        """Carrega todos os meios de pagamento ativos do banco de dados"""
        self.cursor.execute("SELECT nome FROM meios_pagamento WHERE ativo = 1 ORDER BY nome")
        return [meio[0] for meio in self.cursor.fetchall()]
    
    def criar_widgets(self):
        """Cria todos os elementos da interface gráfica"""
        
        # Frame principal dividido em duas partes
        self.frame_principal = ttk.Frame(self.root)
        self.frame_principal.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Frame lateral esquerdo (formulário)
        self.frame_form = ttk.Frame(self.frame_principal, style='TFrame')
        self.frame_form.pack(side=tk.LEFT, fill=tk.BOTH, padx=10, pady=10)
        
        # Frame direito (tabela e gráficos)
        self.frame_tabela = ttk.Frame(self.frame_principal, style='TFrame')
        self.frame_tabela.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Título
        titulo = ttk.Label(self.frame_form, text="Controle Financeiro", 
                          font=('Arial', 16, 'bold'), style='TLabel')
        titulo.grid(row=0, column=0, columnspan=2, pady=10)
        
        # Formulário de entrada
        # ID (Hidden)
        ttk.Label(self.frame_form, text="ID:").grid(row=1, column=0, sticky=tk.W, pady=5)
        id_entry = ttk.Entry(self.frame_form, textvariable=self.id_despesa, state='readonly', width=5)
        id_entry.grid(row=1, column=1, sticky=tk.W, pady=5)
        
        # Descrição
        ttk.Label(self.frame_form, text="Descrição:").grid(row=2, column=0, sticky=tk.W, pady=5)
        ttk.Entry(self.frame_form, textvariable=self.descricao, width=30).grid(row=2, column=1, sticky=tk.W, pady=5)
        
        # Meio de Pagamento (ComboBox)
        ttk.Label(self.frame_form, text="Meio de Pagamento:").grid(row=3, column=0, sticky=tk.W, pady=5)
        self.combo_meio_pagamento = ttk.Combobox(self.frame_form, textvariable=self.meio_pagamento, width=20)
        self.combo_meio_pagamento['values'] = self.carregar_meios_pagamento()
        self.combo_meio_pagamento.grid(row=3, column=1, sticky=tk.W, pady=5)
        
        # Conta de Despesa (ComboBox)
        ttk.Label(self.frame_form, text="Categoria:").grid(row=4, column=0, sticky=tk.W, pady=5)
        self.combo_conta_despesa = ttk.Combobox(self.frame_form, textvariable=self.conta_despesa, width=20)
        self.combo_conta_despesa['values'] = self.carregar_categorias()
        self.combo_conta_despesa.grid(row=4, column=1, sticky=tk.W, pady=5)
        
        # Valor
        ttk.Label(self.frame_form, text="Valor (R$):").grid(row=5, column=0, sticky=tk.W, pady=5)
        vcmd = (self.root.register(validar_entrada_numerica), '%P')
        ttk.Entry(self.frame_form, textvariable=self.valor, width=15, validate='key', validatecommand=vcmd).grid(row=5, column=1, sticky=tk.W, pady=5)
        
        # Número de Parcelas
        ttk.Label(self.frame_form, text="Parcelas:").grid(row=6, column=0, sticky=tk.W, pady=5)
        ttk.Spinbox(self.frame_form, from_=1, to=48, textvariable=self.num_parcelas, width=5).grid(row=6, column=1, sticky=tk.W, pady=5)
        self.num_parcelas.set(1)  # Valor padrão
        
        # Data de Pagamento
      #  ttk.Label(self.frame_form, text="Data de Pagamento:").grid(row=7, column=0, sticky=tk.W, pady=5)
      #  self.data_entry = DateEntry(self.frame_form, width=12, background=self.cor_primaria,
      #                             foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
       # self.data_entry.grid(row=7, column=1, sticky=tk.W, pady=5)
      #  self.data_entry.set_date(datetime.now())
        
        self.data_entry = DateEntry(self.frame_form, width=12, background=self.cor_primaria,
                                   foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy',
                                   locale='pt_BR') # Padrão de data alterado para dd/mm/yyyy e locale pt_BR adicionado
                                   
        self.data_entry.grid(row=7, column=1, sticky=tk.W, pady=5)
        self.data_entry.set_date(datetime.now()) # Define a data atual ao iniciar
        
        
        # Frame dos Botões
        botoes_frame = ttk.Frame(self.frame_form)
        botoes_frame.grid(row=8, column=0, columnspan=2, pady=15)
        
        # Botões de ação
        self.btn_salvar = ttk.Button(botoes_frame, text="Salvar", command=self.salvar_despesa)
        self.btn_salvar.grid(row=0, column=0, padx=5)
        
        self.btn_atualizar = ttk.Button(botoes_frame, text="Atualizar", command=self.atualizar_despesa)
        self.btn_atualizar.grid(row=0, column=1, padx=5)
        self.btn_atualizar['state'] = 'disabled'
        
        self.btn_excluir = ttk.Button(botoes_frame, text="Excluir", command=self.excluir_despesa)
        self.btn_excluir.grid(row=0, column=2, padx=5)
        self.btn_excluir['state'] = 'disabled'
        
        self.btn_limpar = ttk.Button(botoes_frame, text="Limpar", command=self.limpar_campos)
        self.btn_limpar.grid(row=0, column=3, padx=5)

       # self.btn_conf = ttk.Button(botoes_frame, text="Configuração", command=self.abrir_gerenciador)
       # self.btn_conf.grid(row=0, column=4, padx=5)

                      
        # Frame de Pesquisa
        frame_pesquisa = ttk.LabelFrame(self.frame_form, text="Pesquisar")
        frame_pesquisa.grid(row=9, column=0, columnspan=2, sticky=tk.W+tk.E, pady=10)
        
        self.pesquisa_termo = tk.StringVar()
        ttk.Entry(frame_pesquisa, textvariable=self.pesquisa_termo, width=20).grid(row=0, column=0, padx=5, pady=5)
        ttk.Button(frame_pesquisa, text="Buscar", command=self.pesquisar_despesa).grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(frame_pesquisa, text="Mostrar Todos", command=self.carregar_despesas).grid(row=0, column=2, padx=5, pady=5)
        
        # Frame de Relatórios
        frame_relatorios = ttk.LabelFrame(self.frame_form, text="Relatórios")
        frame_relatorios.grid(row=10, column=0, columnspan=2, sticky=tk.W+tk.E, pady=10)
        
        ttk.Button(frame_relatorios, text="Relatório Mensal", 
                  command=self.gerar_relatorio_mensal).grid(row=0, column=0, padx=5, pady=5)
        ttk.Button(frame_relatorios, text="Relatório por Categoria", 
                  command=self.gerar_relatorio_categoria).grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(frame_relatorios, text="Exportar Excel", 
                  command=self.exportar_excel).grid(row=0, column=2, padx=5, pady=5)
       # ttk.Button(frame_relatorios, text="Relatório Meios de Pagamento", 
       #           command=self.gerar_relatorio_meios_pagamento).grid(row=0, column=3, padx=5, pady=5)
        
        # Tabela de despesas
        self.tree_frame = ttk.LabelFrame(self.frame_tabela, text="Registros de Despesas")
        self.tree_frame.pack(fill=tk.BOTH, expand=True)
        
        # Scrollbar para a tabela
        scrollbar = ttk.Scrollbar(self.tree_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Configurar a tabela (Treeview)
        self.tabela = ttk.Treeview(self.tree_frame, yscrollcommand=scrollbar.set)
        
        # Definir colunas
        self.tabela['columns'] = ('id', 'descricao', 'meio_pagamento', 'conta_despesa', 
                                 'valor', 'parcelas', 'data')
        
        # Formatar colunas
        self.tabela.column('#0', width=0, stretch=tk.NO)  # Coluna fantasma
        self.tabela.column('id', anchor=tk.CENTER, width=50)
        self.tabela.column('descricao', anchor=tk.W, width=200)
        self.tabela.column('meio_pagamento', anchor=tk.W, width=120)
        self.tabela.column('conta_despesa', anchor=tk.W, width=120)
        self.tabela.column('valor', anchor=tk.E, width=100)
        self.tabela.column('parcelas', anchor=tk.CENTER, width=70)
        self.tabela.column('data', anchor=tk.CENTER, width=100)
        
        # Definir cabeçalhos
        self.tabela.heading('#0', text='', anchor=tk.CENTER)
        self.tabela.heading('id', text='ID', anchor=tk.CENTER)
        self.tabela.heading('descricao', text='Descrição', anchor=tk.CENTER)
        self.tabela.heading('meio_pagamento', text='Meio Pagamento', anchor=tk.CENTER)
        self.tabela.heading('conta_despesa', text='Categoria', anchor=tk.CENTER)
        self.tabela.heading('valor', text='Valor (R$)', anchor=tk.CENTER)
        self.tabela.heading('parcelas', text='Parcelas', anchor=tk.CENTER)
        self.tabela.heading('data', text='Data', anchor=tk.CENTER)

        # === INÍCIO NOVA ALTERAÇÃO: Ativar ordenação e menu de contexto ===
        # Atribuir comando de ordenação para cada cabeçalho
        for col_id in self.tabela['columns']:
            # A função `heading` é usada tanto para configurar quanto para ler.
            # Aqui, lemos o texto atual para não o perder.
            text = self.tabela.heading(col_id, 'text')
            self.tabela.heading(col_id, text=text, command=lambda _col=col_id: self._sort_by_column(_col))
        
        # Criar menu de contexto para seleção de colunas
        self.header_context_menu = tk.Menu(self.root, tearoff=0)
        for col_id, (text, var) in self._all_columns.items():
            self.header_context_menu.add_checkbutton(
                label=text,
                variable=var,
                command=self._update_visible_columns
            )
        
        # Vincular o clique com o botão direito ao evento que mostra o menu
        self.tabela.bind('<Button-3>', self._show_header_context_menu)
        # === FIM NOVA ALTERAÇÃO ===
        
        # Vincular evento de clique na tabela
        self.tabela.bind("<ButtonRelease-1>", self.selecionar_item)
        
        # Posicionar a tabela e conectar scrollbar
        self.tabela.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        scrollbar.config(command=self.tabela.yview)
        
        # Frame para gráficos
        # INÍCIO DA MODIFICAÇÃO: Título do LabelFrame é genérico agora
        self.frame_grafico = ttk.LabelFrame(self.frame_tabela, text="Resumo de Gastos")
        # FIM DA MODIFICAÇÃO
        self.frame_grafico.pack(fill=tk.BOTH, expand=True, pady=10)

        # INÍCIO DA MODIFICAÇÃO: Adiciona controles de seleção de mês
        frame_controles_grafico = ttk.Frame(self.frame_grafico)
        frame_controles_grafico.pack(fill=tk.X, padx=5, pady=2)
        
        ttk.Label(frame_controles_grafico, text="Mês:").pack(side=tk.LEFT, padx=(0, 5))
        self.combo_mes_grafico = ttk.Combobox(frame_controles_grafico, state="readonly", width=10)
        self.combo_mes_grafico.pack(side=tk.LEFT)
        self.combo_mes_grafico.bind("<<ComboboxSelected>>", self.atualizar_grafico)
        # FIM DA MODIFICAÇÃO
        
        # INÍCIO DA NOVA ALTERAÇÃO: Adiciona o Label para o total do mês
        label_total_mes = ttk.Label(frame_controles_grafico, textvariable=self.total_mes_selecionado_var, font=('Arial', 9, 'bold'))
        label_total_mes.pack(side=tk.LEFT, padx=(15, 0))
        # FIM DA NOVA ALTERAÇÃO


        # Criar gráfico inicial
        self.figura_grafico = plt.Figure(figsize=(5, 4), dpi=100)
        self.canvas_grafico = FigureCanvasTkAgg(self.figura_grafico, self.frame_grafico)
        self.canvas_grafico.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
        # Inicializar o gráfico
        self.atualizar_grafico()

    # === INÍCIO NOVA ALTERAÇÃO: Métodos para ordenação e visibilidade ===
    def _sort_by_column(self, col):
        """Ordena a tabela pela coluna clicada."""
        # Pega todos os itens da tabela
        items = [(self.tabela.set(k, col), k) for k in self.tabela.get_children('')]

        # Determina a direção da ordenação
        reverse = self._sort_state['col'] == col and not self._sort_state['reverse']
        self._sort_state = {'col': col, 'reverse': reverse}

        # Cria uma função chave para converter os dados antes de ordenar
        def get_sort_key(item):
            value = item[0]
            if col == 'valor':
                # Converte 'R$ 1.234,56' para o número 1234.56
                try:
                    return float(value.replace('R$', '').replace('.', '').replace(',', '.').strip())
                except (ValueError, AttributeError):
                    return 0.0
            elif col == 'data':
                # Converte 'dd/mm/AAAA' para um objeto de data
                try:
                    return datetime.strptime(value, '%d/%m/%Y')
                except (ValueError, TypeError):
                    return datetime.min # Retorna uma data mínima para valores ruins/vazios
            elif col in ['id', 'parcelas']:
                # Converte para inteiro
                try:
                    return int(value)
                except (ValueError, TypeError):
                    return 0
            else: # Para descrição, categoria, etc., ordena como texto (ignorando maiúsculas/minúsculas)
                return str(value).lower()

        # Ordena a lista de itens
        items.sort(key=get_sort_key, reverse=reverse)

        # Reorganiza os itens na tabela
        for index, (val, k) in enumerate(items):
            self.tabela.move(k, '', index)

        # Atualiza os cabeçalhos para mostrar um indicador de ordenação (▲/▼)
        for c in self.tabela['columns']:
            current_text = self.tabela.heading(c, 'text').replace(' ▲', '').replace(' ▼', '')
            self.tabela.heading(c, text=current_text)
        
        new_header_text = self.tabela.heading(col, 'text') + (' ▼' if reverse else ' ▲')
        self.tabela.heading(col, text=new_header_text)

    def _show_header_context_menu(self, event):
        """Exibe o menu de contexto ao clicar com o botão direito no cabeçalho."""
        if self.tabela.identify_region(event.x, event.y) == 'heading':
            self.header_context_menu.post(event.x_root, event.y_root)

    def _update_visible_columns(self):
        """Atualiza quais colunas estão visíveis na tabela."""
        visible_cols = [col_id for col_id, (_, var) in self._all_columns.items() if var.get()]
        
        self.tabela['displaycolumns'] = visible_cols
    # === FIM NOVA ALTERAÇÃO ===


    def abrir_gerenciador(self):
        """Abre o programa 2 dentro de uma nova janela Tkinter."""
        # opcionalmente pode usar Toplevel() se quiser que compartilhe a mesma root
        # # def abrir_gerenciador(self):
        # """Abre a janela de configurações"""
        config_window = configuracao.GerenciadorConfiguracoes(self.root, False)
        
        # Opcional: aguardar a janela de configuração fechar antes de atualizar dados
        self.root.wait_window(config_window.root)
        
        # Atualizar dados após fechar a janela de configuração
        self.combo_meio_pagamento['values'] = self.carregar_meios_pagamento()
        self.combo_conta_despesa['values'] = self.carregar_categorias()

    def abrir_gerenciador2(self):
        """Abre o programa 2 dentro de uma nova janela Tkinter."""
        # opcionalmente pode usar Toplevel() se quiser que compartilhe a mesma root
        # # def abrir_gerenciador(self):
        # """Abre a janela de configurações"""
        config_window = MENUBD.GerenciadorConfiguracoes2(self.root, False)
        
        # Opcional: aguardar a janela de configuração fechar antes de atualizar dados
        self.root.wait_window(config_window.root)
        
        # Atualizar dados após fechar a janela de configuração
        self.combo_meio_pagamento['values'] = self.carregar_meios_pagamento()
        self.combo_conta_despesa['values'] = self.carregar_categorias()
    

    # ### OTIMIZAÇÃO 1: Criar um método reutilizável para abrir o calendário ###
    def _abrir_calendario_selecao(self, parent_dialog, entry_var):
        """Abre uma janela Toplevel com um widget de calendário para selecionar uma data."""
        
        def pegar_data():
            """Pega a data selecionada e atualiza a variável de entrada."""
            data_selecionada = cal.selection_get()
            entry_var.set(data_selecionada.strftime("%d/%m/%Y"))
            top.destroy()
        
        top = tk.Toplevel(parent_dialog)
        top.title("Selecionar Data")
        
        # Tenta obter a data atual do campo de entrada para iniciar o calendário
        try:
            data_atual = datetime.strptime(entry_var.get(), "%d/%m/%Y")
        except ValueError:
            data_atual = datetime.now()

        # ### CORREÇÃO 2: Usar o widget 'Calendar' importado corretamente ###
        cal = Calendar(top, selectmode='day', date_pattern='dd/mm/yyyy',
                       year=data_atual.year, month=data_atual.month, day=data_atual.day,
                       locale='pt_BR')
        cal.pack(pady=10, padx=10)
        
        ttk.Button(top, text="Selecionar", command=pegar_data).pack(pady=5)
        
        # Centralizar e focar a janela do calendário
        top.transient(parent_dialog)
        top.wait_visibility()
        top.grab_set()
        top.focus_set()


    def gerar_relatorio_entre_datas(self):
        """Gera um relatório de despesas entre duas datas selecionadas"""
        try:
            dialog = tk.Toplevel(self.root)
            dialog.title("Relatório Entre Datas")
            dialog.geometry("400x200") # Reduzido um pouco
            dialog.resizable(False, False)
            
            frame = ttk.Frame(dialog, padding="10")
            frame.pack(fill=tk.BOTH, expand=True)
            
            data_atual = datetime.now()
            data_inicial_var = tk.StringVar(value=data_atual.strftime("%d/%m/%Y"))
            data_final_var = tk.StringVar(value=data_atual.strftime("%d/%m/%Y"))
            
            ttk.Label(frame, text="Data Inicial:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
            data_inicial_entry = ttk.Entry(frame, textvariable=data_inicial_var, width=15)
            data_inicial_entry.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)
            
            ttk.Label(frame, text="Data Final:").grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
            data_final_entry = ttk.Entry(frame, textvariable=data_final_var, width=15)
            data_final_entry.grid(row=1, column=1, padx=5, pady=5, sticky=tk.W)

            # ### OTIMIZAÇÃO 2: Chamar o método reutilizável ###
            ttk.Button(frame, text="📅", width=3, 
                       command=lambda: self._abrir_calendario_selecao(dialog, data_inicial_var)).grid(row=0, column=2, padx=5)
            ttk.Button(frame, text="📅", width=3, 
                       command=lambda: self._abrir_calendario_selecao(dialog, data_final_var)).grid(row=1, column=2, padx=5)
            
            def gerar():
                try:
                    data_inicial = datetime.strptime(data_inicial_var.get(), "%d/%m/%Y").strftime("%Y-%m-%d")
                    data_final = datetime.strptime(data_final_var.get(), "%d/%m/%Y").strftime("%Y-%m-%d")
                    
                    if data_inicial > data_final:
                        messagebox.showerror("Erro", "Data inicial deve ser anterior à data final.", parent=dialog)
                        return
                        
                    self.mostrar_relatorio_entre_datas(data_inicial, data_final)
                    dialog.destroy()
                except ValueError:
                    messagebox.showerror("Erro", "Formato de data inválido. Use DD/MM/AAAA.", parent=dialog)
            
            ttk.Button(frame, text="Gerar Relatório", command=gerar).grid(row=3, column=0, columnspan=3, pady=20)
            
            dialog.transient(self.root)
            dialog.grab_set()
            dialog.wait_window()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar relatório: {e}")

    def mostrar_relatorio_entre_datas(self, data_inicial, data_final):
        """Exibe o relatório para o período entre as datas selecionadas"""
        try:
            # Formatar datas para exibição
            data_inicial_formatada = datetime.strptime(data_inicial, "%Y-%m-%d").strftime("%d/%m/%Y")
            data_final_formatada = datetime.strptime(data_final, "%Y-%m-%d").strftime("%d/%m/%Y")
            
            # Consultar dados do período
            self.cursor.execute("""
                SELECT 
                    conta_despesa, 
                    SUM(valor) as total_valor,
                    COUNT(*) as total_registros
                FROM v_despesas_compat
                WHERE data_pagamento BETWEEN ? AND ?
                AND user_id = ?
                GROUP BY conta_despesa
                ORDER BY total_valor DESC
            """, (data_inicial, data_final, self.sessao.user_id))
            
            resultados = self.cursor.fetchall()
            
            if not resultados:
                messagebox.showinfo("Relatório", f"Nenhuma despesa encontrada no período de {data_inicial_formatada} a {data_final_formatada}.")
                return
                
            # Calcular total geral
            total_geral = sum(row[1] for row in resultados)
            
            # Criar nova janela para o relatório
            janela_relatorio = tk.Toplevel(self.root)
            janela_relatorio.title(f"Relatório de {data_inicial_formatada} a {data_final_formatada}")
            janela_relatorio.geometry("900x600")
            
            # Frame principal
            frame_principal = ttk.Frame(janela_relatorio, padding="10")
            frame_principal.pack(fill=tk.BOTH, expand=True)
            
            # Título
            titulo = ttk.Label(frame_principal, 
                            text=f"Relatório de Despesas - {data_inicial_formatada} a {data_final_formatada}",
                            font=('Arial', 16, 'bold'))
            titulo.pack(pady=10)
            
            # Total geral
            total_label = ttk.Label(frame_principal, 
                                text=f"Total de Despesas: R$ {total_geral:.2f}".replace('.', ','),
                                font=('Arial', 12))
            total_label.pack(pady=5)
            
            # Frame com abas
            notebook = ttk.Notebook(frame_principal)
            notebook.pack(fill=tk.BOTH, expand=True, pady=10)
            
            # Aba de tabela
            tab_tabela = ttk.Frame(notebook)
            notebook.add(tab_tabela, text="Tabela")
            
            # Aba de gráfico
            tab_grafico = ttk.Frame(notebook)
            notebook.add(tab_grafico, text="Gráfico")
            
            # Criar tabela na primeira aba
            tree_frame = ttk.Frame(tab_tabela)
            tree_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
            
            # Scrollbar para a tabela
            scrollbar = ttk.Scrollbar(tree_frame)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # Tabela de resultados
            colunas = ('categoria', 'total', 'registros', 'percentual')
            tabela = ttk.Treeview(tree_frame, columns=colunas, show='headings', yscrollcommand=scrollbar.set)
            
            # Cabeçalhos
            tabela.heading('categoria', text='Categoria')
            tabela.heading('total', text='Total (R$)')
            tabela.heading('registros', text='Qtd. Registros')
            tabela.heading('percentual', text='% do Total')
            
            # Larguras
            tabela.column('categoria', width=150)
            tabela.column('total', width=100, anchor=tk.E)
            tabela.column('registros', width=100, anchor=tk.CENTER)
            tabela.column('percentual', width=100, anchor=tk.CENTER)
            
            # Preencher tabela
            for row in resultados:
                categoria, total, registros = row
                percentual = (total / total_geral) * 100
                
                tabela.insert('', tk.END, values=(
                    categoria,
                    f"R$ {total:.2f}".replace('.', ','),
                    registros,
                    f"{percentual:.1f}%"
                ))
                
            tabela.pack(fill=tk.BOTH, expand=True)
            scrollbar.config(command=tabela.yview)
            
            # Criar gráfico na segunda aba
            figura = plt.Figure(figsize=(7, 5), dpi=100)
            ax = figura.add_subplot(111)
            
            # Extrair dados para o gráfico
            categorias = [row[0] for row in resultados]
            valores = [row[1] for row in resultados]
            
            # Criar gráfico de pizza
            wedges, texts, autotexts = ax.pie(
                valores, 
                labels=None,
                autopct='%1.1f%%',
                startangle=90,
                shadow=False,
            )
            
            # Estilo do gráfico
            ax.set_title(f'Distribuição de Despesas - {data_inicial_formatada} a {data_final_formatada}')
            ax.axis('equal')  # Garantir que o gráfico seja circular
            
            # Adicionar legenda
            ax.legend(wedges, categorias, title="Categorias", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))
            
            # Adicionar o gráfico ao frame
            canvas = FigureCanvasTkAgg(figura, tab_grafico)
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

            # Função para exportar para Excel
            def exportar_para_excel():
                try:
                    # Solicitar local para salvar o arquivo
                    arquivo = filedialog.asksaveasfilename(
                        defaultextension=".xlsx",
                        filetypes=[("Excel files", "*.xlsx")],
                        title="Salvar Relatório"
                    )
                    
                    if not arquivo:
                        return
                    
                    # Criar planilha
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Relatório de Despesas"
                    
                    # Título
                    ws['A1'] = f"Relatório de Despesas - Período: {data_inicial_formatada} a {data_final_formatada}"
                    ws.merge_cells('A1:D1')
                    ws['A1'].font = Font(size=14, bold=True)
                    
                    # Total
                    ws['A2'] = f"Total de Despesas: R$ {total_geral:.2f}".replace('.', ',')
                    ws.merge_cells('A2:D2')
                    ws['A2'].font = Font(size=12)
                    
                    # Cabeçalhos
                    cabecalhos = ["Categoria", "Total (R$)", "Qtd. Registros", "% do Total"]
                    for col, cabecalho in enumerate(cabecalhos, start=1):
                        cell = ws.cell(row=4, column=col, value=cabecalho)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                    
                    # Dados
                    for row_idx, (categoria, total, registros) in enumerate(resultados, start=5):
                        percentual = (total / total_geral) * 100
                        
                        ws.cell(row=row_idx, column=1, value=categoria)
                        ws.cell(row=row_idx, column=2, value=f"R$ {total:.2f}".replace('.', ','))
                        ws.cell(row=row_idx, column=3, value=registros)
                        ws.cell(row=row_idx, column=4, value=f"{percentual:.1f}%")
                    
                    # Ajustar larguras
                    for col in range(1, 5):
                        ws.column_dimensions[get_column_letter(col)].width = 20
                    
                    # Salvar arquivo
                    wb.save(arquivo)
                    messagebox.showinfo("Exportar", f"Relatório exportado com sucesso para {arquivo}")
                    
                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao exportar relatório: {e}")
            
            # Botão para exportar
            ttk.Button(frame_principal, text="Exportar para Excel", 
                    command=exportar_para_excel).pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exibir relatório: {e}")        

    ##########################

    def gerar_relatorio_mensal_periodo(self):
        """Gera um relatório de despesas totais somadas por mês entre um intervalo de datas"""
        try:
            dialog = tk.Toplevel(self.root)
            dialog.title("Relatório de Despesas por Mês")
            dialog.geometry("400x200") # Reduzido um pouco
            dialog.resizable(False, False)
            
            frame = ttk.Frame(dialog, padding="10")
            frame.pack(fill=tk.BOTH, expand=True)
            
            data_atual = datetime.now()
            data_inicial_var = tk.StringVar(value=f"01/01/{data_atual.year}")
            data_final_var = tk.StringVar(value=data_atual.strftime("%d/%m/%Y"))
            
            ttk.Label(frame, text="Data Inicial:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
            data_inicial_entry = ttk.Entry(frame, textvariable=data_inicial_var, width=15)
            data_inicial_entry.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)
            
            ttk.Label(frame, text="Data Final:").grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
            data_final_entry = ttk.Entry(frame, textvariable=data_final_var, width=15)
            data_final_entry.grid(row=1, column=1, padx=5, pady=5, sticky=tk.W)
            
            # ### OTIMIZAÇÃO 3: Chamar o método reutilizável também aqui ###
            ttk.Button(frame, text="📅", width=3, 
                       command=lambda: self._abrir_calendario_selecao(dialog, data_inicial_var)).grid(row=0, column=2, padx=5)
            ttk.Button(frame, text="📅", width=3, 
                       command=lambda: self._abrir_calendario_selecao(dialog, data_final_var)).grid(row=1, column=2, padx=5)

            def gerar():
                try:
                    data_inicial = datetime.strptime(data_inicial_var.get(), "%d/%m/%Y").strftime("%Y-%m-%d")
                    data_final = datetime.strptime(data_final_var.get(), "%d/%m/%Y").strftime("%Y-%m-%d")
                    
                    if data_inicial > data_final:
                        messagebox.showerror("Erro", "Data inicial deve ser anterior à data final.", parent=dialog)
                        return
                        
                    self.mostrar_relatorio_mensal_periodo(data_inicial, data_final)
                    dialog.destroy()
                except ValueError:
                    messagebox.showerror("Erro", "Formato de data inválido. Use DD/MM/AAAA.", parent=dialog)
            
            ttk.Button(frame, text="Gerar Relatório", command=gerar).grid(row=3, column=0, columnspan=3, pady=20)
            
            dialog.transient(self.root)
            dialog.grab_set()
            dialog.wait_window()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar relatório: {e}")

    def mostrar_relatorio_mensal_periodo(self, data_inicial, data_final):
        """Exibe o relatório de despesas totais somadas por mês dentro do período selecionado"""
        try:
            # Formatar datas para exibição
            data_inicial_formatada = datetime.strptime(data_inicial, "%Y-%m-%d").strftime("%d/%m/%Y")
            data_final_formatada = datetime.strptime(data_final, "%Y-%m-%d").strftime("%d/%m/%Y")
            
            # Consultar despesas agrupadas por mês
            self.cursor.execute("""
                SELECT 
                    strftime('%Y-%m', data_pagamento) as ano_mes,
                    strftime('%m/%Y', data_pagamento) as mes_ano_formatado,
                    SUM(valor) as total_valor,
                    COUNT(*) as total_registros
                FROM v_despesas_compat
                WHERE data_pagamento BETWEEN ? AND ?
                AND user_id = ?
                GROUP BY ano_mes
                ORDER BY ano_mes
            """, (data_inicial, data_final, self.sessao.user_id))
            
            resultados = self.cursor.fetchall()
            
            if not resultados:
                messagebox.showinfo("Relatório", f"Nenhuma despesa encontrada no período de {data_inicial_formatada} a {data_final_formatada}.")
                return
                
            # Calcular total geral
            total_geral = sum(row[2] for row in resultados)
            
            # Criar nova janela para o relatório
            janela_relatorio = tk.Toplevel(self.root)
            janela_relatorio.title(f"Relatório Mensal - {data_inicial_formatada} a {data_final_formatada}")
            janela_relatorio.geometry("900x600")
            
            # Frame principal
            frame_principal = ttk.Frame(janela_relatorio, padding="10")
            frame_principal.pack(fill=tk.BOTH, expand=True)
            
            # Título
            titulo = ttk.Label(frame_principal, 
                            text=f"Despesas Mensais - {data_inicial_formatada} a {data_final_formatada}",
                            font=('Arial', 16, 'bold'))
            titulo.pack(pady=10)
            
            # Total geral
            total_label = ttk.Label(frame_principal, 
                                text=f"Total de Despesas: R$ {total_geral:.2f}".replace('.', ','),
                                font=('Arial', 12))
            total_label.pack(pady=5)
            
            # Frame com abas
            notebook = ttk.Notebook(frame_principal)
            notebook.pack(fill=tk.BOTH, expand=True, pady=10)
            
            # Aba de tabela
            tab_tabela = ttk.Frame(notebook)
            notebook.add(tab_tabela, text="Tabela")
            
            # Aba de gráfico
            tab_grafico = ttk.Frame(notebook)
            notebook.add(tab_grafico, text="Gráfico")
            
            # Criar tabela na primeira aba
            tree_frame = ttk.Frame(tab_tabela)
            tree_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
            
            # Scrollbar para a tabela
            scrollbar = ttk.Scrollbar(tree_frame)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # Tabela de resultados
            colunas = ('mes_ano', 'total', 'registros', 'percentual')
            tabela = ttk.Treeview(tree_frame, columns=colunas, show='headings', yscrollcommand=scrollbar.set)
            
            # Cabeçalhos
            tabela.heading('mes_ano', text='Mês/Ano')
            tabela.heading('total', text='Total (R$)')
            tabela.heading('registros', text='Qtd. Registros')
            tabela.heading('percentual', text='% do Total')
            
            # Larguras
            tabela.column('mes_ano', width=100, anchor=tk.CENTER)
            tabela.column('total', width=150, anchor=tk.E)
            tabela.column('registros', width=100, anchor=tk.CENTER)
            tabela.column('percentual', width=100, anchor=tk.CENTER)
            
            # Preencher tabela
            for row in resultados:
                ano_mes, mes_ano_formatado, total, registros = row
                percentual = (total / total_geral) * 100
                
                tabela.insert('', tk.END, values=(
                    mes_ano_formatado,
                    f"R$ {total:.2f}".replace('.', ','),
                    registros,
                    f"{percentual:.1f}%"
                ))
                
            tabela.pack(fill=tk.BOTH, expand=True)
            scrollbar.config(command=tabela.yview)
            
            # Criar gráfico de barras na segunda aba
            figura = plt.Figure(figsize=(7, 5), dpi=100)
            ax = figura.add_subplot(111)
            
            # Extrair dados para o gráfico
            meses = [row[1] for row in resultados]
            valores = [row[2] for row in resultados]
            
            # Criar gráfico de barras
            barras = ax.bar(meses, valores, color=self.cor_primaria)
            
            # Adicionar rótulos com valores
            for barra in barras:
                altura = barra.get_height()
                ax.text(barra.get_x() + barra.get_width()/2., altura + 0.05*max(valores),
                    f'R${altura:.2f}'.replace('.', ','),
                    ha='center', va='bottom', fontsize=9, rotation=0)
            
            # Estilo do gráfico
            ax.set_title(f'Despesas Mensais - {data_inicial_formatada} a {data_final_formatada}')
            ax.set_xlabel('Mês/Ano')
            ax.set_ylabel('Valor Total (R$)')
            ax.grid(True, linestyle='--', alpha=0.7, axis='y')
            
            # Rotacionar rótulos do eixo x para melhor visualização
            plt.setp(ax.get_xticklabels(), rotation=45, ha='right')
            
            # Ajustar layout
            figura.tight_layout()
            
            # Adicionar o gráfico ao frame
            canvas = FigureCanvasTkAgg(figura, tab_grafico)
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
            
            # Adicionar aba de gráfico de linha também
            tab_grafico_linha = ttk.Frame(notebook)
            notebook.add(tab_grafico_linha, text="Evolução")
            
            # Criar gráfico de linha na terceira aba
            figura_linha = plt.Figure(figsize=(7, 5), dpi=100)
            ax_linha = figura_linha.add_subplot(111)
            
            # Criar gráfico de linha
            ax_linha.plot(meses, valores, marker='o', linestyle='-', color=self.cor_primaria, linewidth=2)
            
            # Adicionar rótulos
            for i, valor in enumerate(valores):
                ax_linha.annotate(f'R${valor:.2f}'.replace('.', ','), 
                        (meses[i], valores[i]), 
                        textcoords="offset points",
                        xytext=(0,10), 
                        ha='center',
                        fontsize=8)
            
            # Estilo do gráfico
            ax_linha.set_title(f'Evolução de Despesas Mensais')
            ax_linha.set_xlabel('Mês/Ano')
            ax_linha.set_ylabel('Valor Total (R$)')
            ax_linha.grid(True, linestyle='--', alpha=0.7)
            
            # Rotacionar rótulos do eixo x
            plt.setp(ax_linha.get_xticklabels(), rotation=45, ha='right')
            
            # Ajustar layout
            figura_linha.tight_layout()
            
            # Adicionar o gráfico ao frame
            canvas_linha = FigureCanvasTkAgg(figura_linha, tab_grafico_linha)
            canvas_linha.get_tk_widget().pack(fill=tk.BOTH, expand=True)

            # Função para exportar para Excel
            def exportar_para_excel():
                try:
                    # Solicitar local para salvar o arquivo
                    arquivo = filedialog.asksaveasfilename(
                        defaultextension=".xlsx",
                        filetypes=[("Excel files", "*.xlsx")],
                        title="Salvar Relatório Mensal"
                    )
                    
                    if not arquivo:
                        return
                    
                    # Criar planilha
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Despesas Mensais"
                    
                    # Título
                    ws['A1'] = f"Relatório de Despesas Mensais - Período: {data_inicial_formatada} a {data_final_formatada}"
                    ws.merge_cells('A1:D1')
                    ws['A1'].font = Font(size=14, bold=True)
                    
                    # Total
                    ws['A2'] = f"Total de Despesas: R$ {total_geral:.2f}".replace('.', ',')
                    ws.merge_cells('A2:D2')
                    ws['A2'].font = Font(size=12)
                    
                    # Cabeçalhos
                    cabecalhos = ["Mês/Ano", "Total (R$)", "Qtd. Registros", "% do Total"]
                    for col, cabecalho in enumerate(cabecalhos, start=1):
                        cell = ws.cell(row=4, column=col, value=cabecalho)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                    
                    # Dados
                    for row_idx, (categoria, total, registros) in enumerate(resultados, start=5):
                        percentual = (total / total_geral) * 100
                        
                        ws.cell(row=row_idx, column=1, value=categoria)
                        ws.cell(row=row_idx, column=2, value=f"R$ {total:.2f}".replace('.', ','))
                        ws.cell(row=row_idx, column=3, value=registros)
                        ws.cell(row=row_idx, column=4, value=f"{percentual:.1f}%")
                    
                    # Ajustar larguras
                    for col in range(1, 5):
                        ws.column_dimensions[get_column_letter(col)].width = 20
                    
                    # Salvar arquivo
                    wb.save(arquivo)
                    messagebox.showinfo("Exportar", f"Relatório exportado com sucesso para {arquivo}")
                    
                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao exportar relatório: {e}")
            
            # Botão para exportar
            ttk.Button(frame_principal, text="Exportar para Excel", 
                    command=exportar_para_excel).pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exibir relatório: {e}")
        """Gera um relatório de despesas totais somadas por mês entre um intervalo de datas"""
        try:
            dialog = tk.Toplevel(self.root)
            dialog.title("Relatório de Despesas por Mês")
            dialog.geometry("400x200") # Reduzido um pouco
            dialog.resizable(False, False)
            
            frame = ttk.Frame(dialog, padding="10")
            frame.pack(fill=tk.BOTH, expand=True)
            
            data_atual = datetime.now()
            data_inicial_var = tk.StringVar(value=f"01/01/{data_atual.year}")
            data_final_var = tk.StringVar(value=data_atual.strftime("%d/%m/%Y"))
            
            ttk.Label(frame, text="Data Inicial:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
            data_inicial_entry = ttk.Entry(frame, textvariable=data_inicial_var, width=15)
            data_inicial_entry.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)
            
            ttk.Label(frame, text="Data Final:").grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
            data_final_entry = ttk.Entry(frame, textvariable=data_final_var, width=15)
            data_final_entry.grid(row=1, column=1, padx=5, pady=5, sticky=tk.W)
            
            # ### OTIMIZAÇÃO 3: Chamar o método reutilizável também aqui ###
            ttk.Button(frame, text="📅", width=3, 
                       command=lambda: self._abrir_calendario_selecao(dialog, data_inicial_var)).grid(row=0, column=2, padx=5)
            ttk.Button(frame, text="📅", width=3, 
                       command=lambda: self._abrir_calendario_selecao(dialog, data_final_var)).grid(row=1, column=2, padx=5)

            def gerar():
                try:
                    data_inicial = datetime.strptime(data_inicial_var.get(), "%d/%m/%Y").strftime("%Y-%m-%d")
                    data_final = datetime.strptime(data_final_var.get(), "%d/%m/%Y").strftime("%Y-%m-%d")
                    
                    if data_inicial > data_final:
                        messagebox.showerror("Erro", "Data inicial deve ser anterior à data final.", parent=dialog)
                        return
                        
                    self.mostrar_relatorio_mensal_periodo(data_inicial, data_final)
                    dialog.destroy()
                except ValueError:
                    messagebox.showerror("Erro", "Formato de data inválido. Use DD/MM/AAAA.", parent=dialog)
            
            ttk.Button(frame, text="Gerar Relatório", command=gerar).grid(row=3, column=0, columnspan=3, pady=20)
            
            dialog.transient(self.root)
            dialog.grab_set()
            dialog.wait_window()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar relatório: {e}")

    def mostrar_relatorio_mensal_periodo(self, data_inicial, data_final):
        """Exibe o relatório de despesas totais somadas por mês dentro do período selecionado"""
        try:
            # Formatar datas para exibição
            data_inicial_formatada = datetime.strptime(data_inicial, "%Y-%m-%d").strftime("%d/%m/%Y")
            data_final_formatada = datetime.strptime(data_final, "%Y-%m-%d").strftime("%d/%m/%Y")
            
            # Consultar despesas agrupadas por mês
            self.cursor.execute("""
                SELECT 
                    strftime('%Y-%m', data_pagamento) as ano_mes,
                    strftime('%m/%Y', data_pagamento) as mes_ano_formatado,
                    SUM(valor) as total_valor,
                    COUNT(*) as total_registros
                FROM v_despesas_compat
                WHERE data_pagamento BETWEEN ? AND ?
                AND user_id = ?
                GROUP BY ano_mes
                ORDER BY ano_mes
            """, (data_inicial, data_final, self.sessao.user_id))
            
            resultados = self.cursor.fetchall()
            
            if not resultados:
                messagebox.showinfo("Relatório", f"Nenhuma despesa encontrada no período de {data_inicial_formatada} a {data_final_formatada}.")
                return
                
            # Calcular total geral
            total_geral = sum(row[2] for row in resultados)
            
            # Criar nova janela para o relatório
            janela_relatorio = tk.Toplevel(self.root)
            janela_relatorio.title(f"Relatório Mensal - {data_inicial_formatada} a {data_final_formatada}")
            janela_relatorio.geometry("900x600")
            
            # Frame principal
            frame_principal = ttk.Frame(janela_relatorio, padding="10")
            frame_principal.pack(fill=tk.BOTH, expand=True)
            
            # Título
            titulo = ttk.Label(frame_principal, 
                            text=f"Despesas Mensais - {data_inicial_formatada} a {data_final_formatada}",
                            font=('Arial', 16, 'bold'))
            titulo.pack(pady=10)
            
            # Total geral
            total_label = ttk.Label(frame_principal, 
                                text=f"Total de Despesas: R$ {total_geral:.2f}".replace('.', ','),
                                font=('Arial', 12))
            total_label.pack(pady=5)
            
            # Frame com abas
            notebook = ttk.Notebook(frame_principal)
            notebook.pack(fill=tk.BOTH, expand=True, pady=10)
            
            # Aba de tabela
            tab_tabela = ttk.Frame(notebook)
            notebook.add(tab_tabela, text="Tabela")
            
            # Aba de gráfico
            tab_grafico = ttk.Frame(notebook)
            notebook.add(tab_grafico, text="Gráfico")
            
            # Criar tabela na primeira aba
            tree_frame = ttk.Frame(tab_tabela)
            tree_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
            
            # Scrollbar para a tabela
            scrollbar = ttk.Scrollbar(tree_frame)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # Tabela de resultados
            colunas = ('mes_ano', 'total', 'registros', 'percentual')
            tabela = ttk.Treeview(tree_frame, columns=colunas, show='headings', yscrollcommand=scrollbar.set)
            
            # Cabeçalhos
            tabela.heading('mes_ano', text='Mês/Ano')
            tabela.heading('total', text='Total (R$)')
            tabela.heading('registros', text='Qtd. Registros')
            tabela.heading('percentual', text='% do Total')
            
            # Larguras
            tabela.column('mes_ano', width=100, anchor=tk.CENTER)
            tabela.column('total', width=150, anchor=tk.E)
            tabela.column('registros', width=100, anchor=tk.CENTER)
            tabela.column('percentual', width=100, anchor=tk.CENTER)
            
            # Preencher tabela
            for row in resultados:
                ano_mes, mes_ano_formatado, total, registros = row
                percentual = (total / total_geral) * 100
                
                tabela.insert('', tk.END, values=(
                    mes_ano_formatado,
                    f"R$ {total:.2f}".replace('.', ','),
                    registros,
                    f"{percentual:.1f}%"
                ))
                
            tabela.pack(fill=tk.BOTH, expand=True)
            scrollbar.config(command=tabela.yview)
            
            # Criar gráfico de barras na segunda aba
            figura = plt.Figure(figsize=(7, 5), dpi=100)
            ax = figura.add_subplot(111)
            
            # Extrair dados para o gráfico
            meses = [row[1] for row in resultados]
            valores = [row[2] for row in resultados]
            
            # Criar gráfico de barras
            barras = ax.bar(meses, valores, color=self.cor_primaria)
            
            # Adicionar rótulos com valores
            for barra in barras:
                altura = barra.get_height()
                ax.text(barra.get_x() + barra.get_width()/2., altura + 0.05*max(valores),
                    f'R${altura:.2f}'.replace('.', ','),
                    ha='center', va='bottom', fontsize=9, rotation=0)
            
            # Estilo do gráfico
            ax.set_title(f'Despesas Mensais - {data_inicial_formatada} a {data_final_formatada}')
            ax.set_xlabel('Mês/Ano')
            ax.set_ylabel('Valor Total (R$)')
            ax.grid(True, linestyle='--', alpha=0.7, axis='y')
            
            # Rotacionar rótulos do eixo x para melhor visualização
            plt.setp(ax.get_xticklabels(), rotation=45, ha='right')
                WHERE data_pagamento BETWEEN ? AND ?
                AND user_id = ?
                ORDER BY data_pagamento
            """, (primeiro_dia, ultimo_dia, self.sessao.user_id))
            
            df_detalhes = pd.DataFrame(self.cursor.fetchall(), 
                                     columns=['ID', 'Descrição', 'Meio de Pagamento', 'Categoria', 
                                             'Valor (R$)', 'Parcelas', 'Data de Pagamento'])
            
            # Solicitar local para salvar o arquivo
            from tkinter import filedialog
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Salvar Como"
            )
            
            if not file_path:
                return  # Usuário cancelou
                
            # Criar arquivo Excel com múltiplas planilhas
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df_resumo.to_excel(writer, sheet_name='Resumo', index=False)
                df_detalhes.to_excel(writer, sheet_name='Detalhes', index=False)
            
            messagebox.showinfo("Exportação Concluída", 
                             f"Relatório exportado com sucesso para {file_path}")
            
        except Exception as e:
            messagebox.showerror("Erro na Exportação", f"Erro ao exportar relatório: {e}")
    
    def exportar_categoria_excel(self, categoria):
        """Exporta todas as despesas para um arquivo Excel"""
        try:
            # Consultar resumo por mês
# Fixed garbage line 2278
                SELECT 
                    strftime('%m/%Y', data_pagamento) as 'Mês/Ano',
                    SUM(valor) as 'Total Valor',
                    COUNT(*) as 'Quantidade'
                FROM v_despesas_compat
                WHERE conta_despesa = ?
                AND user_id = ?
                GROUP BY strftime('%Y-%m', data_pagamento)
                ORDER BY strftime('%Y-%m', data_pagamento) DESC
            """, (categoria, self.sessao.user_id))
            
            # Converter para DataFrame
            df_resumo = pd.DataFrame(self.cursor.fetchall(), 
                                   columns=['Mês/Ano', 'Total (R$)', 'Quantidade'])
            
            # Consultar detalhes das despesas
            self.cursor.execute("""
                SELECT 
                    id as ID,
                    descricao as Descrição,
                    meio_pagamento as 'Meio de Pagamento',
                    valor as 'Valor (R$)',
                    num_parcelas as Parcelas,
                    strftime('%d/%m/%Y', data_pagamento) as 'Data de Pagamento'
                FROM v_despesas_compat
                WHERE conta_despesa = ?
                AND user_id = ?
                ORDER BY data_pagamento DESC
            """, (categoria, self.sessao.user_id))
            
            df_detalhes = pd.DataFrame(self.cursor.fetchall(), 
                                     columns=['ID', 'Descrição', 'Meio de Pagamento', 
                                             'Valor (R$)', 'Parcelas', 'Data de Pagamento'])
            
            # Solicitar local para salvar o arquivo
            from tkinter import filedialog
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Salvar Como"
            )
            
            if not file_path:
                return  # Usuário cancelou
                
            # Criar arquivo Excel com múltiplas planilhas
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df_resumo.to_excel(writer, sheet_name=f'Resumo - {categoria}', index=False)
                df_detalhes.to_excel(writer, sheet_name='Detalhes', index=False)
            
            messagebox.showinfo("Exportação Concluída", 
                             f"Relatório exportado com sucesso para {file_path}")
            ax.set_title(f'Evolução de Despesas - {categoria}')
            ax.set_xlabel('Mês/Ano')
            ax.set_ylabel('Valor Total (R$)')
            ax.grid(True, linestyle='--', alpha=0.7)
            
            # Rotacionar rótulos do eixo x
            plt.setp(ax.get_xticklabels(), rotation=45, ha='right')
            
            # Ajustar layout
            figura.tight_layout()
            
            # Adicionar o gráfico ao frame
            canvas = FigureCanvasTkAgg(figura, tab_grafico)
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
            
            # Botão para exportar
            ttk.Button(frame_principal, text="Exportar para Excel", 
    def exportar_excel(self):
        """Exporta todas as despesas para um arquivo Excel"""
        try:
            # Verificar se pandas está instalado
            try:
                import pandas as pd
            except ImportError:
                resposta = messagebox.askyesno(
                    "Biblioteca Necessária", 
                    "É necessário instalar a biblioteca pandas para exportar para Excel. Deseja instalar agora?"
                )
                if resposta:
                    import subprocess
                    subprocess.check_call(['pip', 'install', 'pandas', 'openpyxl'])
                    import pandas as pd
                else:
                    return
            
            # Consultar todas as despesas
            self.cursor.execute("""
                SELECT 
                    id, descricao, meio_pagamento, conta_despesa, valor, 
                    num_parcelas, strftime('%d/%m/%Y', data_pagamento) as data_pagamento,
                    strftime('%d/%m/%Y', data_registro) as data_registro
                FROM v_despesas_compat
                WHERE user_id = ?
                ORDER BY data_pagamento DESC
            """, (self.sessao.user_id,))
            
            # Converter resultado para DataFrame
            colunas = ['ID', 'Descrição', 'Meio de Pagamento', 'Categoria', 
                     'Valor (R$)', 'Parcelas', 'Data de Pagamento', 'Data de Registro']
            
            df = pd.DataFrame(self.cursor.fetchall(), columns=colunas)
            
            # Solicitar local para salvar o arquivo
            from tkinter import filedialog
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Salvar Como"
            )
            
            if not file_path:
                return  # Usuário cancelou
                
            # Exportar para Excel
            df.to_excel(file_path, index=False, sheet_name='Despesas')
            
            messagebox.showinfo("Exportação Concluída", 
                             f"Dados exportados com sucesso para {file_path}")
            
        except Exception as e:
            messagebox.showerror("Erro na Exportação", f"Erro ao exportar dados: {e}")
    
    def exportar_relatorio_excel(self, mes, ano):
        """Exporta o relatório mensal para Excel"""
        try:
            # Calcular início e fim do mês
            primeiro_dia = f"{ano}-{mes:02d}-01"
            ultimo_dia = f"{ano}-{mes:02d}-{calendar.monthrange(ano, mes)[1]}"
            
            # Consultar dados do mês
            self.cursor.execute("""
                SELECT 
                    conta_despesa as Categoria, 
                    SUM(valor) as 'Total Valor',
                    COUNT(*) as 'Quantidade'
                FROM v_despesas_compat
                WHERE data_pagamento BETWEEN ? AND ?
                AND user_id = ?
                GROUP BY conta_despesa
                ORDER BY 'Total Valor' DESC
            """, (primeiro_dia, ultimo_dia, self.sessao.user_id))
            
            # Converter para DataFrame
            df_resumo = pd.DataFrame(self.cursor.fetchall(), 
                                   columns=['Categoria', 'Total (R$)', 'Quantidade'])
            
            # Consultar detalhes das despesas
            self.cursor.execute("""
                SELECT 
                    id as ID,
                    descricao as Descrição,
                    meio_pagamento as 'Meio de Pagamento',
                    conta_despesa as Categoria,
                    valor as 'Valor (R$)',
                    num_parcelas as Parcelas,
                    strftime('%d/%m/%Y', data_pagamento) as 'Data de Pagamento'
                FROM v_despesas_compat
                WHERE data_pagamento BETWEEN ? AND ?
                AND user_id = ?
                ORDER BY data_pagamento
            """, (primeiro_dia, ultimo_dia, self.sessao.user_id))
            
            df_detalhes = pd.DataFrame(self.cursor.fetchall(), 
                                     columns=['ID', 'Descrição', 'Meio de Pagamento', 'Categoria', 
                                             'Valor (R$)', 'Parcelas', 'Data de Pagamento'])
            
            # Solicitar local para salvar o arquivo
            from tkinter import filedialog
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Salvar Como"
            )
            
            if not file_path:
                return  # Usuário cancelou
                
            # Criar arquivo Excel com múltiplas planilhas
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df_resumo.to_excel(writer, sheet_name='Resumo', index=False)
                df_detalhes.to_excel(writer, sheet_name='Detalhes', index=False)
            
            messagebox.showinfo("Exportação Concluída", 
                             f"Relatório exportado com sucesso para {file_path}")
            
        except Exception as e:
            messagebox.showerror("Erro na Exportação", f"Erro ao exportar relatório: {e}")
    
    def exportar_categoria_excel(self, categoria):
        """Exporta o relatório por categoria para Excel"""
        try:
            # Consultar resumo por mês
            self.cursor.execute("""
                SELECT 
                    strftime('%m/%Y', data_pagamento) as 'Mês/Ano',
                    SUM(valor) as 'Total Valor',
                    COUNT(*) as 'Quantidade'
                FROM v_despesas_compat
                WHERE conta_despesa = ?
                AND user_id = ?
                GROUP BY strftime('%Y-%m', data_pagamento)
                ORDER BY strftime('%Y-%m', data_pagamento) DESC
            """, (categoria, self.sessao.user_id))
            
            # Converter para DataFrame
            df_resumo = pd.DataFrame(self.cursor.fetchall(), 
                                   columns=['Mês/Ano', 'Total (R$)', 'Quantidade'])
            
            # Consultar detalhes das despesas
            self.cursor.execute("""
                SELECT 
                    id as ID,
                    descricao as Descrição,
                    meio_pagamento as 'Meio de Pagamento',
                    valor as 'Valor (R$)',
                    num_parcelas as Parcelas,
                    strftime('%d/%m/%Y', data_pagamento) as 'Data de Pagamento'
                FROM v_despesas_compat
                WHERE conta_despesa = ?
                AND user_id = ?
                ORDER BY data_pagamento DESC
            """, (categoria, self.sessao.user_id))
            
            df_detalhes = pd.DataFrame(self.cursor.fetchall(), 
                                     columns=['ID', 'Descrição', 'Meio de Pagamento', 
                                             'Valor (R$)', 'Parcelas', 'Data de Pagamento'])
            
            # Solicitar local para salvar o arquivo
            from tkinter import filedialog
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Salvar Como"
            )
            
            if not file_path:
                return  # Usuário cancelou
                
            # Criar arquivo Excel com múltiplas planilhas
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df_resumo.to_excel(writer, sheet_name=f'Resumo - {categoria}', index=False)
                df_detalhes.to_excel(writer, sheet_name='Detalhes', index=False)
            
            messagebox.showinfo("Exportação Concluída", 
                             f"Relatório exportado com sucesso para {file_path}")

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exibir relatrio: {e}")



        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exibir relatório: {e}")

# Função principal
def main():
    root = tk.Tk()
    app = SistemaFinanceiro(root)
    root.protocol("WM_DELETE_WINDOW", app.on_closing)
    root.mainloop()

if __name__ == "__main__":
    main()
